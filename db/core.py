"""
实验室设备管理系统 - 数据库层
提供 SQLite 数据库初始化、CRUD 操作和统计查询接口
"""

import sqlite3
import os
import sys
import json
import shutil
import inspect
import unicodedata
from functools import lru_cache, wraps
from datetime import date, timedelta, datetime

# 项目根目录：本文件位于 <root>/db/core.py，上提两级还原为 <root>
# （保持与原 database.py 处于项目根时完全一致的路径解析）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 路径
DB_DIR = os.path.join(PROJECT_ROOT, 'data')
DB_PATH = os.path.join(DB_DIR, 'lab_manager.db')
DB_BACKUP_DIR = os.path.join(DB_DIR, 'backups')
DB_REPAIR_STATUS_FILE = os.path.join(DB_DIR, 'db_repair_status.json')
SAMPLE_SYNC_REQUEST_FILE = os.path.join(DB_DIR, 'windows_samples_sync_request.json')
SAMPLE_SYNC_APPLIED_FILE = os.path.join(DB_DIR, 'windows_samples_sync_applied.json')
# 原始名单文件路径
NAMES_FILE = os.path.join(PROJECT_ROOT,
                          'SainStore实验室文件', '原始名单.xlsx')

_DB_READY_CACHE = {"checked": False, "path": None}


def _filename_readability_score(name):
    if not name:
        return -999

    score = 0
    for ch in str(name):
        code = ord(ch)
        if ch == "\ufffd":
            score -= 10
        elif 0x2500 <= code <= 0x257F:
            score -= 6
        elif unicodedata.category(ch).startswith("C") and ch not in ("\t", "\n", "\r"):
            score -= 4
        elif "\u4e00" <= ch <= "\u9fff":
            score += 3
        elif ch.isalnum():
            score += 1
        elif ch in " ._()-[]{}&+,，（）【】|":
            score += 1
    return score


def _repair_filename_mojibake(name):
    if not name:
        return name

    name = str(name).strip()
    segments = [seg.strip() for seg in name.split("|")]
    repaired_segments = []

    for segment in segments:
        if not segment:
            continue

        candidates = {segment}
        for source_encoding in ("cp437", "latin1"):
            try:
                raw_bytes = segment.encode(source_encoding)
            except Exception:
                continue

            for target_encoding in ("utf-8", "gbk", "gb18030", "big5"):
                try:
                    repaired = raw_bytes.decode(target_encoding).strip()
                    if repaired:
                        candidates.add(repaired)
                except Exception:
                    continue

        repaired_segments.append(max(candidates, key=_filename_readability_score))

    return " | ".join(repaired_segments) if repaired_segments else name


def _write_db_repair_status(payload):
    try:
        os.makedirs(DB_DIR, exist_ok=True)
        with open(DB_REPAIR_STATUS_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _validate_sqlite_file(path):
    if not os.path.exists(path):
        return False, "missing"
    if os.path.getsize(path) <= 0:
        return False, "empty file"

    conn = None
    try:
        conn = sqlite3.connect(path)
        result = conn.execute("PRAGMA quick_check").fetchone()
        status = (result[0] if result else "").strip()
        if status.lower() == "ok":
            return True, "ok"
        return False, status or "quick_check failed"
    except sqlite3.Error as e:
        return False, str(e)
    finally:
        if conn is not None:
            conn.close()


def _latest_valid_backup():
    if not os.path.isdir(DB_BACKUP_DIR):
        return None

    candidates = []
    for name in os.listdir(DB_BACKUP_DIR):
        if name.endswith(".db"):
            path = os.path.join(DB_BACKUP_DIR, name)
            candidates.append((os.path.getmtime(path), path))

    for _, path in sorted(candidates, reverse=True):
        ok, _ = _validate_sqlite_file(path)
        if ok:
            return path
    return None


def _ensure_database_ready():
    cache_key = os.path.abspath(DB_PATH)
    if _DB_READY_CACHE["checked"] and _DB_READY_CACHE["path"] == cache_key:
        return

    os.makedirs(DB_DIR, exist_ok=True)
    if not _DB_READY_CACHE.get("wal_cleaned"):
        for _suf in ("-wal", "-shm"):
            try:
                os.remove(DB_PATH + _suf)
            except Exception:
                pass
        _DB_READY_CACHE["wal_cleaned"] = True
    _DB_READY_CACHE["checked"] = True
    _DB_READY_CACHE["path"] = cache_key
    try:
        init_db()
        _c = get_connection()
        try:
            _c.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        except Exception:
            pass
        _c.close()
    except Exception as _e:
        import traceback as _tb
        try:
            with open(os.path.join(DB_DIR, "initdb_error.log"), "a", encoding="utf-8") as _f:
                _f.write(f"[{datetime.now()}] init_db EXC: {repr(_e)}\n{_tb.format_exc()}\n")
        except Exception:
            pass
    ok, reason = _validate_sqlite_file(DB_PATH)
    if ok or reason == "missing":
        _DB_READY_CACHE["checked"] = True
        _DB_READY_CACHE["path"] = cache_key
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    corrupt_copy = None
    if os.path.exists(DB_PATH):
        corrupt_copy = os.path.join(DB_DIR, f"lab_manager.corrupt.{timestamp}.db")
        try:
            shutil.move(DB_PATH, corrupt_copy)
        except Exception:
            corrupt_copy = None
            try:
                os.remove(DB_PATH)
            except Exception:
                pass

    restored_from = _latest_valid_backup()
    if restored_from:
        shutil.copy2(restored_from, DB_PATH)
        _write_db_repair_status({
            "handled_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": "restored_from_backup",
            "db_path": DB_PATH,
            "reason": reason,
            "backup_path": restored_from,
            "corrupt_copy": corrupt_copy or "",
        })
    else:
        _write_db_repair_status({
            "handled_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": "reinitialized_empty_database",
            "db_path": DB_PATH,
            "reason": reason,
            "backup_path": "",
            "corrupt_copy": corrupt_copy or "",
        })

    _DB_READY_CACHE["checked"] = True
    _DB_READY_CACHE["path"] = cache_key


# ==================== 全局远程数据源（Mac 直读 Win 生产库） ====================

# 模块级缓存：避免每次 get_connection 都重新解析远程路径
_remote_db_cache = {
    "path": None,          # 远程库绝对路径（或 None）
    "resolved_at": None,   # 解析时间戳
    "usable": False,       # 是否可用（连接测试通过）
    "mode": None,          # "rw" / "ro" / None
    "error": None,         # 最近一次失败原因
}

_REMOTE_CACHE_TTL = 60  # 路径缓存有效期（秒），超过后重新探测


def _get_global_remote_db_path(force_refresh: bool = False):
    """返回全局远程生产库路径（带 TTL 缓存）。

    仅 Mac(darwin) 环境尝试解析；Win(win32) 本地库就是生产库，无需远程。
    返回 (path_str, is_usable, mode_str) 或 (None, False, None)。
    """
    import time as _time
    cache = _remote_db_cache

    # 命令行强制本地
    if os.environ.get("LAB_AUDIT_LOCAL") == "1":
        return None, False, None

    # Mac 上显式启用生产模式（qms_launch.sh 设 FORCE_PRODUCTION=1）时，
    # 默认不再尝试远程 Windows 生产库。本地 Mac 经 SMB 反复打开远程 SQLite
    # 会导致每次切页 2-3 秒白屏；若确实需要远程可设 LAB_AUDIT_REMOTE=1。
    if (sys.platform == "darwin" and
            os.environ.get("FORCE_PRODUCTION") == "1" and
            os.environ.get("LAB_AUDIT_REMOTE") != "1"):
        return None, False, None

    # 缓存有效且未强制刷新 → 直接返回
    if (not force_refresh and cache["resolved_at"] and
            (_time.time() - cache["resolved_at"]) < _REMOTE_CACHE_TTL):
        if cache["usable"]:
            return cache["path"], True, cache["mode"]
        # 上次不可用，但还在缓存期内 → 不再重复试（避免每查询都超时）
        return None, False, cache.get("error", "远程库不可达")

    # 需要重新解析
    resolved = None
    try:
        resolved = _resolve_remote_audit_db()
    except Exception:
        pass

    if not resolved:
        cache.update(path=None, usable=False, mode=None,
                     error="未找到远程库路径", resolved_at=_time.time())
        return None, False, "未找到远程库路径"

    # 测试连接可用性
    mode = None
    try:
        test_conn = sqlite3.connect(f"file:{resolved}?mode=rw", uri=True, timeout=5)
        test_conn.execute("SELECT 1").fetchone()
        test_conn.close()
        mode = "rw"
        usable = True
    except Exception:
        # 读写失败 → 尝试只读
        try:
            test_ro = sqlite3.connect(f"file:{resolved}?mode=ro", uri=True, timeout=5)
            test_ro.execute("SELECT 1").fetchone()
            test_ro.close()
            mode = "ro"
            usable = True
        except Exception as e:
            usable = False

    cache.update(
        path=resolved,
        usable=usable,
        mode=mode,
        error=None if usable else f"连接失败({str(e)[:60]})",
        resolved_at=_time.time(),
    )
    return resolved, usable, mode


def get_connection(remote_fallback: bool = True):
    """获取数据库连接（性能优化：WAL模式 + 内存缓存 + mmap）。

    全局远程数据源支持：
    - Mac 开发环境：自动尝试直连 Win 生产库（优先 rw，锁库时降级 ro）
    - Win 生产环境：直接使用本地库（本机即生产库）
    - 远程不可达：自动回退本地库

    Args:
        remote_fallback: 为 True 时远程不可达自动降级本地；
                         为 False 时仅返回远程连接或抛异常（用于必须读生产数据的场景）
    """
    _ensure_database_ready()
    os.makedirs(DB_DIR, exist_ok=True)

    # ---- Mac 环境：尝试远程生产库 ----
    if sys.platform == "darwin":
        rpath, rusable, rmode = _get_global_remote_db_path()
        if rusable and rpath:
            try:
                conn = sqlite3.connect(f"file:{rpath}?mode={rmode}", uri=True, timeout=10)
                conn.row_factory = sqlite3.Row
                conn.execute("PRAGMA foreign_keys = ON")
                conn.execute("PRAGMA journal_mode=WAL")          # 远程 WAL 不影响本地
                conn.execute("PRAGMA synchronous=NORMAL")
                conn.execute("PRAGMA cache_size=-8000")
                conn.execute("PRAGMA mmap_size=268435456")
                conn.execute("PRAGMA temp_store=MEMORY")
                return conn
            except Exception:
                pass  # 连接瞬间失效 → 降级本地

    # ---- 本地库（默认/回退） ----
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        conn.execute("PRAGMA journal_mode=DELETE")   # 直接写主库，避免 WAL 导致 pull 镜像不同步
    except Exception:
        pass
    conn.execute("PRAGMA synchronous=NORMAL")       # 降低 fsync 开销
    conn.execute("PRAGMA cache_size=-8000")         # 8MB 页缓存
    conn.execute("PRAGMA mmap_size=268435456")      # 256MB mmap(加速大文件读取)
    conn.execute("PRAGMA temp_store=MEMORY")        # 临时表/排序存内存
    return conn


def get_global_datasource_status() -> dict:
    """返回当前全局数据源状态（供 sidebar / 各页面指示器使用）。

    返回 dict:
        source:     "production" | "remote" | "local"
        label:      显示文案
        detail:     详情（路径/错误原因）
        mode:       "rw" | "ro" | None
        icon:       "🟢" | "🟡" | "💻" | "❌"
        is_remote:  bool
    """
    if sys.platform == "win32" or os.environ.get("FORCE_PRODUCTION") == "1":
        return {
            "source": "production",
            "label": "本机生产库（实时）",
            "detail": f"{DB_PATH}",
            "mode": "rw",
            "icon": "🟢",
            "is_remote": False,
        }

    # Mac 环境：检查远程
    rpath, rusable, rmode = _get_global_remote_db_path()
    if rusable and rpath:
        mode_label = {"rw": "读写", "ro": "只读"}.get(rmode, rmode or "")
        return {
            "source": "remote",
            "label": f"Win 生产库（{mode_label}）",
            "detail": rpath,
            "mode": rmode,
            "icon": "🟢" if rmode == "rw" else "🟡",
            "is_remote": True,
        }

    # 回退本地
    err = _remote_db_cache.get("error", "") if _remote_db_cache.get("resolved_at") else "未检测"
    return {
        "source": "local",
        "label": "本地库（已自动回退）",
        "detail": f"{DB_PATH} | 原因: {err}",
        "mode": None,
        "icon": "💻",
        "is_remote": False,
    }


def invalidate_remote_cache():
    """强制清除远程库缓存，下次 get_connection 重新探测。
    用于手动切换数据源或网络恢复后触发刷新。
    """
    _remote_db_cache.update(
        path=None, usable=False, mode=None,
        error=None, resolved_at=None
    )


# ==================== 操作审计日志 ====================

def _detect_deployment():
    """判断当前运行环境：Win 生产服务器 or Mac 本地开发。"""
    if os.environ.get("FORCE_PRODUCTION", "").strip() == "1":
        return "Win-生产"
    return "Mac-本地"


def _current_actor():
    """从 Streamlit 会话安全读取操作人信息；非 Streamlit 环境返回系统。"""
    actor = {
        "operator": "系统",
        "network": "未知",
        "deployment": _detect_deployment(),
        "ip": "",
    }
    try:
        import streamlit as st
        ss = st.session_state
        email = (ss.get("user_email") or "").strip()
        name = (ss.get("user_name") or "").strip()
        actor["operator"] = name or email or "匿名"
        actor["network"] = ss.get("network_zone", "未知")
        actor["ip"] = ss.get("client_ip", "")
    except Exception:
        pass

    # 本地开发环境兜底：尝试检测本机 IP
    if not actor["ip"]:
        try:
            import socket
            hostname = socket.gethostname()
            ip = socket.gethostbyname(hostname)
            if ip and ip != "127.0.0.1":
                actor["ip"] = f"{ip} ({hostname})"
            else:
                actor["ip"] = "127.0.0.1 (localhost)"
        except Exception:
            actor["ip"] = "127.0.0.1"
    return actor


def log_operation(action, target_table, record_id=None, detail=""):
    """写入一条操作审计记录（同事内外网操作可追溯）。"""
    try:
        actor = _current_actor()
        conn = get_connection()
        conn.execute(
            """INSERT INTO operation_log
               (operator, action, target_table, record_id, detail, network, deployment, ip_address)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                actor["operator"], action, target_table,
                int(record_id) if record_id is not None else -1,
                str(detail)[:500], actor["network"], actor["deployment"], actor["ip"],
            ),
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


def audit(action, table, capture_new_id=False):
    """装饰器：自动记录数据写入操作。update/delete 提取 id 参数；insert 可选取 MAX(id)。"""
    # 参数名黑名单：这些是内部/ID 参数，不计入详情
    _SKIP_PARAMS = frozenset({
        "record_id", "report_id", "change_id", "outbound_id",
        "self", "cls", "conn",
    })

    def deco(func):
        try:
            sig = inspect.signature(func)
        except Exception:
            return func

        @wraps(func)
        def wrapper(*args, **kwargs):
            result = func(*args, **kwargs)
            try:
                ba = sig.bind(*args, **kwargs)
                ba.apply_defaults()
                rid = None
                # ── 提取记录 ID ──
                for pname, val in ba.arguments.items():
                    if pname in ("record_id", "report_id", "change_id", "outbound_id") or pname.endswith("_id"):
                        rid = val
                        break
                if rid is None and capture_new_id:
                    try:
                        c = get_connection()
                        rid = c.execute(f"SELECT MAX(id) FROM {table}").fetchone()[0]
                        c.close()
                    except Exception:
                        rid = None
                # ── 自动生成 detail（从函数参数中提取业务字段）──
                detail_parts = []
                for pname, val in ba.arguments.items():
                    if pname in _SKIP_PARAMS or pname.endswith("_id"):
                        continue
                    if val is None or val == "" or val == -1:
                        continue
                    # 截断超长值
                    s_val = str(val)[:80]
                    detail_parts.append(f"{pname}={s_val}")
                auto_detail = ", ".join(detail_parts) if detail_parts else ""
                log_operation(action, table, record_id=rid, detail=auto_detail)
            except Exception:
                pass
            return result
        return wrapper
    return deco


# ==================== 误删找回（回收站）基础设施 ====================

# 表 → 中文名，用于回收站可读展示
_RECOVERABLE_TABLE_CN = {
    "categories": "设备分类",
    "equipment": "设备台账",
    "users": "用户管理",
    "maintenance_records": "维护记录",
    "samples": "样品管理",
    "change_records": "变更管理",
    "inspection_reports": "检验报告",
    "changelog": "版本记录",
    "activity_log": "活动日志",
    "borrow_records": "借用归还",
}

# 每张表生成一句话摘要时的候选字段（按优先级）
_SUMMARY_FIELDS = {
    "change_records": ["bu", "brand", "sku", "change_reason"],
    "equipment": ["name", "model", "serial_number"],
    "users": ["name", "email", "department"],
    "samples": ["name", "brand", "sku", "model"],
    "maintenance_records": ["equipment_id", "maintenance_type", "maintenance_date"],
    "inspection_reports": ["filename", "sku", "report_no"],
    "borrow_records": ["equipment_id", "borrower", "borrow_date"],
    "categories": ["name"],
    "changelog": ["version", "title"],
    "activity_log": ["user", "action", "detail"],
}


def _summarize_row(table, row):
    """从整行 dict 生成一句话摘要，便于回收站辨认是哪条记录。"""
    try:
        parts = []
        for f in _SUMMARY_FIELDS.get(table, []):
            v = row.get(f)
            if v not in (None, "", -1):
                parts.append(str(v))
        if not parts:
            for k, v in row.items():
                if k == "id":
                    continue
                if v not in (None, "", -1):
                    parts.append(f"{k}={v}")
                if len(parts) >= 3:
                    break
        s = " / ".join(parts)
        return s[:200] if s else "(无摘要)"
    except Exception:
        return "(无摘要)"


def _save_deleted_record(table, record_id, row_dict):
    """把被删除的整行写入回收站 deleted_records。全程 try/except，绝不阻断业务。"""
    try:
        actor = _current_actor()
        conn = get_connection()
        conn.execute(
            """INSERT INTO deleted_records
               (source_table, record_id, record_json, summary,
                deleted_by, deleted_network, deleted_deployment)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                table,
                int(record_id) if record_id is not None else -1,
                json.dumps(row_dict, ensure_ascii=False, default=str),
                _summarize_row(table, row_dict),
                actor["operator"], actor["network"], actor["deployment"],
            ),
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


def recoverable(table, id_column="id"):
    """装饰器：删除前抓取整行快照，删除成功后写入回收站，支持后续一键还原。

    - 与 @audit 相互独立、可叠加（@audit 在外、@recoverable 紧贴 def）。
    - 自动识别第一个 *_id / *_ids 参数为记录 ID；支持批量（list/tuple/set）。
    - 快照在删除「之前」采集，因此即使函数内部连带删除文件，也能留存元数据。
    - 仅当被装饰函数返回成功（元组首元素为真，或非元组返回真值）时才落库。
    """
    def deco(func):
        try:
            sig = inspect.signature(func)
        except Exception:
            return func

        @wraps(func)
        def wrapper(*args, **kwargs):
            snapshot = []  # [(record_id, row_dict), ...]
            try:
                ba = sig.bind(*args, **kwargs)
                ba.apply_defaults()
                idval = None
                for pname, val in ba.arguments.items():
                    if pname in ("self", "cls", "conn"):
                        continue
                    if pname.endswith("_id") or pname.endswith("_ids"):
                        idval = val
                        break
                if idval is None:
                    for pname, val in ba.arguments.items():
                        if pname not in ("self", "cls", "conn"):
                            idval = val
                            break
                ids = list(idval) if isinstance(idval, (list, tuple, set)) else (
                    [idval] if idval is not None else [])
                if ids:
                    c = get_connection()
                    for rid in ids:
                        try:
                            r = c.execute(
                                f"SELECT * FROM {table} WHERE {id_column}=?", (rid,)
                            ).fetchone()
                            if r:
                                snapshot.append((rid, dict(r)))
                        except Exception:
                            pass
                    c.close()
            except Exception:
                pass

            result = func(*args, **kwargs)

            try:
                ok = True
                if isinstance(result, tuple) and len(result) >= 1:
                    ok = bool(result[0])
                elif result is not None:
                    ok = bool(result)
                if ok and snapshot:
                    for rid, row in snapshot:
                        _save_deleted_record(table, rid, row)
            except Exception:
                pass
            return result
        return wrapper
    return deco


def get_deleted_records(source_table="", include_restored=False, keyword="",
                        date_from="", date_to="", limit=500):
    """查询回收站记录（默认仅未还原）。表不存在时安全返回 []。"""
    try:
        conn = get_connection()
        conds, params = [], []
        if not include_restored:
            conds.append("restored = 0")
        if source_table:
            conds.append("source_table = ?"); params.append(source_table)
        if keyword:
            conds.append("(summary LIKE ? OR record_json LIKE ?)")
            params += [f"%{keyword}%", f"%{keyword}%"]
        if date_from:
            conds.append("date(deleted_at) >= date(?)"); params.append(date_from)
        if date_to:
            conds.append("date(deleted_at) <= date(?)"); params.append(date_to)
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        rows = conn.execute(
            f"SELECT * FROM deleted_records {where} ORDER BY id DESC LIMIT ?",
            params + [int(limit)],
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception:
        return []


def restore_deleted_record(deleted_id):
    """把回收站中的一条记录还原回原表。

    - 优先按原始 ID 还原（保住附件引用/外键关系）；原 ID 已被占用时以新 ID 插入。
    - 还原成功后标记 restored=1，避免重复还原，并记一条审计。
    返回 (success: bool, msg: str)。
    """
    try:
        conn = get_connection()
        row = conn.execute(
            "SELECT * FROM deleted_records WHERE id=?", (deleted_id,)
        ).fetchone()
        if not row:
            conn.close()
            return False, "回收站记录不存在"
        rec = dict(row)
        if rec.get("restored"):
            conn.close()
            return False, "该记录已还原，无需重复操作"
        table = rec["source_table"]
        try:
            data = json.loads(rec["record_json"] or "{}")
        except Exception:
            conn.close()
            return False, "记录数据损坏，无法解析"
        if not data:
            conn.close()
            return False, "记录数据为空"

        cols_info = conn.execute(f"PRAGMA table_info({table})").fetchall()
        if not cols_info:
            conn.close()
            return False, f"目标表 {table} 不存在"
        valid_cols = {c[1] for c in cols_info}
        payload = {k: v for k, v in data.items() if k in valid_cols}
        if not payload:
            conn.close()
            return False, "无可还原的字段"

        orig_id = payload.get("id")
        exists = False
        if orig_id is not None:
            exists = conn.execute(
                f"SELECT 1 FROM {table} WHERE id=?", (orig_id,)
            ).fetchone() is not None
        if exists:
            payload.pop("id", None)

        cols = list(payload.keys())
        ph = ",".join("?" for _ in cols)
        conn.execute(
            f"INSERT INTO {table} ({','.join(cols)}) VALUES ({ph})",
            [payload[c] for c in cols],
        )
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        actor = _current_actor()
        conn.execute(
            "UPDATE deleted_records SET restored=1, restored_by=?, "
            "restored_at=datetime('now','localtime') WHERE id=?",
            (actor["operator"], deleted_id),
        )
        conn.commit()
        conn.close()
        log_operation("还原记录", table, record_id=new_id,
                      detail=f"从回收站还原 deleted_id={deleted_id} 原ID={orig_id}")
        tip = f"（原 ID {orig_id} 已被占用，已用新 ID {new_id} 还原）" if exists else ""
        return True, f"已还原到「{_RECOVERABLE_TABLE_CN.get(table, table)}」{tip}"
    except Exception as e:
        return False, f"还原失败：{str(e)[:120]}"


def purge_deleted_records(before_date="", only_restored=False):
    """清理回收站：删除指定日期之前的记录（或仅清理已还原的）。返回清理条数。"""
    try:
        conn = get_connection()
        conds, params = [], []
        if before_date:
            conds.append("date(deleted_at) < date(?)"); params.append(before_date)
        if only_restored:
            conds.append("restored = 1")
        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        cnt = conn.execute(
            f"SELECT COUNT(*) FROM deleted_records {where}", params
        ).fetchone()[0]
        conn.execute(f"DELETE FROM deleted_records {where}", params)
        conn.commit()
        conn.close()
        return cnt
    except Exception:
        return 0


def _apply_pending_samples_sync(db_conn):
    """在 Win 端启动时自动消费待同步的样品数据包。"""
    if not os.path.exists(SAMPLE_SYNC_REQUEST_FILE):
        return

    handled_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(SAMPLE_SYNC_REQUEST_FILE, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception as exc:
        with open(SAMPLE_SYNC_APPLIED_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "handled_at": handled_at,
                "success": False,
                "message": f"读取样品同步包失败: {exc}",
            }, f, ensure_ascii=False, indent=2)
        return

    success, message, stats = merge_samples_sync_package(payload, db_conn=db_conn)
    applied_payload = {
        "handled_at": handled_at,
        "success": success,
        "message": message,
        "stats": stats or {},
        "sample_count": payload.get("sample_count", 0),
        "generated_at": payload.get("generated_at", ""),
        "type": payload.get("type", ""),
        "version": payload.get("version", ""),
    }
    with open(SAMPLE_SYNC_APPLIED_FILE, "w", encoding="utf-8") as f:
        json.dump(applied_payload, f, ensure_ascii=False, indent=2)

    if success:
        try:
            os.remove(SAMPLE_SYNC_REQUEST_FILE)
        except FileNotFoundError:
            pass


@lru_cache(maxsize=1)
def init_db():
    """初始化数据库表结构和种子数据（进程级缓存，仅首次调用时执行）"""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.executescript('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            description TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS equipment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            model TEXT DEFAULT '',
            serial_number TEXT UNIQUE,
            category_id INTEGER,
            location TEXT DEFAULT '',
            status TEXT DEFAULT '可用' CHECK(status IN ('可用','借出','维修中','报废')),
            purchase_date TEXT DEFAULT '',
            price REAL DEFAULT 0,
            supplier TEXT DEFAULT '',
            warranty_expiry TEXT DEFAULT '',
            description TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            department TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            email TEXT DEFAULT '',
            role TEXT DEFAULT '普通用户' CHECK(role IN ('管理员','普通用户')),
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS borrow_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            borrow_date TEXT NOT NULL,
            expected_return_date TEXT DEFAULT '',
            actual_return_date TEXT DEFAULT '',
            test_standard TEXT DEFAULT '',
            test_start_time TEXT DEFAULT '',
            test_end_time TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            status TEXT DEFAULT '借出中' CHECK(status IN ('借出中','已归还','逾期','已出库')),
            purpose TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (equipment_id) REFERENCES equipment(id),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS maintenance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_id INTEGER NOT NULL,
            maintenance_date TEXT NOT NULL,
            maintenance_type TEXT DEFAULT '定期保养' CHECK(maintenance_type IN ('定期保养','故障维修','校准','其他')),
            description TEXT DEFAULT '',
            cost REAL DEFAULT 0,
            technician TEXT DEFAULT '',
            next_maintenance_date TEXT DEFAULT '',
            status TEXT DEFAULT '已完成' CHECK(status IN ('已完成','进行中','计划中')),
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (equipment_id) REFERENCES equipment(id)
        );

        CREATE TABLE IF NOT EXISTS samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bg TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            sample_name TEXT DEFAULT '',
            sign_date TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            location TEXT DEFAULT '',
            expiry_date TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS factory_registrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            register_date TEXT DEFAULT '',
            factory_name TEXT DEFAULT '',
            onsite_staff TEXT DEFAULT '',
            trip_type TEXT DEFAULT '',
            trip_days INTEGER DEFAULT 0,
            po_no TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            product_project TEXT DEFAULT '',
            is_empty_run TEXT DEFAULT '否',
            is_recheck TEXT DEFAULT '否',
            is_delay TEXT DEFAULT '否',
            delay_days INTEGER DEFAULT 0,
            return_reason TEXT DEFAULT '',
            inspection_result TEXT DEFAULT '待定',
            purpose TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS change_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bu TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            change_reason TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            attachments TEXT DEFAULT '',
            change_date TEXT DEFAULT '',
            notify_person TEXT DEFAULT '',
            confirm_date TEXT DEFAULT '',
            confirm_person TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS inspection_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_type TEXT DEFAULT '',
            inspector TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            bg TEXT DEFAULT '',
            bu TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            filename TEXT DEFAULT '',
            file_path TEXT DEFAULT '',
            image_paths TEXT DEFAULT '',
            nas_report_path TEXT DEFAULT '',
            nas_picture_path TEXT DEFAULT '',
            nas_staging_path TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            status TEXT DEFAULT '待审核' CHECK(status IN ('待审核','已通过','已驳回')),
            reviewer TEXT DEFAULT '',
            review_comment TEXT DEFAULT '',
            reject_reason TEXT DEFAULT '',
            inspection_date TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_email TEXT NOT NULL,
            user_name TEXT DEFAULT '',
            action TEXT NOT NULL,
            category TEXT DEFAULT 'system',
            detail TEXT DEFAULT '',
            page TEXT DEFAULT '',
            ip_address TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS changelog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            version TEXT NOT NULL,
            title TEXT DEFAULT '',
            description TEXT DEFAULT '',
            changes TEXT DEFAULT '',
            category TEXT DEFAULT '优化',
            created_by TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS sample_outbound (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sample_id INTEGER NOT NULL,
            qty INTEGER DEFAULT 1,
            out_date TEXT DEFAULT '',
            borrower TEXT DEFAULT '',
            department TEXT DEFAULT '',
            reason TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (sample_id) REFERENCES samples(id)
        );

        CREATE TABLE IF NOT EXISTS operation_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            operator TEXT DEFAULT '',
            action TEXT DEFAULT '',
            target_table TEXT DEFAULT '',
            record_id INTEGER DEFAULT -1,
            detail TEXT DEFAULT '',
            network TEXT DEFAULT '未知',
            deployment TEXT DEFAULT '未知',
            ip_address TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        -- 误删找回（回收站）：删除任意业务记录前先在此留存整行快照，支持一键还原
        CREATE TABLE IF NOT EXISTS deleted_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_table TEXT DEFAULT '',
            record_id INTEGER DEFAULT -1,
            record_json TEXT DEFAULT '',
            summary TEXT DEFAULT '',
            deleted_by TEXT DEFAULT '',
            deleted_network TEXT DEFAULT '未知',
            deleted_deployment TEXT DEFAULT '未知',
            deleted_at TEXT DEFAULT (datetime('now','localtime')),
            restored INTEGER DEFAULT 0,
            restored_by TEXT DEFAULT '',
            restored_at TEXT DEFAULT '',
            extra_json TEXT DEFAULT ''
        );
    ''')

    # 数据库迁移：兼容旧表结构
    try:
        cursor.execute("ALTER TABLE activity_log ADD COLUMN category TEXT DEFAULT 'system'")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE activity_log ADD COLUMN page TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN brand TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN expiry_date TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE change_records ADD COLUMN supplier TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE change_records ADD COLUMN notify_person TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN out_status TEXT DEFAULT '在库'")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN stock_qty INTEGER DEFAULT 1")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE sample_outbound ADD COLUMN is_returned INTEGER DEFAULT 0")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE sample_outbound ADD COLUMN return_date TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE change_records ADD COLUMN rd_team TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE change_records ADD COLUMN sku_confirm_status TEXT DEFAULT '{}'")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE change_records ADD COLUMN overall_status TEXT DEFAULT '待确认'")
    except:
        pass

    # 兼容历史错误逻辑：
    # 旧版本把“推送对象”误写入了 confirm_person/confirm_date。
    # 对仍处于待确认且没有 SKU 确认明细的记录，自动迁移到 notify_person，并清空伪确认信息。
    cursor.execute("""
        UPDATE change_records
        SET notify_person = CASE
                WHEN COALESCE(notify_person, '') = '' THEN COALESCE(confirm_person, '')
                ELSE notify_person
            END,
            confirm_person = '',
            confirm_date = ''
        WHERE COALESCE(overall_status, '待确认') = '待确认'
          AND COALESCE(sku_confirm_status, '{}') = '{}'
          AND COALESCE(confirm_person, '') != ''
    """)

    # 对已有 SKU 确认明细、但表面 confirm_person / confirm_date 为空的历史数据做一次回填，
    # 方便列表展示、导出和后续统计保持一致。
    rows = cursor.execute("""
        SELECT id, sku_confirm_status, confirm_person, confirm_date
        FROM change_records
        WHERE COALESCE(sku_confirm_status, '{}') != '{}'
          AND COALESCE(overall_status, '') IN ('部分确认', '全部确认')
          AND (COALESCE(confirm_person, '') = '' OR COALESCE(confirm_date, '') = '')
    """).fetchall()
    for row in rows:
        derived_person, derived_date = _extract_change_confirmation_summary(dict(row))
        if derived_person or derived_date:
            cursor.execute("""
                UPDATE change_records
                SET confirm_person = CASE
                        WHEN COALESCE(confirm_person, '') = '' THEN ?
                        ELSE confirm_person
                    END,
                    confirm_date = CASE
                        WHEN COALESCE(confirm_date, '') = '' THEN ?
                        ELSE confirm_date
                    END
                WHERE id = ?
            """, (derived_person, derived_date, row['id']))
    try:
        cursor.execute("ALTER TABLE inspection_reports ADD COLUMN supplier TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE inspection_reports ADD COLUMN reject_reason TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE inspection_reports ADD COLUMN nas_report_path TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE inspection_reports ADD COLUMN nas_picture_path TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE inspection_reports ADD COLUMN nas_staging_path TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE inspection_reports ADD COLUMN inspection_date TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN brand TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN sku TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN product_name TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN record_type TEXT DEFAULT 'usage'")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN usage_result TEXT DEFAULT ''")
    except:
        pass

    # 驻厂登记字段迁移：status → inspection_result，supplier 废弃
    try:
        cursor.execute("ALTER TABLE factory_registrations RENAME COLUMN status TO inspection_result")
    except:
        pass
    # supplier 列保留不删（SQLite 旧版不支持 DROP COLUMN），新代码不再读写

    # 仅在表为空时插入种子数据
    cursor.execute("SELECT COUNT(*) FROM categories")
    if cursor.fetchone()[0] == 0:
        seed_data(cursor)

    _apply_pending_samples_sync(conn)
    conn.commit()
    conn.close()
    conn.close()


def seed_data(cursor):
    """插入初始种子数据 - 来源: 实验室设备配备一览表1(1).xlsx (27台设备)"""
    categories = [
        ('力学测试设备', '按键寿命、插拔力、插拔寿命、跌落、振动、摇摆、拉力、纽扣拉力、硬度计'),
        ('环境测试设备', '盐雾、淋雨IPX、恒温恒湿、冷热冲击、温度巡检'),
        ('电学测试设备', 'ESD静电枪、安全性能综合分析仪、阻抗测试'),
        ('电源与电池测试', '直流电源、稳压电源、电池测试、PD负载、模拟电池'),
        ('几何量测与辅助', '二次元影像测量、发热模组温度箱'),
    ]
    cursor.executemany("INSERT INTO categories (name, description) VALUES (?, ?)", categories)

    equipment_list = [
        # ==================== 力学测试设备 (cat 1) ====================
        ('按键寿命试验机', 'LS-AJ-400', 'ET0006', 1, '实验室1', '可用',
         '2024-06-01', 0, '力试', '', '按键耐久性疲劳测试 (2万~10万次) | IEC61058-1'),
        ('微电脑插拔试验机', 'LS-CB-50', 'ET0003', 1, '实验室1', '可用',
         '2024-06-01', 0, '力试', '', 'USB/Type-C/连接器插拔力测试，50kg量程 | EIA-364-09, USB-IF'),
        ('插拔寿命试验机', 'LS-SM-65', 'ET0004', 1, '实验室1', '可用',
         '2024-06-01', 0, '力试', '', '连接器拔插寿命测试，65mm行程 (5000~10000次) | EIA-364-09'),
        ('单翼跌落试验机', 'LS-DL-150', 'ET0001', 1, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', 'ISTA 6-Amazon SIOC 包装跌落测试，150cm | ISTA 6A 亚马逊原箱发货认证'),
        ('模拟运输振动试验机', 'LS-YS-100', 'ET0007', 1, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', '随机/正弦振动模拟长途物流，100kg负载 | ASTM D4728, GB/T4857.7'),
        ('摇摆试验机', 'LS-YB-600', 'ET0005', 1, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', '线材弯折摇摆±90°疲劳测试 | IEC60335-1 §25.14'),
        ('万能拉力试验机', 'GR-WCJ5T', 'ET0008', 1, '实验室2', '可用',
         '2024-06-01', 0, '固润', '', '拉伸/压缩/弯曲/剥离，5T量程 | ISO527, ASTM D3359'),
        ('纽扣拉力试验机', '', None, 1, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '纽扣/按扣/四合扣拉力测试'),
        ('硬度计', '', 'QA049', 1, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '材料硬度测试 (金属/塑料/橡胶)'),
        # ==================== 环境测试设备 (cat 2) ====================
        ('盐雾试验机', 'LS-UT-60', 'ET0002', 2, '盐雾实验室', '可用',
         '2024-06-01', 0, '力试', '', '中性盐雾腐蚀加速测试 (24H~96H+) | ISO9227, ASTM B117, GB/T10125'),
        ('可编程淋雨试验机', 'LS-IPX3456-512', 'ET0010', 2, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', 'IPX3/4/5/6 防水等级测试 | IEC60529/GB/T4208'),
        ('温湿度可编程试验机', 'LS-TH-800Z', 'ET0009', 2, '实验室3', '可用',
         '2024-06-01', 0, '力试', '', '恒温恒湿 (-45~150℃/0~100%RH) 800L | IEC60068-2-78/2-30'),
        ('热冲击试验机', 'LS-THS-180Z', 'ET0011', 2, '实验室3', '可用',
         '2024-06-01', 0, '力试', '', '高低温骤变冲击 (-70~+150℃, 5min转换) | IEC60068-2-14'),
        ('温度巡检仪', '', 'QA018', 2, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '多点温度实时巡检记录 (2台)'),
        # ==================== 电学测试设备 (cat 3) ====================
        ('ESD静电枪试验机', 'PESD6020', 'ET0012', 3, 'ESD静电房', '可用',
         '2024-06-01', 0, '普锐马', '', '接触放电±4kV/空气放电±8kV，20kV | IEC61000-4-2'),
        ('安全性能综合分析仪', 'AC1651B', 'ET013', 3, '实验室1', '可用',
         '2024-06-01', 0, '安规', '', '耐压(1500V)/绝缘(≥2MΩ)/接地/泄漏(≤0.75mA) | IEC60335-1, IEC62368-1'),
        ('阻抗测试仪', '', 'QA073', 3, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '电阻/阻抗/导通精密测量'),
        # ==================== 电源与电池测试 (cat 4) ====================
        ('电池综合测试仪', '', 'INST0023', 4, '品质部实验室', '可用',
         '2024-09-01', 0, '', '', '电池容量/内阻/充放电/BMS保护测试 | UN38.3, IEC62133, UL2054'),
        ('模拟电池测试仪', '', 'INST0024', 4, '品质部实验室', '可用',
         '2024-09-01', 0, '', '', '模拟电池异常状态(过压/欠压/短路) | IEC62368-1'),
        ('大功率直流电源', '', 'INST0025', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '100A/6000W大功率直流供电 | UL1995, IEC60335-2-40'),
        ('PD负载测试仪', '', 'QA009', 4, '品质部实验室', '可用',
         '2025-03-01', 0, '', '', 'USB-PD 3.0/3.1快充协议测试 (5V~20V) | USB-IF PD规范'),
        ('负载测试仪', '', 'QA001', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '电子负载电流/功率测试'),
        ('直流电源', '', 'QA017', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '直流供电电源 (2台)'),
        ('直流稳压电源', '', 'INST0021', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '高精度直流稳压供电'),
        ('可调节直流稳压电源', '', 'QA044', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '可调电压电流直流稳压输出'),
        # ==================== 几何量测与辅助 (cat 5) ====================
        ('二次元测量仪器', '', None, 5, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '光学影像精密测量 (微米级精度) | ISO10360-7'),
        ('发热模组温度测试箱', '', None, 5, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '加热服装发热模组温升曲线监控 | UL130, IEC60335-2-17'),
    ]
    cursor.executemany(
        """INSERT INTO equipment
        (name, model, serial_number, category_id, location, status,
         purchase_date, price, supplier, warranty_expiry, description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        equipment_list
    )

    users = [
        ('Carl Dong董献民', 'ACE/ACE', 'ACE', 'ACE', '管理员'),
        ('joung.yuan袁毅洪', 'BOC/Aura', 'BOC', 'Aura', '管理员'),
        ('haruna.wei韦梦婷', 'Langis LLC/BigRock', 'Langis LLC', 'BigRock', '管理员'),
        ('amelia.han韩亚南', 'no brand/Epicarry', 'no brand', 'Epicarry', '管理员'),
        ('teddy.li黎晓锋', 'Root/KPL', 'Root', 'KPL', '管理员'),
        ('colin.xu徐胜涛', 'Z_Archived/Kronos', 'Z_Archived', 'Kronos', '管理员'),
        ('lucy.ning宁小连', 'no brand', 'no brand', '', '管理员'),
        ('ken.huang黄海森', 'Orion', '', 'Orion', '管理员'),
        ('lainey.pan潘杨阳', 'Parts', '', 'Parts', '管理员'),
        ('fowler.zhai翟始福', 'RaChat', '', 'RaChat', '管理员'),
        ('leo.wu吴嘉俊', 'Root', 'Root', '', '管理员'),
        ('wenzel.chen陈文钊', 'Root-Misc', 'Root-Misc', '', '管理员'),
        ('bruce.cheng程强', 'TheUnicorn', '', 'TheUnicorn', '管理员'),
    ]
    cursor.executemany(
        "INSERT INTO users (name, department, phone, email, role) VALUES (?, ?, ?, ?, ?)",
        users
    )

    today = date.today()
    borrow_records = [
        (1, 8, str(today - timedelta(days=5)), str(today + timedelta(days=2)),
         'IEC61058-1', '09:00', '17:00', '', '', 'PD5K 蓝牙键盘',
         None, '借出中', '新品按键寿命验证', ''),
        (11, 13, str(today - timedelta(days=10)), str(today - timedelta(days=3)),
         'IEC60529 IPX5', '10:00', '16:00', 'TURBRO', 'TB-JXH-001', 'TURBRO接线盒',
         str(today - timedelta(days=3)), '已归还', 'TURBRO接线盒防水测试', '测试完成，数据正常'),
        (5, 12, str(today - timedelta(days=15)), str(today - timedelta(days=8)),
         'ASTM D4728', '08:30', '12:00', '', '', '房车空调外包装',
         str(today - timedelta(days=9)), '已归还', '出口包装振动测试', ''),
    ]
    cursor.executemany(
        """INSERT INTO borrow_records
        (equipment_id, user_id, borrow_date, expected_return_date,
         test_standard, test_start_time, test_end_time, brand, sku, product_name,
         actual_return_date, status, purpose, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        borrow_records
    )

    maintenance_records = [
        (10, str(today - timedelta(days=60)), '定期保养', '盐雾试验机月度保养：清洗喷嘴、检查加热管、补充盐溶液',
         0, 'Carl Dong董献民', str(today + timedelta(days=300)), '已完成', '设备运行正常，盐雾沉降量符合标准'),
        (15, str(today - timedelta(days=10)), '校准', 'ESD静电枪年度计量校准（±4kV接触/±8kV空气）',
         800.00, '计量所', str(today + timedelta(days=355)), '已完成', '校准证书编号: CAL-2026-ESD-001'),
        (16, str(today - timedelta(days=90)), '校准', '安规综合分析仪年度校准（耐压/绝缘/泄漏电流）',
         600.00, '计量所', str(today + timedelta(days=270)), '已完成', '校准证书编号: CAL-2026-SFT-001'),
        (12, str(today + timedelta(days=15)), '定期保养', '温湿度可编程试验机季度保养：清洁滤网、检查制冷系统',
         0, 'teddy.li黎晓锋', None, '计划中', '按季度保养计划执行'),
        (11, str(today + timedelta(days=30)), '定期保养', '淋雨试验机月度保养：检查水泵压力、清洁各喷头',
         0, 'bruce.cheng程强', None, '计划中', '按月度保养计划执行'),
    ]
    cursor.executemany(
        """INSERT INTO maintenance_records
        (equipment_id, maintenance_date, maintenance_type, description,
         cost, technician, next_maintenance_date, status, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        maintenance_records
    )


# ==================== 分类操作 ====================

def get_categories():
    """获取所有设备分类"""
    conn = get_connection()
    rows = conn.execute("SELECT * FROM categories ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_category(name, description=''):
    """添加设备分类"""
    conn = get_connection()
    try:
        conn.execute("INSERT INTO categories (name, description) VALUES (?, ?)", (name, description))
        conn.commit()
        return True, "分类添加成功"
    except sqlite3.IntegrityError:
        return False, "分类名称已存在"
    finally:
        conn.close()


def update_category(cat_id, name, description):
    """更新设备分类"""
    conn = get_connection()
    try:
        conn.execute("UPDATE categories SET name=?, description=? WHERE id=?", (name, description, cat_id))
        conn.commit()
        return True, "分类更新成功"
    except sqlite3.IntegrityError:
        return False, "分类名称已存在"
    finally:
        conn.close()


@recoverable("categories")
def delete_category(cat_id):
    """删除设备分类"""
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM equipment WHERE category_id=?", (cat_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该分类下有 {count} 台设备，无法删除"
    conn.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    conn.commit()
    conn.close()
    return True, "分类删除成功"


# ==================== 设备操作 ====================

def get_equipment(search='', category_id=None, status=None, location='', page=1, per_page=20):
    """查询设备列表（支持搜索和筛选）"""
    conn = get_connection()
    conditions = []
    params = []

    if search:
        conditions.append("(e.name LIKE ? OR e.model LIKE ? OR e.serial_number LIKE ?)")
        params.extend([f'%{search}%'] * 3)
    if category_id:
        conditions.append("e.category_id = ?")
        params.append(category_id)
    if status:
        conditions.append("e.status = ?")
        params.append(status)
    if location:
        conditions.append("e.location LIKE ?")
        params.append(f'%{location}%')

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    total = conn.execute(f"SELECT COUNT(*) FROM equipment e {where}", params).fetchone()[0]

    offset = (page - 1) * per_page
    sql = f"""
        SELECT e.*, c.name as category_name
        FROM equipment e
        LEFT JOIN categories c ON e.category_id = c.id
        {where}
        ORDER BY e.id DESC
        LIMIT ? OFFSET ?
    """
    rows = conn.execute(sql, params + [per_page, offset]).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_equipment_by_id(eq_id):
    """根据 ID 获取设备详情"""
    conn = get_connection()
    row = conn.execute(
        """SELECT e.*, c.name as category_name
           FROM equipment e LEFT JOIN categories c ON e.category_id = c.id
           WHERE e.id=?""", (eq_id,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


@audit("新增设备", "equipment", capture_new_id=True)
def add_equipment(data):
    """添加新设备"""
    conn = get_connection()
    try:
        conn.execute("""
            INSERT INTO equipment (name, model, serial_number, category_id, location, status,
                                   purchase_date, price, supplier, warranty_expiry, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data['name'], data.get('model', ''), data.get('serial_number', ''),
            data.get('category_id'), data.get('location', ''), data.get('status', '可用'),
            data.get('purchase_date', ''), data.get('price', 0), data.get('supplier', ''),
            data.get('warranty_expiry', ''), data.get('description', '')
        ))
        conn.commit()
        return True, "设备添加成功"
    except sqlite3.IntegrityError as e:
        if 'serial_number' in str(e):
            return False, "设备编号已存在"
        return False, f"添加失败: {e}"
    finally:
        conn.close()


@audit("更新设备", "equipment")
def update_equipment(eq_id, data):
    """更新设备信息"""
    conn = get_connection()
    try:
        conn.execute("""
            UPDATE equipment SET name=?, model=?, serial_number=?, category_id=?, location=?,
            status=?, purchase_date=?, price=?, supplier=?, warranty_expiry=?, description=?,
            updated_at=datetime('now','localtime')
            WHERE id=?
        """, (
            data['name'], data.get('model', ''), data.get('serial_number', ''),
            data.get('category_id'), data.get('location', ''), data.get('status', '可用'),
            data.get('purchase_date', ''), data.get('price', 0), data.get('supplier', ''),
            data.get('warranty_expiry', ''), data.get('description', ''), eq_id
        ))
        conn.commit()
        return True, "设备更新成功"
    except sqlite3.IntegrityError:
        return False, "设备编号已存在"
    finally:
        conn.close()


@audit("删除设备", "equipment")
@recoverable("equipment")
def delete_equipment(eq_id):
    """删除设备"""
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM borrow_records WHERE equipment_id=?", (eq_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该设备有 {count} 条借用记录，无法删除（可将其状态设为'报废'）"
    count = conn.execute("SELECT COUNT(*) FROM maintenance_records WHERE equipment_id=?", (eq_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该设备有 {count} 条维护记录，无法删除（可将其状态设为'报废'）"
    conn.execute("DELETE FROM equipment WHERE id=?", (eq_id,))
    conn.commit()
    conn.close()
    return True, "设备删除成功"


def import_equipment_batch(records):
    """批量导入设备"""
    conn = get_connection()
    success = 0
    errors = []
    for i, rec in enumerate(records):
        try:
            conn.execute("""
                INSERT INTO equipment (name, model, serial_number, category_id, location, status,
                                       purchase_date, price, supplier, warranty_expiry, description)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rec.get('name', f'设备{i + 1}'), rec.get('model', ''), rec.get('serial_number', ''),
                rec.get('category_id'), rec.get('location', ''), rec.get('status', '可用'),
                rec.get('purchase_date', ''), rec.get('price', 0), rec.get('supplier', ''),
                rec.get('warranty_expiry', ''), rec.get('description', '')
            ))
            success += 1
        except Exception as e:
            errors.append(f"第{i + 1}行: {e}")
    conn.commit()
    conn.close()
    return success, errors


def get_equipment_for_select():
    """获取设备下拉选项"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, name, serial_number, status FROM equipment WHERE status != '报废' ORDER BY name"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 人员操作 ====================

def get_users(search=''):
    """获取人员列表"""
    conn = get_connection()
    if search:
        rows = conn.execute(
            "SELECT * FROM users WHERE name LIKE ? OR department LIKE ? ORDER BY id",
            (f'%{search}%', f'%{search}%')
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM users ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@audit("新增人员", "users", capture_new_id=True)
def add_user(data):
    """添加人员"""
    conn = get_connection()
    conn.execute(
        "INSERT INTO users (name, department, phone, email, role) VALUES (?, ?, ?, ?, ?)",
        (data['name'], data.get('department', ''), data.get('phone', ''),
         data.get('email', ''), data.get('role', '普通用户'))
    )
    conn.commit()
    conn.close()
    return True, "人员添加成功"


@audit("更新人员", "users")
def update_user(user_id, data):
    """更新人员信息"""
    conn = get_connection()
    conn.execute(
        "UPDATE users SET name=?, department=?, phone=?, email=?, role=? WHERE id=?",
        (data['name'], data.get('department', ''), data.get('phone', ''),
         data.get('email', ''), data.get('role', '普通用户'), user_id)
    )
    conn.commit()
    conn.close()
    return True, "人员更新成功"


@audit("删除人员", "users")
@recoverable("users")
def delete_user(user_id):
    """删除人员"""
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM borrow_records WHERE user_id=?", (user_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该人员有 {count} 条借用记录，无法删除"
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return True, "人员删除成功"


# ==================== 借用操作 ====================

@audit("使用登记", "borrow_records", capture_new_id=True)
def borrow_equipment(equipment_id, user_id, borrow_date, expected_return_date,
                     purpose='', notes='', test_standard='', test_start_time='', test_end_time='',
                     brand='', sku='', product_name=''):
    """使用登记（实验室内部测试预约，不改变设备状态）"""
    conn = get_connection()
    eq = conn.execute("SELECT status FROM equipment WHERE id=?", (equipment_id,)).fetchone()
    if not eq:
        conn.close()
        return False, "设备不存在"

    # 按日期 + 时间段判断冲突；历史记录没有时间时按整天占用处理。
    existing_rows = conn.execute("""
        SELECT id, borrow_date, expected_return_date, test_start_time, test_end_time
        FROM borrow_records
        WHERE equipment_id=? AND status='借出中' AND record_type='usage'
    """, (equipment_id,)).fetchall()

    def _bounds(start_date, end_date, start_time='', end_time=''):
        start = datetime.strptime(f"{start_date} {(start_time or '00:00')[:5]}", "%Y-%m-%d %H:%M")
        end_clock = (end_time or '23:59')[:5]
        end = datetime.strptime(f"{end_date} {end_clock}", "%Y-%m-%d %H:%M")
        return start, end

    new_start, new_end = _bounds(str(borrow_date), str(expected_return_date), test_start_time, test_end_time)
    for existing in existing_rows:
        old_start, old_end = _bounds(
            existing['borrow_date'], existing['expected_return_date'],
            existing['test_start_time'], existing['test_end_time'])
        if new_start < old_end and old_start < new_end:
            conn.close()
            return False, "该设备在此时间段已有测试预约，请调整日期或时间"

    conn.execute(
        """INSERT INTO borrow_records (equipment_id, user_id, borrow_date, expected_return_date,
           purpose, notes, test_standard, test_start_time, test_end_time, brand, sku, product_name, record_type)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'usage')""",
        (equipment_id, user_id, borrow_date, expected_return_date,
         purpose, notes, test_standard, test_start_time, test_end_time,
         brand, sku, product_name)
    )
    # 使用登记不改变设备物理状态，设备仍在实验室
    conn.commit()
    conn.close()
    return True, "使用登记成功"


@audit("借用出库", "borrow_records", capture_new_id=True)
def checkout_equipment(equipment_id, user_id, borrow_date, expected_return_date,
                       purpose='', notes=''):
    """借用设备出库（物理借出实验室）"""
    conn = get_connection()
    eq = conn.execute("SELECT status FROM equipment WHERE id=?", (equipment_id,)).fetchone()
    if not eq:
        conn.close()
        return False, "设备不存在"
    if eq['status'] != '可用':
        conn.close()
        return False, f"设备当前状态为「{eq['status']}」，无法借用出库"

    conn.execute(
        """INSERT INTO borrow_records (equipment_id, user_id, borrow_date, expected_return_date,
           purpose, notes, status, record_type)
           VALUES (?, ?, ?, ?, ?, ?, '已出库', 'borrow')""",
        (equipment_id, user_id, borrow_date, expected_return_date, purpose, notes)
    )
    conn.execute(
        "UPDATE equipment SET status='借出', updated_at=datetime('now','localtime') WHERE id=?",
        (equipment_id,)
    )
    conn.commit()
    conn.close()
    return True, "设备借用出库成功"


def get_active_borrows():
    """获取当前使用中的记录（实验室内部使用登记，type='usage'）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT b.*, e.name as equipment_name, e.serial_number, u.name as user_name, u.department
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        WHERE b.record_type = 'usage'
          AND b.status IN ('使用中','借出中')
          AND (b.test_end_time = '' OR b.test_end_time >= datetime('now','localtime'))
        ORDER BY b.borrow_date DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_equipment_schedule(equipment_id=None):
    """获取设备占用排期，覆盖内部使用预约和物理外借记录。"""
    conn = get_connection()
    params = []
    equipment_clause = ""
    if equipment_id:
        equipment_clause = "AND b.equipment_id = ?"
        params.append(equipment_id)
    rows = conn.execute(f"""
        SELECT b.*, e.name as equipment_name, e.serial_number,
               u.name as user_name, u.department
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        WHERE (
            (b.record_type = 'usage' AND b.status IN ('使用中', '借出中'))
            OR (b.record_type = 'borrow' AND b.status IN ('借出中', '已出库'))
        )
        {equipment_clause}
        ORDER BY e.name, b.borrow_date
    """, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@audit("归还设备", "borrow_records")
def return_equipment(record_id, return_date):
    """归还设备"""
    conn = get_connection()
    record = conn.execute("SELECT * FROM borrow_records WHERE id=?", (record_id,)).fetchone()
    if not record:
        conn.close()
        return False, "借用记录不存在"
    if record['status'] == '已归还':
        conn.close()
        return False, "该设备已归还"

    conn.execute(
        "UPDATE borrow_records SET actual_return_date=?, status='已归还' WHERE id=?",
        (return_date, record_id)
    )
    # 实验室使用登记不改变设备物理状态；只有外借记录归还后才恢复可用。
    if record['record_type'] == 'borrow':
        active_borrow = conn.execute(
            "SELECT COUNT(*) FROM borrow_records WHERE equipment_id=? AND record_type='borrow' AND status='已出库'",
            (record['equipment_id'],)
        ).fetchone()[0]
        if active_borrow == 0:
            conn.execute(
                "UPDATE equipment SET status='可用', updated_at=datetime('now','localtime') WHERE id=?",
                (record['equipment_id'],)
            )
    conn.commit()
    conn.close()
    return True, "归还成功"


def get_borrow_records(status=None, equipment_id=None, user_id=None, page=1, per_page=20):
    """查询借用记录"""
    conn = get_connection()
    conditions = []
    params = []

    if status:
        conditions.append("b.status = ?")
        params.append(status)
    if equipment_id:
        conditions.append("b.equipment_id = ?")
        params.append(equipment_id)
    if user_id:
        conditions.append("b.user_id = ?")
        params.append(user_id)

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    total = conn.execute(f"SELECT COUNT(*) FROM borrow_records b {where}", params).fetchone()[0]
    offset = (page - 1) * per_page

    sql = f"""
        SELECT b.*, e.name as equipment_name, e.serial_number, u.name as user_name, u.department
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        {where}
        ORDER BY b.id DESC
        LIMIT ? OFFSET ?
    """
    rows = conn.execute(sql, params + [per_page, offset]).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def update_usage_result(record_id, result):
    """更新使用登记的结果反馈"""
    conn = get_connection()
    conn.execute(
        "UPDATE borrow_records SET usage_result = ? WHERE id = ?",
        (result, record_id)
    )
    conn.commit()
    conn.close()
    return True, "已更新"


def parse_sku_list(sku_str):
    """解析 SKU 字符串为列表，支持英文逗号、中文逗号、顿号、空格等分隔符"""
    import re
    if not sku_str or not sku_str.strip():
        return []
    # 按逗号(中/英)、顿号、空格拆分，过滤空字符串
    parts = re.split(r'[,，、\s]+', sku_str.strip())
    return [p.strip() for p in parts if p.strip()]


def get_change_stats():
    """获取变更管理统计（基于 overall_status）"""
    conn = get_connection()
    total = conn.execute("SELECT COUNT(*) FROM change_records").fetchone()[0]
    all_confirmed = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE overall_status = '全部确认'"
    ).fetchone()[0]
    partial = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE overall_status = '部分确认'"
    ).fetchone()[0]
    pending = total - all_confirmed - partial
    conn.close()
    return {
        "total": total,
        "confirmed": all_confirmed,
        "partial": partial,
        "unconfirmed": pending
    }


# ==================== 维护操作 ====================

@audit("新增维护", "maintenance_records", capture_new_id=True)
def add_maintenance(data):
    """添加维护记录"""
    conn = get_connection()
    conn.execute("""
        INSERT INTO maintenance_records
        (equipment_id, maintenance_date, maintenance_type, description, cost,
         technician, next_maintenance_date, status, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data['equipment_id'], data['maintenance_date'], data.get('maintenance_type', '定期保养'),
        data.get('description', ''), data.get('cost', 0), data.get('technician', ''),
        data.get('next_maintenance_date', ''), data.get('status', '已完成'), data.get('notes', '')
    ))
    if data.get('status') == '进行中':
        conn.execute(
            "UPDATE equipment SET status='维修中', updated_at=datetime('now','localtime') WHERE id=?",
            (data['equipment_id'],)
        )
    conn.commit()
    conn.close()
    return True, "维护记录添加成功"


@audit("更新维护", "maintenance_records")
def update_maintenance(m_id, data):
    """更新维护记录"""
    conn = get_connection()
    previous = conn.execute(
        "SELECT equipment_id, status FROM maintenance_records WHERE id=?", (m_id,)
    ).fetchone()
    if not previous:
        conn.close()
        return False, "维护记录不存在"
    conn.execute("""
        UPDATE maintenance_records SET equipment_id=?, maintenance_date=?, maintenance_type=?,
        description=?, cost=?, technician=?, next_maintenance_date=?, status=?, notes=?
        WHERE id=?
    """, (
        data['equipment_id'], data['maintenance_date'], data.get('maintenance_type', '定期保养'),
        data.get('description', ''), data.get('cost', 0), data.get('technician', ''),
        data.get('next_maintenance_date', ''), data.get('status', '已完成'), data.get('notes', ''), m_id
    ))
    equipment_id = data['equipment_id']
    if data.get('status') == '进行中':
        conn.execute(
            "UPDATE equipment SET status='维修中', updated_at=datetime('now','localtime') WHERE id=?",
            (equipment_id,)
        )
    elif data.get('status') == '已完成':
        active_maintenance = conn.execute(
            "SELECT COUNT(*) FROM maintenance_records WHERE equipment_id=? AND status='进行中' AND id<>?",
            (equipment_id, m_id)
        ).fetchone()[0]
        active_borrow = conn.execute(
            "SELECT COUNT(*) FROM borrow_records WHERE equipment_id=? AND record_type='borrow' AND status='已出库'",
            (equipment_id,)
        ).fetchone()[0]
        if active_maintenance == 0 and active_borrow == 0:
            conn.execute(
                "UPDATE equipment SET status='可用', updated_at=datetime('now','localtime') WHERE id=?",
                (equipment_id,)
            )
    conn.commit()
    conn.close()
    return True, "维护记录更新成功"


@audit("删除维护", "maintenance_records")
@recoverable("maintenance_records")
def delete_maintenance(m_id):
    """删除维护记录"""
    conn = get_connection()
    conn.execute("DELETE FROM maintenance_records WHERE id=?", (m_id,))
    conn.commit()
    conn.close()
    return True, "维护记录删除成功"


def get_maintenance_records(equipment_id=None, page=1, per_page=20):
    """查询维护记录"""
    conn = get_connection()
    if equipment_id:
        total = conn.execute(
            "SELECT COUNT(*) FROM maintenance_records WHERE equipment_id=?", (equipment_id,)
        ).fetchone()[0]
        offset = (page - 1) * per_page
        rows = conn.execute("""
            SELECT m.*, e.name as equipment_name, e.serial_number
            FROM maintenance_records m
            JOIN equipment e ON m.equipment_id = e.id
            WHERE m.equipment_id = ?
            ORDER BY m.maintenance_date DESC
            LIMIT ? OFFSET ?
        """, (equipment_id, per_page, offset)).fetchall()
    else:
        total = conn.execute("SELECT COUNT(*) FROM maintenance_records").fetchone()[0]
        offset = (page - 1) * per_page
        rows = conn.execute("""
            SELECT m.*, e.name as equipment_name, e.serial_number
            FROM maintenance_records m
            JOIN equipment e ON m.equipment_id = e.id
            ORDER BY m.maintenance_date DESC
            LIMIT ? OFFSET ?
        """, (per_page, offset)).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_upcoming_maintenance(days=30):
    """获取未来 N 天内计划的维护"""
    conn = get_connection()
    today = date.today()
    end_date = today + timedelta(days=days)
    rows = conn.execute("""
        SELECT m.*, e.name as equipment_name, e.serial_number
        FROM maintenance_records m
        JOIN equipment e ON m.equipment_id = e.id
        WHERE m.next_maintenance_date BETWEEN ? AND ?
        ORDER BY m.next_maintenance_date
    """, (str(today), str(end_date))).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 统计操作 ====================

def get_dashboard_stats():
    """获取看板统计数据"""
    conn = get_connection()
    stats = {
        'total': conn.execute("SELECT COUNT(*) FROM equipment").fetchone()[0],
        'available': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='可用'").fetchone()[0],
        'in_use': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='借出中'").fetchone()[0],
        'borrowed': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='借出'").fetchone()[0],
        'maintenance': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='维修中'").fetchone()[0],
        'scrapped': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='报废'").fetchone()[0],
        'total_users': conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        'active_borrows': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='借出中'").fetchone()[0],
        'total_value': conn.execute("SELECT COALESCE(SUM(price), 0) FROM equipment").fetchone()[0],
    }
    conn.close()
    return stats


def get_status_distribution():
    """设备状态分布"""
    conn = get_connection()
    rows = conn.execute("SELECT status, COUNT(*) as count FROM equipment GROUP BY status").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_category_distribution():
    """分类设备统计"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT c.name, COUNT(e.id) as count, COALESCE(SUM(e.price), 0) as total_value
        FROM categories c
        LEFT JOIN equipment e ON c.id = e.category_id
        GROUP BY c.id
        ORDER BY count DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_borrow_stats():
    """借用统计"""
    conn = get_connection()
    stats = {
        'total': conn.execute("SELECT COUNT(*) FROM borrow_records").fetchone()[0],
        'active': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='借出中'").fetchone()[0],
        'returned': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='已归还'").fetchone()[0],
    }
    rows = conn.execute("""
        SELECT strftime('%Y-%m', borrow_date) as month, COUNT(*) as count
        FROM borrow_records
        WHERE borrow_date >= date('now', '-12 months')
        GROUP BY month ORDER BY month
    """).fetchall()
    stats['monthly'] = [dict(r) for r in rows]
    conn.close()
    return stats


def get_maintenance_cost_stats():
    """维护费用统计"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT strftime('%Y-%m', maintenance_date) as month,
               COUNT(*) as count, COALESCE(SUM(cost), 0) as total_cost
        FROM maintenance_records
        WHERE maintenance_date >= date('now', '-12 months')
        GROUP BY month ORDER BY month
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_equipment():
    """获取所有设备（导出用）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT e.*, c.name as category_name
        FROM equipment e LEFT JOIN categories c ON e.category_id = c.id
        ORDER BY e.id
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_borrow_records_export():
    """获取所有借用记录（导出用）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT b.id, e.name as 设备名称, e.serial_number as 设备编号,
               u.name as 借用人, u.department as 部门,
               b.borrow_date as 借用日期, b.expected_return_date as 预计归还,
               b.actual_return_date as 实际归还, b.status as 状态,
               b.purpose as 用途, b.notes as 备注
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        ORDER BY b.id DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_maintenance_export():
    """获取所有维护记录（导出用）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT m.id, e.name as 设备名称, e.serial_number as 设备编号,
               m.maintenance_date as 维护日期, m.maintenance_type as 维护类型,
               m.description as 描述, m.cost as 费用,
               m.technician as 技术人员, m.next_maintenance_date as 下次维护,
               m.status as 状态, m.notes as 备注
        FROM maintenance_records m
        JOIN equipment e ON m.equipment_id = e.id
        ORDER BY m.id DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_recent_borrows(limit=8):
    """获取最近借用记录"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT b.*, e.name as equipment_name, e.serial_number, u.name as user_name
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        ORDER BY b.id DESC
        LIMIT ?
    """, (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 测试标准库 ====================

# 设备关键词 → 关联测试标准列表
EQUIPMENT_STANDARD_MAP = {
    '按键寿命': ['IEC61058-1 (器具开关耐久性)'],
    '插拔力': ['EIA-364-09 (电子连接器插拔力)', 'USB-IF 协会规范'],
    '插拔寿命': ['EIA-364-09 (电子连接器拔插寿命)', 'USB-IF 协会规范'],
    '跌落': ['ISTA 6A (Amazon SIOC 原箱发货认证)'],
    '单翼跌落': ['ISTA 6A (Amazon SIOC 原箱发货认证)'],
    '单臂跌落': ['ISTA 6A (Amazon SIOC 原箱发货认证)'],
    '振动': ['ASTM D4728 (随机振动)', 'GB/T 4857.7 (正弦振动)'],
    '摇摆': ['IEC60335-1 §25.14 (电源线弯折)'],
    '拉力': ['ISO527 (塑料拉伸性能)', 'ASTM D3359 (附着力测试)'],
    '纽扣': ['ASTM D 相关 (纽扣拉力)'],
    '硬度': ['ISO 6508 / ASTM E18 (洛氏硬度)'],
    '盐雾': ['ISO9227 (盐雾试验)', 'ASTM B117 (盐雾试验)', 'GB/T10125 (盐雾试验)'],
    '淋雨': ['IEC60529 (IP代码)', 'GB/T4208 (IP代码)'],
    '防水': ['IEC60529 (IP代码)', 'GB/T4208 (IP代码)'],
    '温湿度': ['IEC60068-2-78 (恒定湿热)', 'IEC60068-2-30 (交变湿热)'],
    '热冲击': ['IEC60068-2-14 (温度变化)', 'GB/T2423.22 (温度变化)'],
    '冷热冲击': ['IEC60068-2-14 (温度变化)', 'GB/T2423.22 (温度变化)'],
    'ESD': ['IEC61000-4-2 (静电放电抗扰度)'],
    '静电': ['IEC61000-4-2 (静电放电抗扰度)'],
    '安规': ['IEC60335-1 (家用电器安全)', 'IEC62368-1 (音视频/IT安全)'],
    '安全性能': ['IEC60335-1 (家用电器安全)', 'IEC62368-1 (音视频/IT安全)'],
    '电池综合': ['UN38.3 (锂电池运输安全)', 'IEC62133 (锂电池安全)', 'UL2054 (北美电池安全)'],
    '模拟电池': ['IEC62368-1 (故障条件测试)'],
    '大功率': ['UL1995 (冷暖设备)', 'IEC60335-2-40 (热泵/空调)'],
    '直流电源': ['IEC62368-1 (电气安全)'],
    'PD负载': ['USB-IF PD 3.0/3.1 规范', 'EN62368-1'],
    '负载测试': ['IEC62368-1 (电气安全)'],
    '阻抗': ['IEC 相关 (阻抗测试)'],
    '直流稳压': ['IEC62368-1 (电气安全)'],
    '可调节': ['IEC62368-1 (电气安全)'],
    '二次元': ['ISO10360-7 (影像测量仪)'],
    '发热模组': ['UL130 (电热服装安全)', 'IEC60335-2-17 (电热毯/服装)'],
    '温度巡检': ['IEC60068-2-78 (恒定湿热)'],
}

# 全部标准列表（用于下拉选择）
ALL_STANDARDS = sorted(set(
    std for standards in EQUIPMENT_STANDARD_MAP.values() for std in standards
))


def find_standards_for_equipment(equipment_name):
    """根据设备名称查找关联的测试标准列表"""
    if not equipment_name:
        return []
    for keyword, standards in EQUIPMENT_STANDARD_MAP.items():
        if keyword in equipment_name:
            return standards
    return []


def get_all_test_standards():
    """获取所有测试标准列表（供下拉选择）"""
    return ALL_STANDARDS


# ==================== 品质人员导入 ====================

def import_personnel_from_excel(filepath=None):
    """从原始名单Excel导入品质人员"""
    import pandas as pd
    if filepath is None:
        filepath = os.path.join(PROJECT_ROOT,
                                'SainStore实验室文件', '原始名单.xlsx')
    if not os.path.exists(filepath):
        return 0, "文件不存在"

    df = pd.read_excel(filepath, engine='openpyxl')
    conn = get_connection()
    count = 0
    for _, row in df.iterrows():
        name = str(row.get('品质人员', '')).strip() if pd.notna(row.get('品质人员')) else ''
        if not name or name == 'nan':
            continue
        bg = str(row.get('BG', '')).strip() if pd.notna(row.get('BG')) else ''
        bu = str(row.get('BU', '')).strip() if pd.notna(row.get('BU')) else ''
        brand = str(row.get('brand', '')).strip() if pd.notna(row.get('brand')) else ''
        dept = f"{bg}/{bu}" if bg and bu else (bg or bu or '品质部')

        existing = conn.execute("SELECT id FROM users WHERE name=?", (name,)).fetchone()
        if existing:
            conn.execute("UPDATE users SET department=?, phone=?, email=? WHERE id=?",
                         (dept, bg, bu, existing[0]))
        else:
            conn.execute("INSERT INTO users (name, department, phone, email, role) VALUES (?, ?, ?, ?, ?)",
                         (name, dept, bg, bu, '管理员'))
        count += 1
    conn.commit()
    conn.close()
    return count, "导入成功"


# ==================== 样品管理 ====================

def import_samples_from_excel(filepath=None, append_only=False):
    """从签样记录Excel导入样品（遍历所有BG工作表）"""
    import openpyxl
    if filepath is None:
        filepath = os.path.join(PROJECT_ROOT,
                                'SainStore实验室文件', '签样记录 （最新）.xlsx')
    if not os.path.exists(filepath):
        return 0, "文件不存在"

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        conn = get_connection()
        count = 0
        skip = {'不是封样可参考', '原IBG'}

        if not append_only:
            conn.execute("DELETE FROM samples")

        for sheet_name in wb.sheetnames:
            if sheet_name in skip:
                continue
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                bg_val = str(ws.cell(r, 2).value or '').strip()
                sku_val = str(ws.cell(r, 3).value or '').strip()
                name_val = str(ws.cell(r, 4).value or '').strip()
                if not name_val or name_val.lower() == 'none':
                    continue
                sign_date = str(ws.cell(r, 5).value or '').strip()
                supplier = str(ws.cell(r, 6).value or '').strip()
                notes = str(ws.cell(r, 7).value or '').strip()
                location = str(ws.cell(r, 8).value or '').strip()

                conn.execute(
                    """INSERT INTO samples (bg, sku, sample_name, sign_date, supplier, notes, location)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (bg_val, sku_val, name_val, sign_date, supplier, notes, location)
                )
                count += 1
        wb.close()
        conn.commit()
        conn.close()
        return count, f"成功导入 {count} 条样品"
    except Exception as e:
        return 0, str(e)


def import_samples_from_workbook_file(file_obj, db_conn=None):
    """从原始签样记录工作簿自动清洗并合并样品数据。"""
    import io
    import openpyxl

    skip_sheets = {'不是封样可参考', '原IBG', 'Sheet1'}
    header_alias = {
        'bg': 'bg',
        'sku': 'sku',
        '样品名称': 'sample_name',
        'sample_name': 'sample_name',
        '签样日期': 'sign_date',
        'sign_date': 'sign_date',
        '供应商': 'supplier',
        'supplier': 'supplier',
        '备注': 'notes',
        'notes': 'notes',
        '样品放置区域': 'location',
        'location': 'location',
        '品牌': 'brand',
        'brand': 'brand',
        '到期日期': 'expiry_date',
        'expiry_date': 'expiry_date',
    }
    required = ['bg', 'sku', 'sample_name', 'sign_date', 'supplier', 'notes', 'location', 'brand', 'expiry_date']

    raw_bytes = file_obj.getvalue() if hasattr(file_obj, 'getvalue') else file_obj.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)

    owns_conn = db_conn is None
    conn = db_conn or get_connection()
    inserted = 0
    updated = 0
    skipped = 0
    parsed_rows = []
    warnings = []

    try:
        current_rows = conn.execute("""
            SELECT id, bg, sku, sample_name, sign_date, supplier, brand, notes,
                   location, expiry_date, out_status, stock_qty
            FROM samples
        """).fetchall()
        existing_map = {}
        fallback_map = {}
        for row in current_rows:
            row_dict = dict(row)
            key = _sample_sync_key(row_dict)
            if key and key not in existing_map:
                existing_map[key] = row_dict
            fallback_key = _sample_fallback_key(row_dict)
            if fallback_key and fallback_key not in fallback_map:
                fallback_map[fallback_key] = row_dict

        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                continue

            ws = wb[sheet_name]
            preview = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
            if not preview:
                continue

            headers = [str(v).strip() if v is not None else '' for v in preview[0]]
            if not any(headers):
                continue

            col_map = {}
            for idx, header in enumerate(headers):
                normalized = header.strip()
                normalized_lower = normalized.lower()
                if normalized in header_alias:
                    col_map[idx] = header_alias[normalized]
                elif normalized_lower in header_alias:
                    col_map[idx] = header_alias[normalized_lower]

            if not {'bg', 'sku', 'sample_name'}.intersection(set(col_map.values())):
                warnings.append(f"工作表「{sheet_name}」未识别到核心列，已跳过。")
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                record = {field: '' for field in required}
                for idx, target_field in col_map.items():
                    if idx < len(row):
                        value = row[idx]
                        record[target_field] = '' if value is None else str(value).strip()

                if not record['bg']:
                    record['bg'] = sheet_name.split('-')[0].strip()

                record['sku'] = _sample_placeholder_if_missing(record['sku'])
                record['supplier'] = _sample_placeholder_if_missing(record['supplier'])
                record['brand'] = _sample_placeholder_if_missing(record['brand'])

                sign_date = record['sign_date'].replace('/', '.').replace('-', '.')
                record['sign_date'] = sign_date
                if sign_date and not record['expiry_date']:
                    try:
                        if len(sign_date) == 7:
                            sign_seed = sign_date + '.01'
                        else:
                            sign_seed = sign_date
                        sign_dt = datetime.strptime(sign_seed[:10].replace('.', '-'), '%Y-%m-%d')
                        record['expiry_date'] = (sign_dt + timedelta(days=365)).strftime('%Y-%m-%d')
                    except Exception:
                        record['expiry_date'] = ''

                key = _sample_sync_key(record)
                # 同名同 BG 同日期但无样品名称的行：追加供应商+备注防止误合并
                if not _sample_has_meaningful_sku(record.get('sku')) and not record['sample_name']:
                    key = f"{key}||{_sample_sync_text(record.get('supplier'))}||{_sample_sync_text(record.get('notes'))}"
                # 冲突时追加计数器，确保 Excel 源数据每一行都有独立记录
                if key in existing_map:
                    key = f"{key}||dup_{inserted + updated}"
                if not key:
                    skipped += 1
                    continue

                # 剔除空记录：sample_name、sku、supplier、sign_date、location、notes 全部为空或占位符
                if _is_empty_sample_record(record):
                    skipped += 1
                    continue

                existing = existing_map.get(key)
                if not existing:
                    fallback_key = _sample_fallback_key(record)
                    fallback_existing = fallback_map.get(fallback_key) if fallback_key else None
                    if fallback_existing and (
                        not _sample_has_meaningful_sku(record.get('sku'))
                        or not _sample_has_meaningful_sku(fallback_existing.get('sku'))
                    ):
                        existing = fallback_existing
                parsed_rows.append(record.copy())
                if existing:
                    update_values = [
                        record['bg'], record['sku'], record['sample_name'], record['sign_date'],
                        record['supplier'], record['brand'], record['notes'], record['location'],
                        record['expiry_date'], existing['id']
                    ]
                    conn.execute("""
                        UPDATE samples
                        SET bg=?, sku=?, sample_name=?, sign_date=?, supplier=?,
                            brand=?, notes=?, location=?, expiry_date=?
                        WHERE id=?
                    """, update_values)
                    updated += 1
                    updated_row = dict(existing)
                    updated_row.update(record)
                    existing_map[key] = updated_row
                    fallback_key = _sample_fallback_key(updated_row)
                    if fallback_key:
                        fallback_map[fallback_key] = updated_row
                else:
                    cursor = conn.execute("""
                        INSERT INTO samples (
                            bg, sku, sample_name, sign_date, supplier, brand, notes,
                            location, expiry_date, out_status, stock_qty
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, '在库', 1)
                    """, (
                        record['bg'], record['sku'], record['sample_name'], record['sign_date'],
                        record['supplier'], record['brand'], record['notes'], record['location'],
                        record['expiry_date']
                    ))
                    inserted += 1
                    inserted_row = record.copy()
                    inserted_row['id'] = cursor.lastrowid
                    existing_map[key] = inserted_row
                    fallback_key = _sample_fallback_key(inserted_row)
                    if fallback_key:
                        fallback_map[fallback_key] = inserted_row

        conn.commit()
        preview_df = None
        if parsed_rows:
            import pandas as pd
            preview_df = pd.DataFrame(parsed_rows)

        return {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "preview_df": preview_df,
            "warnings": warnings,
            "message": f"已自动识别 {len(wb.sheetnames)} 个工作表并完成清洗预处理。",
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        wb.close()
        if owns_conn:
            conn.close()


def get_samples(search='', bg='', page=1, per_page=20):
    """查询样品列表"""
    conn = get_connection()
    conditions = []
    params = []
    if search:
        conditions.append("(sample_name LIKE ? OR sku LIKE ?)")
        params.extend([f'%{search}%'] * 2)
    if bg:
        conditions.append("bg = ?")
        params.append(bg)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    total = conn.execute(f"SELECT COUNT(*) FROM samples {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM samples {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


@audit("新增样品", "samples", capture_new_id=True)
def add_sample(data):
    """添加样品（自动计算一年有效期）"""
    from datetime import datetime as dt, timedelta as td
    sign_date = data.get('sign_date', '')
    expiry_date = ''
    if sign_date:
        try:
            s = sign_date.replace('.', '-')
            if len(s) == 7: s += '-01'
            d = dt.strptime(s[:10], '%Y-%m-%d')
            expiry_date = (d + td(days=365)).strftime('%Y-%m-%d')
        except Exception:
            expiry_date = ''
    conn = get_connection()
    conn.execute("""
        INSERT INTO samples (bg, sku, sample_name, sign_date, supplier, brand, notes, location, expiry_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data.get('bg', ''), data.get('sku', ''), data.get('sample_name', ''),
          sign_date, data.get('supplier', ''),
          data.get('brand', ''),
          data.get('notes', ''), data.get('location', ''), expiry_date))
    conn.commit()
    conn.close()
    return True, "样品添加成功"


@audit("更新样品", "samples")
def update_sample(s_id, data):
    """更新样品，并在签样日期变化时重算一年有效期。"""
    from datetime import datetime as dt, timedelta as td
    sign_date = data.get('sign_date', '')
    expiry_date = ''
    if sign_date:
        try:
            normalized = str(sign_date).replace('.', '-')
            if len(normalized) == 7:
                normalized += '-01'
            sign_dt = dt.strptime(normalized[:10], '%Y-%m-%d')
            expiry_date = (sign_dt + td(days=365)).strftime('%Y-%m-%d')
        except Exception:
            expiry_date = ''
    conn = get_connection()
    conn.execute("""
        UPDATE samples SET bg=?, sku=?, sample_name=?, sign_date=?, supplier=?,
        brand=?, notes=?, location=?, expiry_date=?
        WHERE id=?
    """, (data.get('bg', ''), data.get('sku', ''), data.get('sample_name', ''),
          data.get('sign_date', ''), data.get('supplier', ''), data.get('brand', ''),
          data.get('notes', ''), data.get('location', ''), expiry_date, s_id))
    conn.commit()
    conn.close()
    return True, "样品已更新"


@audit("删除样品", "samples")
@recoverable("samples")
def delete_sample(s_id):
    """删除样品"""
    conn = get_connection()
    conn.execute("DELETE FROM samples WHERE id=?", (s_id,))
    conn.commit()
    conn.close()
    return True, "样品已删除"


# ==================== 驻厂登记 ====================
@audit("新增驻厂登记", "factory_registrations")
def add_factory_registration(data):
    """新增一条驻厂登记。"""
    conn = get_connection()
    conn.execute("""
        INSERT INTO factory_registrations
        (register_date, factory_name, supplier, onsite_staff, trip_type,
         trip_days, po_no, sku, product_project, is_empty_run,
         is_recheck, is_delay, delay_days, return_reason, inspection_result,
         purpose, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get('register_date', ''),
        data.get('factory_name', ''),
        data.get('supplier', ''),
        data.get('onsite_staff', ''),
        data.get('trip_type', ''),
        int(data.get('trip_days', 0) or 0),
        data.get('po_no', ''),
        data.get('sku', ''),
        data.get('product_project', ''),
        data.get('is_empty_run', '否'),
        data.get('is_recheck', '否'),
        data.get('is_delay', '否'),
        int(data.get('delay_days', 0) or 0),
        data.get('return_reason', ''),
        data.get('inspection_result', '待定'),
        data.get('purpose', ''),
        data.get('notes', ''),
    ))
    conn.commit()
    conn.close()
    return True, "驻厂登记已添加"


@audit("更新驻厂登记", "factory_registrations")
def update_factory_registration(fr_id, data):
    """更新驻厂登记。"""
    conn = get_connection()
    conn.execute("""
        UPDATE factory_registrations SET
            register_date=?, factory_name=?, supplier=?, onsite_staff=?,
            trip_type=?, trip_days=?, po_no=?, sku=?, product_project=?,
            is_empty_run=?, is_recheck=?, is_delay=?, delay_days=?,
            return_reason=?, inspection_result=?, purpose=?, notes=?
        WHERE id=?
    """, (
        data.get('register_date', ''),
        data.get('factory_name', ''),
        data.get('supplier', ''),
        data.get('onsite_staff', ''),
        data.get('trip_type', ''),
        int(data.get('trip_days', 0) or 0),
        data.get('po_no', ''),
        data.get('sku', ''),
        data.get('product_project', ''),
        data.get('is_empty_run', '否'),
        data.get('is_recheck', '否'),
        data.get('is_delay', '否'),
        int(data.get('delay_days', 0) or 0),
        data.get('return_reason', ''),
        data.get('inspection_result', '待定'),
        data.get('purpose', ''),
        data.get('notes', ''),
        fr_id,
    ))
    conn.commit()
    conn.close()
    return True, "驻厂登记已更新"


@audit("删除驻厂登记", "factory_registrations")
@recoverable("factory_registrations")
def delete_factory_registration(fr_id):
    """删除驻厂登记（自动进入回收站，可在『误删找回』中还原）。"""
    conn = get_connection()
    conn.execute("DELETE FROM factory_registrations WHERE id=?", (fr_id,))
    conn.commit()
    conn.close()
    return True, "已移入回收站（可在数据记录的『误删找回』中找回）"


def get_factory_registrations(search='', factory='', staff='', inspection_result='',
                             trip_type='', date_from='', date_to='',
                             page=1, per_page=2000):
    """查询驻厂登记列表。"""
    conn = get_connection()
    conditions, params = [], []
    if search:
        conditions.append(
            "(factory_name LIKE ? OR onsite_staff LIKE ? OR po_no LIKE ? "
            "OR sku LIKE ? OR product_project LIKE ?)")
        params.extend([f'%{search}%'] * 5)
    if factory:
        conditions.append("factory_name = ?"); params.append(factory)
    if staff:
        conditions.append("onsite_staff LIKE ?"); params.append(f'%{staff}%')
    if inspection_result:
        conditions.append("inspection_result = ?"); params.append(inspection_result)
    if trip_type:
        conditions.append("trip_type = ?"); params.append(trip_type)
    if date_from:
        conditions.append("date(register_date) >= date(?)"); params.append(date_from)
    if date_to:
        conditions.append("date(register_date) <= date(?)"); params.append(date_to)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    total = conn.execute(
        f"SELECT COUNT(*) FROM factory_registrations {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM factory_registrations {where} "
        f"ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_factory_stats():
    """驻厂登记总览统计。"""
    conn = get_connection()
    total = conn.execute("SELECT COUNT(*) FROM factory_registrations").fetchone()[0]
    onsite = conn.execute(
        "SELECT COUNT(*) FROM factory_registrations WHERE inspection_result='Pass'").fetchone()[0]
    ended = conn.execute(
        "SELECT COUNT(*) FROM factory_registrations WHERE inspection_result='Fail'").fetchone()[0]
    empty = conn.execute(
        "SELECT COUNT(*) FROM factory_registrations WHERE is_empty_run='是'").fetchone()[0]
    recheck = conn.execute(
        "SELECT COUNT(*) FROM factory_registrations WHERE is_recheck='是'").fetchone()[0]
    delay = conn.execute(
        "SELECT COUNT(*) FROM factory_registrations WHERE is_delay='是'").fetchone()[0]
    conn.close()
    return {
        'total': total, 'onsite': onsite, 'ended': ended,
        'empty': empty, 'recheck': recheck, 'delay': delay,
    }


def get_factory_freq_by_staff():
    """按人员统计驻厂次数与累计天数。"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT onsite_staff, trip_days FROM factory_registrations").fetchall()
    conn.close()
    agg = {}
    for r in rows:
        staff_raw = r['onsite_staff'] or ''
        try:
            days = int(r['trip_days'] or 0)
        except Exception:
            days = 0
        for name in [s.strip() for s in
                      staff_raw.replace('\n', ',').split(',') if s.strip()]:
            if name not in agg:
                agg[name] = {'staff': name, 'count': 0, 'days': 0}
            agg[name]['count'] += 1
            agg[name]['days'] += days
    return sorted(agg.values(), key=lambda x: -x['count'])


def get_factory_freq_by_factory():
    """按工厂统计被驻厂次数。"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT factory_name, COUNT(*) AS c FROM factory_registrations "
        "GROUP BY factory_name ORDER BY c DESC").fetchall()
    conn.close()
    return [{'factory': r['factory_name'], 'count': r['c']} for r in rows]


def get_factory_freq_by_month():
    """按月统计驻厂次数。"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT substr(register_date,1,7) AS m, COUNT(*) AS c "
        "FROM factory_registrations GROUP BY m ORDER BY m").fetchall()
    conn.close()
    return [{'month': r['m'] or '未知', 'count': r['c']} for r in rows]


def get_factory_freq_by_type():
    """按出差类型统计。"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT trip_type, COUNT(*) AS c FROM factory_registrations "
        "GROUP BY trip_type ORDER BY c DESC").fetchall()
    conn.close()
    return [{'type': r['trip_type'] or '未填写', 'count': r['c']} for r in rows]

def _sample_sync_text(value):
    """同步时统一清洗文本值，避免 None/空白影响业务键匹配。"""
    if value is None:
        return ''
    return str(value).strip()


def _sample_placeholder_if_missing(value, placeholder='/'):
    """把空值/None/nan 等缺失内容统一替换成占位符。"""
    text = _sample_sync_text(value)
    if not text or text.lower() in {'none', 'nan', 'null'}:
        return placeholder
    return text


def _sample_has_meaningful_sku(value):
    text = _sample_sync_text(value)
    return bool(text) and text not in {'/'} and text.lower() not in {'none', 'nan', 'null'}


def _is_empty_sample_record(record):
    """判断样品记录是否为空记录（除 brand/bg 外所有关键字段均为空或占位符）"""
    meaningful_fields = ['sample_name', 'sku', 'supplier', 'sign_date', 'location', 'notes']
    for field in meaningful_fields:
        value = _sample_sync_text(record.get(field, ''))
        if value and value not in {'/', 'None', 'none', 'nan', 'null'}:
            return False
    return True

def _sample_sync_key(row):
    """为跨环境同步生成稳定业务键，不依赖自增 ID。"""
    parts = [
        _sample_sync_text(row.get('bg', '')).lower(),
        _sample_sync_text(row.get('sku', '')).lower(),
        _sample_sync_text(row.get('sample_name', '')).lower(),
        _sample_sync_text(row.get('sign_date', '')).replace('.', '-').lower(),
    ]
    if not any(parts[1:3]):
        return ''
    return '||'.join(parts)


def _sample_fallback_key(row):
    """SKU 缺失时使用的兜底业务键，避免同一条旧数据被重复插入。"""
    parts = [
        _sample_sync_text(row.get('bg', '')).lower(),
        _sample_sync_text(row.get('sample_name', '')).lower(),
        _sample_sync_text(row.get('sign_date', '')).replace('.', '-').lower(),
    ]
    if not any(parts[1:2]):
        return ''
    return '||'.join(parts)


def build_samples_sync_package():
    """导出样品同步包，供异地 Web 导入合并。"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT bg, sku, sample_name, sign_date, supplier, brand, notes,
               location, expiry_date, out_status, stock_qty
        FROM samples
        ORDER BY id ASC
    """).fetchall()
    conn.close()

    samples = []
    for row in rows:
        record = dict(row)
        for key, value in list(record.items()):
            if key == 'stock_qty':
                try:
                    record[key] = int(value or 0)
                except Exception:
                    record[key] = 0
            else:
                record[key] = _sample_sync_text(value)
        samples.append(record)

    payload = {
        "type": "samples_sync_package",
        "version": 1,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_db": os.path.basename(DB_PATH),
        "sample_count": len(samples),
        "samples": samples,
    }
    return payload


def merge_samples_sync_package(payload, db_conn=None):
    """
    合并样品同步包。
    保留当前库里的出库状态/库存，重点同步样品基础档案和新增记录。
    """
    if not isinstance(payload, dict):
        return False, "同步包格式不正确", {}
    if payload.get("type") != "samples_sync_package":
        return False, "不是有效的样品同步包", {}

    incoming_rows = payload.get("samples", [])
    if not isinstance(incoming_rows, list) or not incoming_rows:
        return False, "同步包里没有样品数据", {}

    owns_conn = db_conn is None
    conn = db_conn or get_connection()
    try:
        current_rows = conn.execute("""
            SELECT id, bg, sku, sample_name, sign_date, supplier, brand, notes,
                   location, expiry_date, out_status, stock_qty
            FROM samples
        """).fetchall()
        existing_map = {}
        fallback_map = {}
        for row in current_rows:
            row_dict = dict(row)
            key = _sample_sync_key(row_dict)
            if key and key not in existing_map:
                existing_map[key] = row_dict
            fallback_key = _sample_fallback_key(row_dict)
            if fallback_key and fallback_key not in fallback_map:
                fallback_map[fallback_key] = row_dict

        inserted = 0
        updated = 0
        skipped = 0
        duplicate_keys = 0
        seen_incoming = set()
        update_fields = ['bg', 'sku', 'sample_name', 'sign_date', 'supplier', 'brand', 'notes', 'location', 'expiry_date']

        for raw_row in incoming_rows:
            if not isinstance(raw_row, dict):
                skipped += 1
                continue

            incoming = {}
            for field in update_fields + ['out_status', 'stock_qty']:
                value = raw_row.get(field, '')
                if field == 'stock_qty':
                    try:
                        incoming[field] = int(value or 0)
                    except Exception:
                        incoming[field] = 0
                else:
                    incoming[field] = _sample_sync_text(value)

            key = _sample_sync_key(incoming)
            if not key:
                skipped += 1
                continue
            if key in seen_incoming:
                duplicate_keys += 1
                continue
            seen_incoming.add(key)

            existing = existing_map.get(key)
            if not existing:
                fallback_key = _sample_fallback_key(incoming)
                fallback_existing = fallback_map.get(fallback_key) if fallback_key else None
                if fallback_existing and (
                    not _sample_has_meaningful_sku(incoming.get('sku'))
                    or not _sample_has_meaningful_sku(fallback_existing.get('sku'))
                ):
                    existing = fallback_existing
            if existing:
                changed = False
                update_values = []
                for field in update_fields:
                    new_value = incoming[field]
                    old_value = _sample_sync_text(existing.get(field, ''))
                    update_values.append(new_value)
                    if new_value != old_value:
                        changed = True
                if changed:
                    conn.execute("""
                        UPDATE samples
                        SET bg=?, sku=?, sample_name=?, sign_date=?, supplier=?,
                            brand=?, notes=?, location=?, expiry_date=?
                        WHERE id=?
                    """, update_values + [existing['id']])
                    updated += 1
                    updated_row = dict(existing)
                    updated_row.update(incoming)
                    existing_map[key] = updated_row
                    fallback_key = _sample_fallback_key(updated_row)
                    if fallback_key:
                        fallback_map[fallback_key] = updated_row
                else:
                    skipped += 1
                continue

            cursor = conn.execute("""
                INSERT INTO samples (
                    bg, sku, sample_name, sign_date, supplier, brand, notes,
                    location, expiry_date, out_status, stock_qty
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                incoming['bg'], incoming['sku'], incoming['sample_name'], incoming['sign_date'],
                incoming['supplier'], incoming['brand'], incoming['notes'], incoming['location'],
                incoming['expiry_date'], incoming['out_status'] or '在库', incoming['stock_qty'] or 1
            ))
            inserted += 1
            inserted_row = incoming.copy()
            inserted_row['id'] = cursor.lastrowid
            existing_map[key] = inserted_row
            fallback_key = _sample_fallback_key(inserted_row)
            if fallback_key:
                fallback_map[fallback_key] = inserted_row

        if owns_conn:
            conn.commit()
        summary = {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "duplicate_keys": duplicate_keys,
            "incoming": len(incoming_rows),
        }
        return True, f"同步完成：新增 {inserted} 条，更新 {updated} 条，跳过 {skipped} 条", summary
    except Exception as e:
        if owns_conn:
            conn.rollback()
        return False, f"同步失败：{e}", {}
    finally:
        if owns_conn:
            conn.close()


# ==================== 样品出库 ====================

@audit("样品出库", "sample_outbound", capture_new_id=True)
def sample_outbound(sample_id, qty=1, out_date='', borrower='', department='', reason='', notes=''):
    """样品出库登记（写记录 + 扣库存 两步原子操作）"""
    conn = get_connection()
    try:
        # 检查样品状态：仅支持「在库」样品
        current = conn.execute(
            "SELECT out_status, stock_qty FROM samples WHERE id=?", (sample_id,)
        ).fetchone()
        if not current:
            conn.close()
            return False, "样品不存在"

        # 在库检查（out_status 为 已出库 且 stock_qty <= 0 则不可出库）
        if current['out_status'] == '已出库' and (current['stock_qty'] or 0) <= 0:
            conn.close()
            return False, "该样品已出库，请先归还再操作"

        stock_qty = current['stock_qty'] or 0

        # 1/2 写入出库记录
        conn.execute("""
            INSERT INTO sample_outbound (sample_id, qty, out_date, borrower, department, reason, notes, is_returned)
            VALUES (?, ?, ?, ?, ?, ?, ?, 0)
        """, (sample_id, qty, out_date, borrower, department, reason, notes))

        # 2/2 扣减样品库存
        new_qty = max(0, stock_qty - int(qty))
        new_status = '已出库' if new_qty <= 0 else '在库'
        conn.execute(
            "UPDATE samples SET out_status=?, stock_qty=? WHERE id=?",
            (new_status, new_qty, sample_id)
        )

        conn.commit()
        conn.close()
        return True, f"出库成功（剩余库存：{new_qty}）"
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, f"出库异常：{e}"


@audit("样品归还", "samples")
def sample_return(sample_id):
    """样品归还（改回在库状态 + 恢复库存）"""
    conn = get_connection()
    try:
        # 找到该样品最近一条未归还的出库记录，获取归还数量
        latest_out = conn.execute(
            "SELECT id, qty FROM sample_outbound WHERE sample_id=? AND (is_returned IS NULL OR is_returned=0) ORDER BY id DESC LIMIT 1",
            (sample_id,)
        ).fetchone()

        return_qty = latest_out['qty'] if latest_out else 1

        # 恢复样品状态和库存
        current = conn.execute("SELECT stock_qty FROM samples WHERE id=?", (sample_id,)).fetchone()
        if current:
            new_qty = (current['stock_qty'] or 0) + return_qty
            conn.execute(
                "UPDATE samples SET out_status='在库', stock_qty=? WHERE id=?",
                (new_qty, sample_id)
            )

        # 标记出库记录为已归还
        if latest_out:
            from datetime import date
            conn.execute(
                "UPDATE sample_outbound SET is_returned=1, return_date=? WHERE id=?",
                (str(date.today()), latest_out['id'])
            )

        conn.commit()
        conn.close()
        return True, "已归还"
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, f"归还异常：{e}"


@audit("按记录归还样品", "sample_outbound")
def sample_return_by_record(outbound_id):
    """根据出库记录 ID 一键归还（更新出库记录 + 恢复主表库存，原子操作）"""
    conn = get_connection()
    try:
        # 获取出库记录
        rec = conn.execute(
            "SELECT id, sample_id, qty, is_returned FROM sample_outbound WHERE id=?",
            (outbound_id,)
        ).fetchone()
        if not rec:
            conn.close()
            return False, "出库记录不存在"
        if rec['is_returned']:
            conn.close()
            return False, "该记录已归还，不可重复操作"

        sample_id = rec['sample_id']
        return_qty = rec['qty'] or 1

        # 获取样品当前库存
        current = conn.execute("SELECT stock_qty FROM samples WHERE id=?", (sample_id,)).fetchone()
        if not current:
            conn.close()
            return False, "样品不存在"

        new_qty = (current['stock_qty'] or 0) + return_qty

        # 1/2 更新出库记录为已归还
        from datetime import date
        conn.execute(
            "UPDATE sample_outbound SET is_returned=1, return_date=? WHERE id=?",
            (str(date.today()), outbound_id)
        )

        # 2/2 恢复样品库存和状态
        conn.execute(
            "UPDATE samples SET out_status='在库', stock_qty=? WHERE id=?",
            (new_qty, sample_id)
        )

        conn.commit()
        conn.close()
        return True, f"归还成功（库存 +{return_qty}，当前：{new_qty}）"
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, f"归还异常：{e}"


def get_outbound_records(sample_id=None, limit=100):
    """查询出库记录"""
    conn = get_connection()
    if sample_id:
        rows = conn.execute("""
            SELECT o.*, s.sample_name, s.sku, s.bg
            FROM sample_outbound o JOIN samples s ON o.sample_id = s.id
            WHERE o.sample_id = ? ORDER BY o.id DESC LIMIT ?
        """, (sample_id, limit)).fetchall()
    else:
        rows = conn.execute("""
            SELECT o.*, s.sample_name, s.sku, s.bg
            FROM sample_outbound o JOIN samples s ON o.sample_id = s.id
            ORDER BY o.id DESC LIMIT ?
        """, (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 变更管理 ====================

def _normalize_change_text(value):
    """统一清洗变更记录字段，避免 nan / None 混入业务键"""
    if value is None:
        return ''
    text = str(value).strip()
    return '' if text.lower() == 'nan' else text


def _extract_supplier_and_reason(change_reason, supplier=''):
    """从变更原因中提取供应商，并返回清理后的原因文本"""
    import re

    cleaned_reason = _normalize_change_text(change_reason)[:500]
    cleaned_supplier = _normalize_change_text(supplier)

    if cleaned_reason and not cleaned_supplier:
        patterns = [
            (r'[；;]?\s*变更供应商[：:]\s*([^；;]+?)(?:[；;\s]+|$)',
             lambda m, raw: (m.group(1).strip().rstrip('；; '), raw[:m.start()] + raw[m.end():])),
            (r'^变更供应商[：:]\s*(.+?)\s{2,}',
             lambda m, raw: (m.group(1).strip(), raw[m.end():])),
            (r'^供应商[：:]\s*(.+?)\s{2,}',
             lambda m, raw: (m.group(1).strip(), raw[m.end():])),
            (r'[；;]?\s*供应商[：:]\s*([^；;]+?)(?:[；;\s]+|$)',
             lambda m, raw: (m.group(1).strip().rstrip('；; '), raw[:m.start()] + raw[m.end():])),
        ]
        for pattern, handler in patterns:
            match = re.search(pattern, cleaned_reason)
            if match:
                cleaned_supplier, cleaned_reason = handler(match, cleaned_reason)
                break

        cleaned_reason = re.sub(r'[；;]{2,}', '；', cleaned_reason)
        cleaned_reason = cleaned_reason.strip('；; ')

    return cleaned_supplier, cleaned_reason


def _change_identity(record):
    """变更记录的业务唯一键，用于导入覆盖而不是重复追加"""
    return (
        _normalize_change_text(record.get('bu', '')),
        _normalize_change_text(record.get('brand', '')),
        _normalize_change_text(record.get('sku', '')),
        _normalize_change_text(record.get('change_reason', '')),
        _normalize_change_text(record.get('change_date', '')),
    )


def _infer_change_overall_status(record, fallback_status='待确认'):
    """根据导入的确认信息推导整体确认状态"""
    explicit_status = _normalize_change_text(record.get('overall_status', fallback_status))
    sku_confirm_status = _normalize_change_text(record.get('sku_confirm_status', '{}'))
    confirm_person = _normalize_change_text(record.get('confirm_person', ''))
    confirm_date = _normalize_change_text(record.get('confirm_date', ''))

    if sku_confirm_status and sku_confirm_status != '{}':
        return explicit_status or '待确认'

    if confirm_person or confirm_date:
        return '全部确认'

    return explicit_status or '待确认'


def _extract_change_confirmation_summary(record: dict) -> tuple[str, str]:
    """从 SKU 明细确认 JSON 中提取一条可展示的确认人/确认日期摘要。"""
    confirm_person = _normalize_change_text(record.get('confirm_person', ''))
    confirm_date = _normalize_change_text(record.get('confirm_date', ''))
    if confirm_person or confirm_date:
        return confirm_person, confirm_date

    raw_status = record.get('sku_confirm_status', '{}')
    try:
        status_data = json.loads(raw_status) if isinstance(raw_status, str) else (raw_status or {})
    except Exception:
        return '', ''

    latest_person = ''
    latest_date = ''

    def _pick(person: str, dt: str) -> None:
        nonlocal latest_person, latest_date
        person = _normalize_change_text(person)
        dt = _normalize_change_text(dt)
        if not person and not dt:
            return
        if dt >= latest_date:
            latest_person = person
            latest_date = dt

    for value in (status_data or {}).values():
        if not isinstance(value, dict):
            continue

        # 兼容旧格式：{"SKU": {"status": true, "confirmer": "...", "date": "..."}}
        if 'status' in value:
            if value.get('status'):
                _pick(value.get('confirmer', ''), value.get('date', ''))
            continue

        # 新格式：{"SKU": {"供应商A": {"status": true, ...}, ...}}
        for supplier_value in value.values():
            if isinstance(supplier_value, dict) and supplier_value.get('status'):
                _pick(supplier_value.get('confirmer', ''), supplier_value.get('date', ''))

    return latest_person, latest_date


def _upsert_change_records(conn, records):
    """按业务键同步变更记录：同键更新，且自动清理重复项"""
    inserted = 0
    updated = 0
    deduped = 0

    for raw_record in records:
        record = {
            'bu': _normalize_change_text(raw_record.get('bu', '')),
            'brand': _normalize_change_text(raw_record.get('brand', '')),
            'sku': _normalize_change_text(raw_record.get('sku', '')),
            'change_reason': _normalize_change_text(raw_record.get('change_reason', '')),
            'supplier': _normalize_change_text(raw_record.get('supplier', '')),
            'change_date': _normalize_change_text(raw_record.get('change_date', ''))[:19],
            'notify_person': _normalize_change_text(raw_record.get('notify_person', '')),
            'confirm_date': _normalize_change_text(raw_record.get('confirm_date', ''))[:19],
            'confirm_person': _normalize_change_text(raw_record.get('confirm_person', '')),
        }

        if not record['brand']:
            continue

        match_values = _change_identity(record)
        existing_rows = conn.execute("""
            SELECT id, attachments, rd_team, sku_confirm_status, overall_status, notify_person
            FROM change_records
            WHERE bu=? AND brand=? AND sku=? AND change_reason=? AND change_date=?
            ORDER BY id ASC
        """, match_values).fetchall()

        if existing_rows:
            keep_row = existing_rows[0]
            attachments = _normalize_change_text(raw_record.get('attachments')) if 'attachments' in raw_record else keep_row['attachments']
            rd_team = _normalize_change_text(raw_record.get('rd_team')) if 'rd_team' in raw_record else keep_row['rd_team']
            notify_person = (
                _normalize_change_text(raw_record.get('notify_person'))
                if 'notify_person' in raw_record else keep_row['notify_person']
            )
            sku_confirm_status = (
                _normalize_change_text(raw_record.get('sku_confirm_status'))
                if 'sku_confirm_status' in raw_record else keep_row['sku_confirm_status']
            ) or '{}'
            overall_status = _infer_change_overall_status({
                **raw_record,
                'confirm_person': record['confirm_person'],
                'confirm_date': record['confirm_date'],
                'sku_confirm_status': sku_confirm_status,
                'overall_status': (
                    _normalize_change_text(raw_record.get('overall_status'))
                    if 'overall_status' in raw_record else keep_row['overall_status']
                ) or '待确认',
            })

            conn.execute("""
                UPDATE change_records
                SET bu=?, brand=?, sku=?, change_reason=?, supplier=?, attachments=?,
                    change_date=?, notify_person=?, confirm_date=?, confirm_person=?, rd_team=?,
                    sku_confirm_status=?, overall_status=?
                WHERE id=?
            """, (
                record['bu'], record['brand'], record['sku'], record['change_reason'],
                record['supplier'], attachments, record['change_date'], notify_person,
                record['confirm_date'], record['confirm_person'], rd_team, sku_confirm_status, overall_status,
                keep_row['id']
            ))
            updated += 1

            duplicate_ids = [row['id'] for row in existing_rows[1:]]
            if duplicate_ids:
                placeholders = ",".join("?" for _ in duplicate_ids)
                conn.execute(f"DELETE FROM change_records WHERE id IN ({placeholders})", duplicate_ids)
                deduped += len(duplicate_ids)
        else:
            conn.execute("""
                INSERT INTO change_records (
                    bu, brand, sku, change_reason, supplier, attachments, change_date, notify_person,
                    confirm_date, confirm_person, rd_team, sku_confirm_status, overall_status
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                record['bu'], record['brand'], record['sku'], record['change_reason'],
                record['supplier'], _normalize_change_text(raw_record.get('attachments', '')),
                record['change_date'], record['notify_person'], record['confirm_date'], record['confirm_person'],
                _normalize_change_text(raw_record.get('rd_team', '')),
                _normalize_change_text(raw_record.get('sku_confirm_status', '{}')) or '{}',
                _infer_change_overall_status(raw_record),
            ))
            inserted += 1

    conn.commit()
    return inserted, updated, deduped


def import_changes_from_excel(filepath=None):
    """从产品变更汇总表Excel导入变更记录（自动提取供应商）"""
    import pandas as pd
    if filepath is None:
        filepath = os.path.join(PROJECT_ROOT,
                                'SainStore实验室文件', '产品变更汇总表 .xlsx')
    if not os.path.exists(filepath):
        return 0, "文件不存在"

    df = pd.read_excel(filepath, engine='openpyxl', header=1)
    conn = get_connection()
    records = []
    for _, row in df.iterrows():
        bu = _normalize_change_text(row.get('BU', '')) if pd.notna(row.get('BU')) else ''
        brand = _normalize_change_text(row.get('品牌', '')) if pd.notna(row.get('品牌')) else ''
        if not brand or brand == 'nan':
            continue

        supplier, raw_reason = _extract_supplier_and_reason(
            row.get('变更原因及内容', ''),
            ''
        )
        records.append({
            'bu': bu,
            'brand': brand,
            'sku': _normalize_change_text(row.get('SKU', '')) if pd.notna(row.get('SKU')) else '',
            'change_reason': raw_reason,
            'supplier': supplier,
            'change_date': _normalize_change_text(row.get('变更时间', '')) if pd.notna(row.get('变更时间')) else '',
            'notify_person': _normalize_change_text(row.get('待确认人', '')) if pd.notna(row.get('待确认人')) else '',
            'confirm_date': _normalize_change_text(row.get('确认时间', '')) if pd.notna(row.get('确认时间')) else '',
            'confirm_person': _normalize_change_text(row.get('确认人', '')) if pd.notna(row.get('确认人')) else '',
        })

    inserted, updated, deduped = _upsert_change_records(conn, records)
    conn.close()
    affected = inserted + updated
    if affected == 0:
        return 0, "没有识别到可导入的变更记录"
    return affected, f"导入完成：新增 {inserted} 条，覆盖更新 {updated} 条，清理重复 {deduped} 条"


def import_changes_dataframe(import_df, db_conn=None):
    """从模板 DataFrame 导入变更记录，按业务键覆盖更新"""
    records = []
    for _, row in import_df.iterrows():
        supplier, change_reason = _extract_supplier_and_reason(
            row.get('change_reason', ''),
            row.get('supplier', '')
        )
        records.append({
            'bu': row.get('bu', ''),
            'brand': row.get('brand', ''),
            'sku': row.get('sku', ''),
            'supplier': supplier,
            'change_reason': change_reason,
            'change_date': row.get('change_date', ''),
            'notify_person': row.get('notify_person', ''),
            'confirm_person': row.get('confirm_person', ''),
            'confirm_date': row.get('confirm_date', ''),
        })

    own_conn = db_conn is None
    conn = db_conn or get_connection()
    inserted, updated, deduped = _upsert_change_records(conn, records)
    if own_conn:
        conn.close()

    affected = inserted + updated
    if affected == 0:
        return 0, "没有识别到可导入的变更记录"
    return affected, f"模板导入完成：新增 {inserted} 条，覆盖更新 {updated} 条，清理重复 {deduped} 条"


def get_changes(search='', bu='', brand='', page=1, per_page=20, search_sku='', search_content='', search_supplier=''):
    """查询变更记录（支持多字段搜索）"""
    conn = get_connection()
    conditions = []
    params = []

    # 拼接所有搜索条件
    search_parts = []
    if search:
        search_parts.append(search)
    if search_sku:
        conditions.append("sku LIKE ?")
        params.append(f'%{search_sku}%')
    if search_content:
        conditions.append("change_reason LIKE ?")
        params.append(f'%{search_content}%')
    if search_supplier:
        conditions.append("supplier LIKE ?")
        params.append(f'%{search_supplier}%')
    if search:
        conditions.append("(brand LIKE ? OR change_reason LIKE ? OR supplier LIKE ?)")
        params.extend([f'%{search}%'] * 3)

    if bu:
        conditions.append("bu = ?")
        params.append(bu)
    if brand:
        conditions.append("brand = ?")
        params.append(brand)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    total = conn.execute(f"SELECT COUNT(*) FROM change_records {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM change_records {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    result = []
    for row in rows:
        item = dict(row)
        derived_person, derived_date = _extract_change_confirmation_summary(item)
        if derived_person and not _normalize_change_text(item.get('confirm_person', '')):
            item['confirm_person'] = derived_person
        if derived_date and not _normalize_change_text(item.get('confirm_date', '')):
            item['confirm_date'] = derived_date
        result.append(item)
    return result, total


@audit("新增变更", "change_records", capture_new_id=True)
def add_change(data):
    """添加变更记录（自动从变更原因提取供应商），返回 (ok, msg, new_id)"""
    conn = get_connection()
    supplier, change_reason = _extract_supplier_and_reason(
        data.get('change_reason', ''),
        data.get('supplier', '')
    )

    cur = conn.execute("""
        INSERT INTO change_records (bu, brand, sku, change_reason, supplier, attachments, change_date, notify_person, confirm_date, confirm_person, rd_team, sku_confirm_status, overall_status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data.get('bu', ''), data.get('brand', ''), data.get('sku', ''),
          change_reason, supplier, data.get('attachments', ''),
          data.get('change_date', ''), data.get('notify_person', ''),
          data.get('confirm_date', ''),
          data.get('confirm_person', '') or '', data.get('rd_team', ''),
          data.get('sku_confirm_status', '{}'),
          data.get('overall_status', '待确认')))
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return True, "变更记录添加成功", new_id


def update_change_attachments(c_id, attachments):
    """仅更新变更记录的附件字段（单字段 UPDATE，不会清空其他列）"""
    conn = get_connection()
    conn.execute(
        "UPDATE change_records SET attachments=? WHERE id=?",
        (attachments, c_id)
    )
    conn.commit()
    conn.close()
    return True, "附件已更新"


@audit("更新变更", "change_records")
def update_change(c_id, data):
    """更新变更记录（含 SKU 确认状态）"""
    conn = get_connection()
    conn.execute("""
        UPDATE change_records SET bu=?, brand=?, sku=?, change_reason=?, supplier=?,
        attachments=?, change_date=?, notify_person=?, confirm_person=?, confirm_date=?, rd_team=?,
        sku_confirm_status=?, overall_status=?
        WHERE id=?
    """, (data.get('bu', ''), data.get('brand', ''), data.get('sku', ''),
          data.get('change_reason', ''), data.get('supplier', ''), data.get('attachments', ''),
          data.get('change_date', ''), data.get('notify_person', ''), data.get('confirm_person', ''),
          data.get('confirm_date', ''), data.get('rd_team', ''),
          data.get('sku_confirm_status', '{}'),
          data.get('overall_status', '待确认'),
          c_id))
    conn.commit()
    conn.close()
    return True, "变更记录已更新"
    conn.commit()
    conn.close()
    return True, "变更记录已更新"


@audit("删除变更", "change_records")
@recoverable("change_records")
def delete_change(c_id):
    """删除变更记录"""
    conn = get_connection()
    conn.execute("DELETE FROM change_records WHERE id=?", (c_id,))
    conn.commit()
    conn.close()
    return True, "变更记录已删除"


@audit("批量删除变更", "change_records")
@recoverable("change_records")
def bulk_delete_changes(change_ids):
    """批量删除变更记录"""
    if not change_ids:
        return False, "请先选择要删除的记录"

    normalized_ids = []
    for change_id in change_ids:
        try:
            normalized_ids.append(int(change_id))
        except (TypeError, ValueError):
            continue

    if not normalized_ids:
        return False, "没有可删除的有效记录"

    conn = get_connection()
    placeholders = ",".join("?" for _ in normalized_ids)
    conn.execute(f"DELETE FROM change_records WHERE id IN ({placeholders})", normalized_ids)
    conn.commit()
    conn.close()
    return True, f"已删除 {len(normalized_ids)} 条变更记录"


def cleanup_duplicate_changes():
    """按业务键清理重复的变更记录，保留最早的一条"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT id, bu, brand, sku, change_reason, change_date
        FROM change_records
        ORDER BY bu, brand, sku, change_reason, change_date, id
    """).fetchall()

    seen_keys = set()
    duplicate_ids = []
    for row in rows:
        key = (
            row['bu'] or '',
            row['brand'] or '',
            row['sku'] or '',
            row['change_reason'] or '',
            row['change_date'] or '',
        )
        if key in seen_keys:
            duplicate_ids.append(row['id'])
        else:
            seen_keys.add(key)

    if duplicate_ids:
        placeholders = ",".join("?" for _ in duplicate_ids)
        conn.execute(f"DELETE FROM change_records WHERE id IN ({placeholders})", duplicate_ids)
        conn.commit()

    conn.close()
    return True, f"已清理 {len(duplicate_ids)} 条重复记录"


def backfill_change_confirmation_status():
    """把已有确认人/确认日期的旧数据回填为已确认状态"""
    conn = get_connection()
    cursor = conn.execute("""
        UPDATE change_records
        SET overall_status = '全部确认'
        WHERE COALESCE(overall_status, '') = '待确认'
          AND (COALESCE(confirm_person, '') != '' OR COALESCE(confirm_date, '') != '')
          AND COALESCE(sku_confirm_status, '{}') = '{}'
    """)
    conn.commit()
    updated = cursor.rowcount if cursor.rowcount is not None else 0
    conn.close()
    return True, f"已回填 {updated} 条变更记录的确认状态"


# ==================== 检验报告 ====================

@audit("上传检验报告", "inspection_reports", capture_new_id=True)
def add_inspection_report(data):
    """添加检验报告"""
    conn = get_connection()
    conn.execute("""
        INSERT INTO inspection_reports (report_type, inspector, product_name, bg, bu, brand, sku,
                                        filename, file_path, image_paths, nas_report_path, nas_picture_path,
                                        nas_staging_path, supplier, status, reviewer, inspection_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data.get('report_type', ''), data.get('inspector', ''),
          data.get('product_name', ''), data.get('bg', ''), data.get('bu', ''),
          data.get('brand', ''), data.get('sku', ''), data.get('filename', ''),
          data.get('file_path', ''), data.get('image_paths', ''),
          data.get('nas_report_path', ''), data.get('nas_picture_path', ''),
          data.get('nas_staging_path', ''),
          data.get('supplier', ''),
          data.get('status', '待审核'), data.get('reviewer', 'teddy.li黎晓锋'),
          str(data.get('inspection_date', ''))))
    conn.commit()
    conn.close()
    return True, "报告提交成功"


def repair_inspection_report_filenames():
    """修复检验报告历史记录中的乱码文件名。"""
    conn = get_connection()
    rows = conn.execute("SELECT id, filename FROM inspection_reports").fetchall()

    scanned = 0
    updated = 0

    for row in rows:
        rpt_id = row["id"] if isinstance(row, sqlite3.Row) else row[0]
        filename = row["filename"] if isinstance(row, sqlite3.Row) else row[1]
        scanned += 1

        if not filename:
            continue

        repaired = _repair_filename_mojibake(filename)
        if repaired != filename and _filename_readability_score(repaired) > _filename_readability_score(filename):
            conn.execute(
                "UPDATE inspection_reports SET filename = ? WHERE id = ?",
                (repaired, rpt_id),
            )
            updated += 1

    conn.commit()
    conn.close()
    return True, f"已扫描 {scanned} 条报告，修复 {updated} 条乱码文件名"


def update_inspection_report_filename(report_id, filename):
    """手动更新检验报告文件名。"""
    filename = (filename or "").strip()
    if not filename:
        return False, "文件名不能为空"

    conn = get_connection()
    cursor = conn.execute(
        "UPDATE inspection_reports SET filename = ? WHERE id = ?",
        (filename, report_id),
    )
    conn.commit()
    updated = cursor.rowcount if cursor.rowcount is not None else 0
    conn.close()

    if updated <= 0:
        return False, "未找到对应报告记录"
    return True, "文件名已更新"


def get_inspection_reports(status=None, page=1, per_page=20):
    """查询检验报告"""
    conn = get_connection()
    conditions = []
    params = []
    if status:
        conditions.append("status = ?")
        params.append(status)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    total = conn.execute(f"SELECT COUNT(*) FROM inspection_reports {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM inspection_reports {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_unified_reports(status=None, page=1, per_page=20):
    """2026-07-16: 统一查询上传报告 + 在线报告（草稿不进列表）。
    用 UNION ALL 拼接两表，在线报告的 JSON 字段通过 json_extract 抽出对齐。
    返回每条 dict 含 `source` 字段（"upload"/"online"）用于前端路由。"""
    conn = get_connection()
    # 上传报告：可选 status 筛选
    upload_where = f"WHERE status = ?" if status else ""
    # 在线报告：草稿不进列表 + 可选 status 筛选
    online_where = "WHERE status != '草稿'"
    if status:
        online_where += " AND status = ?"

    # 上传报告子查询
    upload_sql = f"""SELECT
        id,
        'upload' AS source,
        '' AS report_no,
        report_type,
        inspector,
        product_name,
        brand,
        sku,
        -- 文件名：优先使用原始上传文件名（与实际文件完全一致），不系统构造
        COALESCE(filename, product_name || '验货报告') AS filename,
        '' AS po,
        file_path,
        bg,
        bu,
        supplier,
        reject_reason,
        inspection_date,
        nas_staging_path,
        image_paths,
        status,
        reviewer,
        review_comment,
        created_at
    FROM inspection_reports
    {upload_where}"""

    # 在线报告子查询：字段映射到同名列 + 从 data_json 拆出品牌/SKU/PO/类型
    # ⚠️ 列顺序必须与 upload_sql 完全一致（UNION ALL 要求两侧列数+顺序相同）
    online_sql = f"""SELECT
        id,
        'online' AS source,
        report_no,
        COALESCE(json_extract(data_json, '$.basic.type'), '在线 QC') AS report_type,
        COALESCE(json_extract(data_json, '$.basic.inspector'), created_by, '') AS inspector,
        COALESCE(json_extract(data_json, '$.basic.product'), product_name, '') AS product_name,
        COALESCE(json_extract(data_json, '$.basic.brand'), '') AS brand,
        COALESCE(json_extract(data_json, '$.basic.sku'), '') AS sku,
        -- [col9] filename：与上传报告一致的位置
        COALESCE(
            json_extract(data_json, '$.basic.report_title'),
            json_extract(data_json, '$.basic.title'),
            ('在线报告_' || report_no)
        ) AS filename,
        -- [col10] po：与上传报告一致的位置
        COALESCE(json_extract(data_json, '$.basic.po'), '') AS po,
        pdf_path AS file_path,
        '' AS bg,
        '' AS bu,
        COALESCE(supplier, '') AS supplier,
        '' AS reject_reason,
        COALESCE(json_extract(data_json, '$.basic.date'), '') AS inspection_date,
        nas_staging_path,
        '' AS image_paths,
        status,
        COALESCE(reviewer, '') AS reviewer,
        COALESCE(review_comment, '') AS review_comment,
        COALESCE(submitted_at, created_at, '') AS created_at
    FROM online_reports
    {online_where}"""

    # 合并参数（上传表的 ? 个数 = 在线表的 ? 个数 = 1 if status else 0）
    all_params = ([status] if status else []) * 2

    # 总数
    count_sql = f"SELECT COUNT(*) FROM ({upload_sql} UNION ALL {online_sql})"
    total = conn.execute(count_sql, all_params).fetchone()[0]

    # 分页查询（SQLite 对 LIMIT/OFFSET 不支持占位符？其实支持，但这里用 format 更安全）
    unified_sql = f"{upload_sql} UNION ALL {online_sql} ORDER BY created_at DESC LIMIT ? OFFSET ?"
    offset = (page - 1) * per_page
    rows = conn.execute(unified_sql, all_params + [per_page, offset]).fetchall()
    result = [dict(r) for r in rows]
    # 后处理：在线报告文件名修正（从data_json提取真实标题，确保与归档命名一致）
    import json as _json
    online_ids = [r['id'] for r in result if r.get('source') == 'online' and r.get('id')]
    if online_ids:
        placeholders = ','.join('?' * len(online_ids))
        raw_rows = conn.execute(
            f"SELECT id, data_json FROM online_reports WHERE id IN ({placeholders})",
            online_ids
        ).fetchall()
        title_map = {}
        for rid, dj in raw_rows:
            try:
                d = _json.loads(dj) if isinstance(dj, str) else (dj or {})
            except:
                d = {}
            basic = d.get('basic', {}) or {}
            ti = d.get('titleInfo', {}) or {}
            title_map[rid] = (
                basic.get('report_title')
                or basic.get('title')
                or ti.get('title')
                or ''
            )
        for r in result:
            if r['id'] in title_map and title_map[r['id']]:
                r['filename'] = title_map[r['id']]
    conn.close()
    return result, total


@audit("更新报告状态", "inspection_reports")
def update_report_status(report_id, status, reject_reason=""):
    """更新报告审核状态"""
    conn = get_connection()
    if reject_reason:
        conn.execute(
            "UPDATE inspection_reports SET status = ?, reject_reason = ? WHERE id = ?",
            (status, reject_reason, report_id)
        )
    else:
        conn.execute(
            "UPDATE inspection_reports SET status = ? WHERE id = ?",
            (status, report_id)
        )
    conn.commit()
    conn.close()
    return True, "状态已更新"


def update_report_images(report_id, image_paths, nas_picture_path):
    """更新报告的检验图片路径（重新上传图片后调用）。"""
    conn = get_connection()
    conn.execute(
        "UPDATE inspection_reports SET image_paths=?, nas_picture_path=? WHERE id=?",
        (image_paths, nas_picture_path, report_id),
    )
    conn.commit()
    conn.close()
    return True


def update_report_file(report_id, filename, file_path, nas_staging_path):
    """更新报告的文件路径（重新上传报告文件后调用）。"""
    conn = get_connection()
    conn.execute(
        "UPDATE inspection_reports SET filename=?, file_path=?, nas_staging_path=? WHERE id=?",
        (filename, file_path, nas_staging_path, report_id),
    )
    conn.commit()
    conn.close()
    return True


def update_report_info(report_id, **kwargs):
    """更新报告的基本信息字段（产品名/品牌/SKU/BG/BU/供应商/检验日期等）。
    只接受表中已存在的字段，忽略未知字段。
    """
    _allowed = {
        'product_name', 'brand', 'sku', 'bg', 'bu', 'supplier',
        'inspection_date', 'report_type', 'reviewer',
    }
    updates = {k: v for k, v in kwargs.items() if k in _allowed and v is not None}
    if not updates:
        return False, "无有效更新字段"
    sets = ", ".join(f"{k} = ?" for k in updates)
    vals = list(updates.values()) + [report_id]
    conn = get_connection()
    conn.execute(f"UPDATE inspection_reports SET {sets} WHERE id = ?", vals)
    conn.commit()
    conn.close()
    return True, f"已更新 {len(updates)} 个字段"


def update_report_review_comment(report_id, comment):
    """保存审核员审批意见。"""
    conn = get_connection()
    conn.execute(
        "UPDATE inspection_reports SET review_comment = ? WHERE id = ?",
        (comment, report_id),
    )
    conn.commit()
    conn.close()
    return True


@audit("删除报告", "inspection_reports")
@recoverable("inspection_reports")
def delete_inspection_report(report_id):
    """删除检验报告记录，并级联清理 NAS / 本地存储的对应文件，避免存档混乱。
    返回 (success: bool, msg: str)。
    """
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id, filename, file_path, image_paths, nas_report_path, "
            "nas_picture_path, nas_staging_path FROM inspection_reports WHERE id = ?",
            (report_id,)
        ).fetchone()
        if not row:
            return False, "报告不存在或已被删除"

        filename = row["filename"] or ""
        file_path = row["file_path"] or ""
        image_paths = row["image_paths"] or ""
        nas_report_path = row["nas_report_path"] or ""
        nas_picture_path = row["nas_picture_path"] or ""
        nas_staging_path = row["nas_staging_path"] or ""
    finally:
        conn.close()

    cleaned = []

    # ── 1. 清理 NAS 上的文件/文件夹（优先）──
    try:
        from nas_client import NAS_AVAILABLE, delete_file
        if NAS_AVAILABLE:
            # 报告正式归档路径（可能是文件夹，如 /QA/.../报告名/）
            for p in [nas_report_path, nas_staging_path]:
                if p:
                    try:
                        if delete_file(p):
                            cleaned.append(f"NAS报告: {p}")
                    except Exception as e:
                        cleaned.append(f"NAS报告删除失败: {p} ({str(e)[:60]})")
            # 图片文件夹（可能是多段用 | 分隔）
            for p in [x.strip() for x in nas_picture_path.split("|") if x.strip()]:
                try:
                    if delete_file(p):
                        cleaned.append(f"NAS图片: {p}")
                except Exception as e:
                    cleaned.append(f"NAS图片删除失败: {p} ({str(e)[:60]})")
    except Exception:
        pass

    # ── 2. 清理本地回退文件 ──
    local_targets = []
    for p in [file_path] + [x.strip() for x in image_paths.split("|") if x.strip()]:
        if p and not p.startswith("/QA/") and os.path.exists(p):
            local_targets.append(p)
    for p in local_targets:
        try:
            if os.path.isdir(p):
                shutil.rmtree(p)
            else:
                os.remove(p)
            cleaned.append(f"本地: {p}")
        except Exception:
            pass

    # ── 3. 删除数据库记录 ──
    conn = get_connection()
    conn.execute("DELETE FROM inspection_reports WHERE id = ?", (report_id,))
    conn.commit()
    conn.close()

    if cleaned:
        return True, f"检验报告已删除，并清理 {len(cleaned)} 个关联文件/文件夹"
    return True, "检验报告已删除（无关联文件需清理）"


def get_report_daily_stats():
    """每日报告统计"""
    conn = get_connection()
    today = date.today()
    total = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE date(created_at)=?",
        (str(today),)
    ).fetchone()[0]
    pending = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE date(created_at)=? AND status='待审核'",
        (str(today),)
    ).fetchone()[0]
    approved = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE date(created_at)=? AND status='已通过'",
        (str(today),)
    ).fetchone()[0]
    conn.close()
    return {'total': total, 'pending': pending, 'approved': approved}


def _get_names_data():
    """获取原始名单数据（优先Excel，fallback到硬编码）"""
    if os.path.exists(NAMES_FILE):
        try:
            return _read_excel_all()
        except Exception:
            pass
    return _names_fallback()


def _read_excel_all():
    """从Excel读取所有列"""
    import openpyxl
    wb = openpyxl.load_workbook(NAMES_FILE, read_only=True, data_only=True)
    ws = wb['Sheet1']
    cols = {1: [], 2: [], 3: [], 4: [], 5: []}
    for r in range(2, ws.max_row + 1):
        for c in cols:
            val = ws.cell(r, c).value
            if val:
                v = str(val).strip()
                if v and v.lower() not in ('no brand', ''):
                    cols[c].append(v)
    wb.close()
    return {k: list(dict.fromkeys(v)) for k, v in cols.items()}


def _names_fallback():
    """硬编码的名单数据（当Excel文件不可用时使用）"""
    return {
        1: ['ACE', 'BOC', 'Langis LLC', 'Root', 'Z_Archived'],
        2: ['ACE', 'Aura', 'BigRock', 'Epicarry', 'KPL', 'Kronos', 'Orion', 'Parts', 'RaChat', 'Root', 'Root-Misc', 'TheUnicorn'],
        3: ['A11N Sports', 'Aireal', 'Airthereal', 'Baofeng', 'BELSIZE', 'BLIZZARD', 'BUYDEEM', 'CREALITY', 'darkFlash', 'DELI', 'Furlihong', 'Genmitsu', 'GLENCREAG', 'iPettie', 'Kronos-Misc', 'Lagute', 'LTC', 'OFFNOVA', 'OPENHEAT', 'Orion-Misc', 'ororo', 'Raddy', 'Radioddity', 'Razorri', 'Redragon', 'Root-Misc', 'Royal Kludge', 'SainSmart', 'SainSmart Jr.', 'SWONDER', 'TOSOT', 'TURBRO', 'WOODSTARTER', 'XIEGU'],
        4: ['Hibiscus', 'Ebony', 'Ace', 'Cactus', 'ET', 'ME', 'QA'],
        5: ['Carl Dong董献民', 'joung.yuan袁毅洪', 'haruna.wei韦梦婷', 'amelia.han韩亚南', 'teddy.li黎晓锋', 'colin.xu徐胜涛', 'lucy.ning宁小连', 'ken.huang黄海森', 'lainey.pan潘杨阳', 'fowler.zhai翟始福', 'leo.wu吴嘉俊', 'wenzel.chen陈文钊', 'bruce.cheng程强'],
    }


def get_bg_list():
    return _get_names_data()[1]

def get_bu_list():
    return _get_names_data()[2]

def get_brand_list():
    return _get_names_data()[3]

def get_rd_teams():
    return _get_names_data()[4]

def get_quality_users_list():
    return _get_names_data()[5]


# ==================== 活动日志 ====================

def _resolve_remote_audit_db():
    """解析审计的远程数据源（Mac 本地实时查看 Win 生产库）。
    优先级：WIN_DB_SMB_PATH 环境变量 > .windows_sync.json 的 target > /Volumes 自动搜索。
    返回 Win 库绝对路径；或 None（走本地库）。
    仅当 Mac(darwin) 或显式 LAB_AUDIT_REMOTE=1 时尝试；LAB_AUDIT_LOCAL=1 强制本地。"""
    import os, sys, json
    if os.environ.get("LAB_AUDIT_LOCAL") == "1":
        return None
    force_remote = os.environ.get("LAB_AUDIT_REMOTE") == "1"
    auto = (sys.platform == "darwin")  # Mac 开发者：默认尝试实时直连 Win
    if not (force_remote or auto):
        return None
    # 1) 显式 SMB 路径
    p = os.environ.get("WIN_DB_SMB_PATH")
    if p and os.path.exists(p):
        return p
    # 2) 从 .windows_sync.json 读已保存的 Win target
    try:
        sync_cfg = os.path.join(PROJECT_ROOT,
                                ".windows_sync.json")
        if os.path.exists(sync_cfg):
            with open(sync_cfg, "r", encoding="utf-8") as _f:
                cfg = json.load(_f)
            t = cfg.get("target_path")
            if t:
                cand = os.path.join(t, "data", "lab_manager.db")
                if os.path.exists(cand):
                    return cand
    except Exception:
        pass
    # 3) /Volumes 下兜底搜索
    try:
        for root, _dirs, files in os.walk("/Volumes"):
            if "lab_manager.db" in files:
                return os.path.join(root, "lab_manager.db")
    except Exception:
        pass
    return None


def _get_audit_connection():
    """返回审计连接：(conn, is_remote)。
    优先只读直连 Win 生产库（实时），任意失败自动回退本地库。"""
    import sqlite3
    remote = _resolve_remote_audit_db()
    if remote:
        try:
            con = sqlite3.connect(f"file:{remote}?mode=ro", uri=True, timeout=10)
            con.row_factory = sqlite3.Row  # 与本地 get_connection 一致，确保 dict(r) 可用
            con.execute("SELECT 1 FROM operation_log LIMIT 1")  # 校验表存在
            return con, True
        except Exception:
            try:
                con.close()
            except Exception:
                pass
    return get_connection(), False


def get_audit_source():
    """返回当前审计数据源：
    'production' = Win 本机生产库（本机运行的库即生产库本身，实时，无需远程）；
    'remote'     = Mac 实时直连 Win 生产库只读；
    'local'      = Mac 读本地库（Win 不可达，已回退）。
    """
    import sys
    # Win 生产环境：本机进程直接读写 data/lab_manager.db，这个库就是生产库，
    # 不应误报为"回退本地"。FORCE_PRODUCTION 为兼容性环境变量兜底。
    if sys.platform == "win32" or os.environ.get("FORCE_PRODUCTION") == "1":
        return "production"
    return "remote" if _resolve_remote_audit_db() else "local"


def get_operation_logs(limit=500, operator="", network="", action="", target_table="",
                       date_from="", date_to=""):
    """查询操作审计日志（支持操作人/网络/动作/数据表/日期筛选）。
    默认在 Mac 本地直连 Win 生产库只读，实现零 pull 实时查看。"""
    try:
        conn, _is_remote = _get_audit_connection()
    except Exception:
        return []
    try:
        conditions = []
        params = []
        if operator:
            conditions.append("operator LIKE ?")
            params.append(f"%{operator}%")
        if network:
            conditions.append("network = ?")
            params.append(network)
        if action:
            conditions.append("action LIKE ?")
            params.append(f"%{action}%")
        if target_table:
            conditions.append("target_table = ?")
            params.append(target_table)
        if date_from:
            conditions.append("created_at >= ?")
            params.append(date_from)
        if date_to:
            conditions.append("created_at <= ?")
            params.append(date_to + " 23:59:59")
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        rows = conn.execute(
            f"SELECT * FROM operation_log {where} ORDER BY id DESC LIMIT ?",
            params + [int(limit)],
        ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []
    finally:
        try:
            conn.close()
        except Exception:
            pass


def log_activity(user_email, action, category='system', detail='', page=''):
    """记录用户活动日志"""
    import re
    conn = get_connection()
    user_name = user_email.split('@')[0] if '@' in user_email else user_email
    # 安全截断，避免超长
    detail = (detail or '')[:500]
    page = (page or '')[:200]
    conn.execute(
        """INSERT INTO activity_log (user_email, user_name, action, category, detail, page)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (user_email, user_name, action, category, detail, page)
    )
    conn.commit()
    conn.close()


def get_activity_logs(limit=200, category='', user_email='', hours=24):
    """获取活动日志"""
    conn = get_connection()
    conditions = ["created_at >= datetime('now', 'localtime', ?)"]
    params = [f'-{hours} hours']
    if category:
        conditions.append("category = ?")
        params.append(category)
    if user_email:
        conditions.append("user_email = ?")
        params.append(user_email)
    where = " AND ".join(conditions)
    sql = f"SELECT * FROM activity_log WHERE {where} ORDER BY id DESC LIMIT ?"
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_online_users(minutes=15):
    """获取最近在线的用户（X分钟内有过活动）"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT user_email, user_name, MAX(created_at) as last_active,
                  COUNT(*) as action_count
           FROM activity_log
           WHERE created_at >= datetime('now', 'localtime', ?)
           GROUP BY user_email, user_name
           ORDER BY last_active DESC""",
        (f'-{minutes} minutes',)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_login_history(limit=50):
    """获取登录历史"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT * FROM activity_log WHERE action = '登录成功'
           ORDER BY id DESC LIMIT ?""", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_daily_stats():
    """获取每日访问统计"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT date(created_at) as day,
               COUNT(DISTINCT user_email) as unique_users,
               COUNT(*) as total_actions,
               SUM(CASE WHEN action = '登录成功' THEN 1 ELSE 0 END) as logins,
               SUM(CASE WHEN category = 'page_view' THEN 1 ELSE 0 END) as page_views,
               SUM(CASE WHEN category = 'data_edit' THEN 1 ELSE 0 END) as data_edits
        FROM activity_log
        WHERE created_at >= date('now', '-7 days')
        GROUP BY day
        ORDER BY day DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_sample_bg_list():
    """获取样品表中实际存在的BG列表"""
    conn = get_connection()
    rows = conn.execute("SELECT DISTINCT bg FROM samples WHERE bg != '' ORDER BY bg").fetchall()
    conn.close()
    return [r['bg'] for r in rows]


def get_page_hotspots(limit=20):
    """获取热门页面访问统计"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT page, COUNT(*) as visit_count,
                  COUNT(DISTINCT user_email) as unique_users
           FROM activity_log WHERE category = 'page_view'
           AND created_at >= datetime('now', '-7 days')
           GROUP BY page ORDER BY visit_count DESC LIMIT ?""", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 版本变动日志 ====================

def add_changelog(version, title, description='', changes='', category='优化', created_by=''):
    """添加版本变动记录"""
    conn = get_connection()
    conn.execute(
        """INSERT INTO changelog (version, title, description, changes, category, created_by)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (version, title, description, changes, category, created_by)
    )
    conn.commit()
    conn.close()


def get_changelogs(limit=50):
    """获取版本变动日志"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM changelog ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@recoverable("changelog")
def delete_changelog(changelog_id):
    """删除版本变动记录"""
    conn = get_connection()
    conn.execute("DELETE FROM changelog WHERE id = ?", (changelog_id,))
    conn.commit()
    conn.close()
    return True, "已删除"


@recoverable("activity_log")
def delete_activity(activity_id):
    """删除活动日志"""
    conn = get_connection()
    conn.execute("DELETE FROM activity_log WHERE id = ?", (activity_id,))
    conn.commit()
    conn.close()
    return True, "已删除"


@recoverable("activity_log")
def delete_activities(activity_ids):
    """批量删除活动日志"""
    conn = get_connection()
    for aid in activity_ids:
        conn.execute("DELETE FROM activity_log WHERE id = ?", (aid,))
    conn.commit()
    conn.close()
    return True, f"已删除 {len(activity_ids)} 条记录"


# ==================== 首页看板聚合函数 ====================

def get_sample_dashboard_stats():
    """获取样品看板统计数据"""
    conn = get_connection()
    today = datetime.now().strftime("%Y-%m-%d")
    thirty_days = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")

    total = conn.execute("SELECT COUNT(*) FROM samples").fetchone()[0]
    in_stock = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE out_status != '已出库' OR out_status IS NULL"
    ).fetchone()[0]
    out_stock = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE out_status = '已出库'"
    ).fetchone()[0]
    near_expiry = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE expiry_date BETWEEN ? AND ? AND expiry_date != ''",
        (today, thirty_days)
    ).fetchone()[0]
    expired = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE expiry_date < ? AND expiry_date != ''",
        (today,)
    ).fetchone()[0]
    conn.close()
    return {
        'total': total, 'in_stock': in_stock, 'out_stock': out_stock,
        'near_expiry': near_expiry, 'expired': expired
    }


def get_sample_bg_distribution():
    """获取样品 BG 分布"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT bg, COUNT(*) as count FROM samples WHERE bg != '' GROUP BY bg ORDER BY count DESC"
    ).fetchall()
    conn.close()
    return [{'bg': r[0], 'count': r[1]} for r in rows]


def get_change_dashboard_stats():
    """获取变更看板统计数据（含 SKU 确认状态分布）"""
    conn = get_connection()
    this_month = datetime.now().strftime("%Y-%m")
    total = conn.execute("SELECT COUNT(*) FROM change_records").fetchone()[0]
    monthly = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE change_date LIKE ?",
        (f"{this_month}%",)
    ).fetchone()[0]
    recent = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE created_at >= datetime('now', '-7 days')"
    ).fetchone()[0]
    all_confirmed = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE overall_status = '全部确认'"
    ).fetchone()[0]
    partial = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE overall_status = '部分确认'"
    ).fetchone()[0]
    conn.close()
    return {
        'total': total, 'this_month': monthly, 'recent_7d': recent,
        'confirmed': all_confirmed, 'partial': partial,
        'unconfirmed': total - all_confirmed - partial
    }


def get_change_bu_distribution():
    """获取变更 BU 分布"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT bu, COUNT(*) as count FROM change_records WHERE bu != '' GROUP BY bu ORDER BY count DESC"
    ).fetchall()
    conn.close()
    return [{'bu': r[0], 'count': r[1]} for r in rows]


def get_inspection_dashboard_stats():
    """获取检验报告看板统计"""
    conn = get_connection()
    total = conn.execute("SELECT COUNT(*) FROM inspection_reports").fetchone()[0]
    pending = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE status = '待审核'"
    ).fetchone()[0]
    approved = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE status = '已通过'"
    ).fetchone()[0]
    rejected = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE status = '已驳回'"
    ).fetchone()[0]
    this_month = datetime.now().strftime("%Y-%m")
    monthly = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE created_at LIKE ?",
        (f"{this_month}%",)
    ).fetchone()[0]
    conn.close()
    return {
        'total': total, 'pending': pending, 'approved': approved,
        'rejected': rejected, 'this_month': monthly
    }


def get_report_type_distribution():
    """获取报告类型分布"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT report_type, COUNT(*) as count FROM inspection_reports WHERE report_type != '' GROUP BY report_type ORDER BY count DESC"
    ).fetchall()
    conn.close()
    return [{'type': r[0], 'count': r[1]} for r in rows]


def get_recent_reports(limit=6):
    """获取最近检验报告"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT id, report_type, inspector, product_name, bg, bu, brand, status, created_at
           FROM inspection_reports ORDER BY id DESC LIMIT ?""",
        (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_recent_changes(limit=6):
    """获取最近变更记录"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT id, bu, brand, sku, change_reason, supplier, change_date, confirm_date, created_at
           FROM change_records ORDER BY id DESC LIMIT ?""",
        (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_expiring_samples(days=30):
    """获取即将过期和已过期的样品（包含过期和即将到期的）"""
    conn = get_connection()
    today = datetime.now().strftime("%Y-%m-%d")
    deadline = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    expired_since = (datetime.now() - timedelta(days=730)).strftime("%Y-%m-%d")
    rows = conn.execute(
        """SELECT id, bg, sku, sample_name, expiry_date, sign_date, supplier, out_status
           FROM samples
           WHERE expiry_date >= ? AND expiry_date <= ? AND expiry_date != ''
             AND (out_status IS NULL OR out_status != '已出库')
           ORDER BY expiry_date ASC""",
        (expired_since, deadline)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_equipment_usage_trend():
    """获取设备使用趋势（最近7天）"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT borrow_date, COUNT(*) as count
           FROM borrow_records
           WHERE borrow_date >= date('now', '-7 days')
           GROUP BY borrow_date ORDER BY borrow_date"""
    ).fetchall()
    conn.close()
    return [{'date': r[0], 'count': r[1]} for r in rows]


@audit("更新借用记录", "borrow_records")
def update_borrow_record(record_id, updates):
    """更新借用记录（支持编辑借用日期/预计归还/借用人/用途等）"""
    conn = get_connection()
    allowed = ['user_id', 'borrow_date', 'expected_return_date',
               'actual_return_date', 'purpose', 'notes', 'status']
    set_parts = []
    params = []
    for k, v in updates.items():
        if k in allowed and v is not None:
            set_parts.append(f"{k} = ?")
            params.append(v)
    if not set_parts:
        conn.close()
        return False, "没有有效的更新字段"
    params.append(record_id)
    sql = f"UPDATE borrow_records SET {', '.join(set_parts)} WHERE id = ?"
    conn.execute(sql, params)
    conn.commit()
    conn.close()
    return True, f"记录 #{record_id} 更新成功"


@audit("删除借用记录", "borrow_records")
@recoverable("borrow_records")
def delete_borrow_record(record_id):
    """删除借用记录（硬删除，仅限未归还的记录）"""
    conn = get_connection()
    row = conn.execute(
        "SELECT status, equipment_id FROM borrow_records WHERE id = ?", (record_id,)
    ).fetchone()
    if not row:
        conn.close()
        return False, "记录不存在"
    status, eq_id = row
    # 如果记录不是已归还，需恢复设备状态
    if status != '已归还':
        conn.execute(
            "UPDATE equipment SET status = '可用' WHERE id = ?", (eq_id,)
        )
    conn.execute("DELETE FROM borrow_records WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()
    return True, f"记录 #{record_id} 已删除"


def get_equipment_usage_frequency(limit=10):
    """获取设备使用频次排行（长期统计）"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT e.name, COUNT(br.id) as usage_count, e.location
           FROM borrow_records br
           JOIN equipment e ON br.equipment_id = e.id
           GROUP BY e.id
           ORDER BY usage_count DESC
           LIMIT ?
        """, (limit,)
    ).fetchall()
    conn.close()
    return [{'equipment': r[0], 'count': r[1], 'location': r[2]} for r in rows]


# ═══════════════════════════════════════════════════════════
# 检验报告 — 审批归档（NAS 暂存 → 正式路径）
# ═══════════════════════════════════════════════════════════

@audit("审核归档报告", "inspection_reports")
def approve_report_with_archival(report_id, brand="", sku="", po="", inspection_date="", reviewer="", comment=""):
    """
    审核通过报告，触发服务器本地暂存 → NAS 正式路径归档。

    流程：
      1. 从 DB 读取 file_path（服务器本地暂存）
      2. 直接读取本地暂存文件
      3. 保留系统原始上传文件名作为正式归档名
      4. 根据 NAS_ROUTING_MAP 获取目标正式路径并上传
      5. 删除本地暂存副本
      6. 更新 DB：正式路径写入 nas_report_path，清空 file_path / nas_staging_path，状态改为「已通过」

    返回:
        (ok_bool, message)
    """
    import os as _os
    import io as _io

    conn = get_connection()
    rpt = conn.execute("SELECT * FROM inspection_reports WHERE id = ?", (report_id,)).fetchone()
    if not rpt:
        conn.close()
        return False, "报告不存在"

    rpt = dict(rpt)
    # 报告文件暂存在服务器本地（file_path）
    local_paths = [p.strip() for p in str(rpt.get('file_path', '') or '').split('|') if p.strip()]
    report_type = rpt.get('report_type', '')

    insp_date = inspection_date or rpt.get('inspection_date', '')
    report_year = insp_date[:4] if insp_date and len(insp_date) >= 4 else ''

    # ── 懒加载 nas_client ──
    try:
        from nas_client import (
            upload_file as _nas_upload,
            delete_file as _nas_delete,
            get_nas_routes as _nas_routes,
            ensure_single_folder as _nas_ensure_folder,
        )
    except ImportError as e:
        conn.close()
        return False, f"NAS 客户端不可用：{e}"

    if not local_paths:
        conn.close()
        return False, "该报告无本地暂存文件，无法归档"

    # ── 1. 读取服务器本地暂存文件 ──
    staged_files = []
    try:
        for lp in local_paths:
            if _os.path.exists(lp):
                with open(lp, 'rb') as _fh:
                    staged_files.append((lp, _fh.read(), _os.path.basename(lp)))
            else:
                conn.close()
                return False, f"本地暂存文件不存在: {lp}"
    except Exception as e:
        conn.close()
        return False, f"读取暂存文件异常: {str(e)[:200]}"

    # ── 2. 保留系统原始上传文件名 ──
    raw_filenames = [
        _os.path.basename(p.strip()) for p in str(rpt.get('filename', '') or '').split('|') if p.strip()
    ]
    if not raw_filenames:
        raw_filenames = [_os.path.basename(p[2] or '') for p in staged_files]
    formal_filename = raw_filenames[0] if raw_filenames else ''
    if not formal_filename:
        formal_filename = f"report_{report_id}.pdf"

    # ── 3. PDF 原文件后追加审核签字页 ──
    # 上传报告来源版式不统一，不能假定签字坐标；追加独立签字页既保留原件，
    # 又让审核人、时间和意见进入正式归档 PDF，避免审核信息只停留在数据库。
    reviewer = (reviewer or "").strip()
    if reviewer and formal_filename.lower().endswith(".pdf"):
        try:
            import fitz as _fitz
            # Signature is appended to the primary PDF; non-PDF attachments
            # remain byte-for-byte unchanged and are archived alongside it.
            src_pdf = _fitz.open(stream=staged_files[0][1], filetype="pdf")
            page = src_pdf.new_page(width=595, height=842)
            page.insert_text((55, 70), "检验报告审核签字页", fontsize=18, color=(0.08, 0.14, 0.23))
            page.insert_text((55, 105), "Inspection Report Approval Record", fontsize=10, color=(0.35, 0.40, 0.47))
            lines = [
                f"报告编号：{rpt.get('id', report_id)}",
                f"产品：{rpt.get('product_name', '') or ''}",
                f"审核结论：通过 / Approved",
                f"审核人签字：{reviewer}",
                f"审核时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"审核意见：{comment or '审核通过'}",
            ]
            y = 165
            for line in lines:
                page.insert_text((55, y), line[:110], fontsize=11, color=(0.08, 0.14, 0.23))
                y += 32
            page.draw_line((55, y + 18), (350, y + 18), color=(0.45, 0.49, 0.55), width=0.8)
            page.insert_text((55, y + 38), "审核人签字 / Reviewer Signature", fontsize=9, color=(0.35, 0.40, 0.47))
            staged_files[0] = (staged_files[0][0], src_pdf.tobytes(garbage=4, deflate=True), staged_files[0][2])
            src_pdf.close()
        except Exception as exc:
            conn.close()
            return False, f"审核签字页生成失败：{str(exc)[:180]}"

    if reviewer:
        conn.execute(
            "UPDATE inspection_reports SET reviewer=?, review_comment=? WHERE id=?",
            (reviewer, comment or "审核通过", report_id),
        )

    # ── 4. 获取正式目标路径 ──
    if report_type and report_year:
        report_base, _ = _nas_routes(report_type, report_year)
    else:
        report_base = "/QA/验货相关文件/其他报告/2026年/"

    report_base_clean = report_base.rstrip('/')
    _nas_ensure_folder(report_base_clean, "")  # ensure base exists
    formal_paths = []
    for idx, (_staging_path, staged_bytes, staged_filename) in enumerate(staged_files):
        formal_name = raw_filenames[idx] if idx < len(raw_filenames) else _os.path.basename(staged_filename or '')
        formal_name = formal_name or f"report_{report_id}_{idx + 1}.pdf"
        ok_up, formal_path = _nas_upload(report_base_clean, formal_name, staged_bytes)
        if not ok_up:
            conn.close()
            return False, f"写入正式路径失败（{formal_name}）: {formal_path}"
        formal_paths.append(formal_path)

    # ── 5. 删除服务器本地暂存副本 ──
    for lp in local_paths:
        if _os.path.exists(lp):
            try:
                _os.remove(lp)
            except Exception:
                pass  # 删除失败不阻断归档

    # ── 6. 更新数据库 ──
    conn.execute("""
        UPDATE inspection_reports
        SET status = '已通过',
            nas_report_path = ?,
            nas_staging_path = '',
            file_path = ''
        WHERE id = ?
    """, (' | '.join(formal_paths), report_id))
    conn.commit()
    conn.close()

    return True, f"归档完成: {' | '.join(formal_paths)}"
