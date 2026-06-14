"""
实验室设备管理系统 - 数据库层
提供 SQLite 数据库初始化、CRUD 操作和统计查询接口
"""

import sqlite3
import os
from datetime import date, timedelta, datetime

# 路径
DB_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
DB_PATH = os.path.join(DB_DIR, 'lab_manager.db')
# 原始名单文件路径
NAMES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          'SainStore实验室文件', '原始名单.xlsx')


def get_connection():
    """获取数据库连接"""
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    """初始化数据库表结构和种子数据"""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.executescript('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            description TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS equipment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            model TEXT DEFAULT '',
            serial_number TEXT UNIQUE,
            category_id INTEGER,
            location TEXT DEFAULT '',
            status TEXT DEFAULT '可用' CHECK(status IN ('可用','借出','维修中','报废')),
            purchase_date TEXT DEFAULT '',
            price REAL DEFAULT 0,
            supplier TEXT DEFAULT '',
            warranty_expiry TEXT DEFAULT '',
            description TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            department TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            email TEXT DEFAULT '',
            role TEXT DEFAULT '普通用户' CHECK(role IN ('管理员','普通用户')),
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS borrow_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            borrow_date TEXT NOT NULL,
            expected_return_date TEXT DEFAULT '',
            actual_return_date TEXT DEFAULT '',
            test_standard TEXT DEFAULT '',
            test_start_time TEXT DEFAULT '',
            test_end_time TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            status TEXT DEFAULT '借出中' CHECK(status IN ('借出中','已归还','逾期','已出库')),
            purpose TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (equipment_id) REFERENCES equipment(id),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS maintenance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_id INTEGER NOT NULL,
            maintenance_date TEXT NOT NULL,
            maintenance_type TEXT DEFAULT '定期保养' CHECK(maintenance_type IN ('定期保养','故障维修','校准','其他')),
            description TEXT DEFAULT '',
            cost REAL DEFAULT 0,
            technician TEXT DEFAULT '',
            next_maintenance_date TEXT DEFAULT '',
            status TEXT DEFAULT '已完成' CHECK(status IN ('已完成','进行中','计划中')),
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (equipment_id) REFERENCES equipment(id)
        );

        CREATE TABLE IF NOT EXISTS samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bg TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            sample_name TEXT DEFAULT '',
            sign_date TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            location TEXT DEFAULT '',
            expiry_date TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS change_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bu TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            change_reason TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            attachments TEXT DEFAULT '',
            change_date TEXT DEFAULT '',
            confirm_date TEXT DEFAULT '',
            confirm_person TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS inspection_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_type TEXT DEFAULT '',
            inspector TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            bg TEXT DEFAULT '',
            bu TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            sku TEXT DEFAULT '',
            filename TEXT DEFAULT '',
            file_path TEXT DEFAULT '',
            image_paths TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            status TEXT DEFAULT '待审核' CHECK(status IN ('待审核','已通过','已驳回')),
            reviewer TEXT DEFAULT '',
            review_comment TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_email TEXT NOT NULL,
            user_name TEXT DEFAULT '',
            action TEXT NOT NULL,
            category TEXT DEFAULT 'system',
            detail TEXT DEFAULT '',
            page TEXT DEFAULT '',
            ip_address TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS changelog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            version TEXT NOT NULL,
            title TEXT DEFAULT '',
            description TEXT DEFAULT '',
            changes TEXT DEFAULT '',
            category TEXT DEFAULT '优化',
            created_by TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS sample_outbound (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sample_id INTEGER NOT NULL,
            qty INTEGER DEFAULT 1,
            out_date TEXT DEFAULT '',
            borrower TEXT DEFAULT '',
            department TEXT DEFAULT '',
            reason TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (sample_id) REFERENCES samples(id)
        );
    ''')

    # 数据库迁移：兼容旧表结构
    try:
        cursor.execute("ALTER TABLE activity_log ADD COLUMN category TEXT DEFAULT 'system'")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE activity_log ADD COLUMN page TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN brand TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN expiry_date TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE change_records ADD COLUMN supplier TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN out_status TEXT DEFAULT '在库'")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE samples ADD COLUMN stock_qty INTEGER DEFAULT 1")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE change_records ADD COLUMN rd_team TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE inspection_reports ADD COLUMN supplier TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN brand TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN sku TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE borrow_records ADD COLUMN product_name TEXT DEFAULT ''")
    except:
        pass

    # 仅在表为空时插入种子数据
    cursor.execute("SELECT COUNT(*) FROM categories")
    if cursor.fetchone()[0] == 0:
        seed_data(cursor)

    conn.commit()
    conn.close()


def seed_data(cursor):
    """插入初始种子数据 - 来源: 实验室设备配备一览表1(1).xlsx (27台设备)"""
    categories = [
        ('力学测试设备', '按键寿命、插拔力、插拔寿命、跌落、振动、摇摆、拉力、纽扣拉力、硬度计'),
        ('环境测试设备', '盐雾、淋雨IPX、恒温恒湿、冷热冲击、温度巡检'),
        ('电学测试设备', 'ESD静电枪、安全性能综合分析仪、阻抗测试'),
        ('电源与电池测试', '直流电源、稳压电源、电池测试、PD负载、模拟电池'),
        ('几何量测与辅助', '二次元影像测量、发热模组温度箱'),
    ]
    cursor.executemany("INSERT INTO categories (name, description) VALUES (?, ?)", categories)

    equipment_list = [
        # ==================== 力学测试设备 (cat 1) ====================
        ('按键寿命试验机', 'LS-AJ-400', 'ET0006', 1, '实验室1', '可用',
         '2024-06-01', 0, '力试', '', '按键耐久性疲劳测试 (2万~10万次) | IEC61058-1'),
        ('微电脑插拔试验机', 'LS-CB-50', 'ET0003', 1, '实验室1', '可用',
         '2024-06-01', 0, '力试', '', 'USB/Type-C/连接器插拔力测试，50kg量程 | EIA-364-09, USB-IF'),
        ('插拔寿命试验机', 'LS-SM-65', 'ET0004', 1, '实验室1', '可用',
         '2024-06-01', 0, '力试', '', '连接器拔插寿命测试，65mm行程 (5000~10000次) | EIA-364-09'),
        ('单翼跌落试验机', 'LS-DL-150', 'ET0001', 1, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', 'ISTA 6-Amazon SIOC 包装跌落测试，150cm | ISTA 6A 亚马逊原箱发货认证'),
        ('模拟运输振动试验机', 'LS-YS-100', 'ET0007', 1, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', '随机/正弦振动模拟长途物流，100kg负载 | ASTM D4728, GB/T4857.7'),
        ('摇摆试验机', 'LS-YB-600', 'ET0005', 1, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', '线材弯折摇摆±90°疲劳测试 | IEC60335-1 §25.14'),
        ('万能拉力试验机', 'GR-WCJ5T', 'ET0008', 1, '实验室2', '可用',
         '2024-06-01', 0, '固润', '', '拉伸/压缩/弯曲/剥离，5T量程 | ISO527, ASTM D3359'),
        ('纽扣拉力试验机', '', None, 1, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '纽扣/按扣/四合扣拉力测试'),
        ('硬度计', '', 'QA049', 1, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '材料硬度测试 (金属/塑料/橡胶)'),
        # ==================== 环境测试设备 (cat 2) ====================
        ('盐雾试验机', 'LS-UT-60', 'ET0002', 2, '盐雾实验室', '可用',
         '2024-06-01', 0, '力试', '', '中性盐雾腐蚀加速测试 (24H~96H+) | ISO9227, ASTM B117, GB/T10125'),
        ('可编程淋雨试验机', 'LS-IPX3456-512', 'ET0010', 2, '实验室2', '可用',
         '2024-06-01', 0, '力试', '', 'IPX3/4/5/6 防水等级测试 | IEC60529/GB/T4208'),
        ('温湿度可编程试验机', 'LS-TH-800Z', 'ET0009', 2, '实验室3', '可用',
         '2024-06-01', 0, '力试', '', '恒温恒湿 (-45~150℃/0~100%RH) 800L | IEC60068-2-78/2-30'),
        ('热冲击试验机', 'LS-THS-180Z', 'ET0011', 2, '实验室3', '可用',
         '2024-06-01', 0, '力试', '', '高低温骤变冲击 (-70~+150℃, 5min转换) | IEC60068-2-14'),
        ('温度巡检仪', '', 'QA018', 2, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '多点温度实时巡检记录 (2台)'),
        # ==================== 电学测试设备 (cat 3) ====================
        ('ESD静电枪试验机', 'PESD6020', 'ET0012', 3, 'ESD静电房', '可用',
         '2024-06-01', 0, '普锐马', '', '接触放电±4kV/空气放电±8kV，20kV | IEC61000-4-2'),
        ('安全性能综合分析仪', 'AC1651B', 'ET013', 3, '实验室1', '可用',
         '2024-06-01', 0, '安规', '', '耐压(1500V)/绝缘(≥2MΩ)/接地/泄漏(≤0.75mA) | IEC60335-1, IEC62368-1'),
        ('阻抗测试仪', '', 'QA073', 3, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '电阻/阻抗/导通精密测量'),
        # ==================== 电源与电池测试 (cat 4) ====================
        ('电池综合测试仪', '', 'INST0023', 4, '品质部实验室', '可用',
         '2024-09-01', 0, '', '', '电池容量/内阻/充放电/BMS保护测试 | UN38.3, IEC62133, UL2054'),
        ('模拟电池测试仪', '', 'INST0024', 4, '品质部实验室', '可用',
         '2024-09-01', 0, '', '', '模拟电池异常状态(过压/欠压/短路) | IEC62368-1'),
        ('大功率直流电源', '', 'INST0025', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '100A/6000W大功率直流供电 | UL1995, IEC60335-2-40'),
        ('PD负载测试仪', '', 'QA009', 4, '品质部实验室', '可用',
         '2025-03-01', 0, '', '', 'USB-PD 3.0/3.1快充协议测试 (5V~20V) | USB-IF PD规范'),
        ('负载测试仪', '', 'QA001', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '电子负载电流/功率测试'),
        ('直流电源', '', 'QA017', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '直流供电电源 (2台)'),
        ('直流稳压电源', '', 'INST0021', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '高精度直流稳压供电'),
        ('可调节直流稳压电源', '', 'QA044', 4, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '可调电压电流直流稳压输出'),
        # ==================== 几何量测与辅助 (cat 5) ====================
        ('二次元测量仪器', '', None, 5, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '光学影像精密测量 (微米级精度) | ISO10360-7'),
        ('发热模组温度测试箱', '', None, 5, '品质部实验室', '可用',
         '2024-06-01', 0, '', '', '加热服装发热模组温升曲线监控 | UL130, IEC60335-2-17'),
    ]
    cursor.executemany(
        """INSERT INTO equipment
        (name, model, serial_number, category_id, location, status,
         purchase_date, price, supplier, warranty_expiry, description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        equipment_list
    )

    users = [
        ('Carl Dong董献民', 'ACE/ACE', 'ACE', 'ACE', '管理员'),
        ('joung.yuan袁毅洪', 'BOC/Aura', 'BOC', 'Aura', '管理员'),
        ('haruna.wei韦梦婷', 'Langis LLC/BigRock', 'Langis LLC', 'BigRock', '管理员'),
        ('amelia.han韩亚南', 'no brand/Epicarry', 'no brand', 'Epicarry', '管理员'),
        ('teddy.li黎晓锋', 'Root/KPL', 'Root', 'KPL', '管理员'),
        ('colin.xu徐胜涛', 'Z_Archived/Kronos', 'Z_Archived', 'Kronos', '管理员'),
        ('lucy.ning宁小连', 'no brand', 'no brand', '', '管理员'),
        ('ken.huang黄海森', 'Orion', '', 'Orion', '管理员'),
        ('lainey.pan潘杨阳', 'Parts', '', 'Parts', '管理员'),
        ('fowler.zhai翟始福', 'RaChat', '', 'RaChat', '管理员'),
        ('leo.wu吴嘉俊', 'Root', 'Root', '', '管理员'),
        ('wenzel.chen陈文钊', 'Root-Misc', 'Root-Misc', '', '管理员'),
        ('bruce.cheng程强', 'TheUnicorn', '', 'TheUnicorn', '管理员'),
    ]
    cursor.executemany(
        "INSERT INTO users (name, department, phone, email, role) VALUES (?, ?, ?, ?, ?)",
        users
    )

    today = date.today()
    borrow_records = [
        (1, 8, str(today - timedelta(days=5)), str(today + timedelta(days=2)),
         'IEC61058-1', '09:00', '17:00', '', '', 'PD5K 蓝牙键盘',
         None, '借出中', '新品按键寿命验证', ''),
        (11, 13, str(today - timedelta(days=10)), str(today - timedelta(days=3)),
         'IEC60529 IPX5', '10:00', '16:00', 'TURBRO', 'TB-JXH-001', 'TURBRO接线盒',
         str(today - timedelta(days=3)), '已归还', 'TURBRO接线盒防水测试', '测试完成，数据正常'),
        (5, 12, str(today - timedelta(days=15)), str(today - timedelta(days=8)),
         'ASTM D4728', '08:30', '12:00', '', '', '房车空调外包装',
         str(today - timedelta(days=9)), '已归还', '出口包装振动测试', ''),
    ]
    cursor.executemany(
        """INSERT INTO borrow_records
        (equipment_id, user_id, borrow_date, expected_return_date,
         test_standard, test_start_time, test_end_time, brand, sku, product_name,
         actual_return_date, status, purpose, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        borrow_records
    )

    maintenance_records = [
        (10, str(today - timedelta(days=60)), '定期保养', '盐雾试验机月度保养：清洗喷嘴、检查加热管、补充盐溶液',
         0, 'Carl Dong董献民', str(today + timedelta(days=300)), '已完成', '设备运行正常，盐雾沉降量符合标准'),
        (15, str(today - timedelta(days=10)), '校准', 'ESD静电枪年度计量校准（±4kV接触/±8kV空气）',
         800.00, '计量所', str(today + timedelta(days=355)), '已完成', '校准证书编号: CAL-2026-ESD-001'),
        (16, str(today - timedelta(days=90)), '校准', '安规综合分析仪年度校准（耐压/绝缘/泄漏电流）',
         600.00, '计量所', str(today + timedelta(days=270)), '已完成', '校准证书编号: CAL-2026-SFT-001'),
        (12, str(today + timedelta(days=15)), '定期保养', '温湿度可编程试验机季度保养：清洁滤网、检查制冷系统',
         0, 'teddy.li黎晓锋', None, '计划中', '按季度保养计划执行'),
        (11, str(today + timedelta(days=30)), '定期保养', '淋雨试验机月度保养：检查水泵压力、清洁各喷头',
         0, 'bruce.cheng程强', None, '计划中', '按月度保养计划执行'),
    ]
    cursor.executemany(
        """INSERT INTO maintenance_records
        (equipment_id, maintenance_date, maintenance_type, description,
         cost, technician, next_maintenance_date, status, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        maintenance_records
    )


# ==================== 分类操作 ====================

def get_categories():
    """获取所有设备分类"""
    conn = get_connection()
    rows = conn.execute("SELECT * FROM categories ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_category(name, description=''):
    """添加设备分类"""
    conn = get_connection()
    try:
        conn.execute("INSERT INTO categories (name, description) VALUES (?, ?)", (name, description))
        conn.commit()
        return True, "分类添加成功"
    except sqlite3.IntegrityError:
        return False, "分类名称已存在"
    finally:
        conn.close()


def update_category(cat_id, name, description):
    """更新设备分类"""
    conn = get_connection()
    try:
        conn.execute("UPDATE categories SET name=?, description=? WHERE id=?", (name, description, cat_id))
        conn.commit()
        return True, "分类更新成功"
    except sqlite3.IntegrityError:
        return False, "分类名称已存在"
    finally:
        conn.close()


def delete_category(cat_id):
    """删除设备分类"""
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM equipment WHERE category_id=?", (cat_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该分类下有 {count} 台设备，无法删除"
    conn.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    conn.commit()
    conn.close()
    return True, "分类删除成功"


# ==================== 设备操作 ====================

def get_equipment(search='', category_id=None, status=None, location='', page=1, per_page=20):
    """查询设备列表（支持搜索和筛选）"""
    conn = get_connection()
    conditions = []
    params = []

    if search:
        conditions.append("(e.name LIKE ? OR e.model LIKE ? OR e.serial_number LIKE ?)")
        params.extend([f'%{search}%'] * 3)
    if category_id:
        conditions.append("e.category_id = ?")
        params.append(category_id)
    if status:
        conditions.append("e.status = ?")
        params.append(status)
    if location:
        conditions.append("e.location LIKE ?")
        params.append(f'%{location}%')

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    total = conn.execute(f"SELECT COUNT(*) FROM equipment e {where}", params).fetchone()[0]

    offset = (page - 1) * per_page
    sql = f"""
        SELECT e.*, c.name as category_name
        FROM equipment e
        LEFT JOIN categories c ON e.category_id = c.id
        {where}
        ORDER BY e.id DESC
        LIMIT ? OFFSET ?
    """
    rows = conn.execute(sql, params + [per_page, offset]).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_equipment_by_id(eq_id):
    """根据 ID 获取设备详情"""
    conn = get_connection()
    row = conn.execute(
        """SELECT e.*, c.name as category_name
           FROM equipment e LEFT JOIN categories c ON e.category_id = c.id
           WHERE e.id=?""", (eq_id,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def add_equipment(data):
    """添加新设备"""
    conn = get_connection()
    try:
        conn.execute("""
            INSERT INTO equipment (name, model, serial_number, category_id, location, status,
                                   purchase_date, price, supplier, warranty_expiry, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data['name'], data.get('model', ''), data.get('serial_number', ''),
            data.get('category_id'), data.get('location', ''), data.get('status', '可用'),
            data.get('purchase_date', ''), data.get('price', 0), data.get('supplier', ''),
            data.get('warranty_expiry', ''), data.get('description', '')
        ))
        conn.commit()
        return True, "设备添加成功"
    except sqlite3.IntegrityError as e:
        if 'serial_number' in str(e):
            return False, "设备编号已存在"
        return False, f"添加失败: {e}"
    finally:
        conn.close()


def update_equipment(eq_id, data):
    """更新设备信息"""
    conn = get_connection()
    try:
        conn.execute("""
            UPDATE equipment SET name=?, model=?, serial_number=?, category_id=?, location=?,
            status=?, purchase_date=?, price=?, supplier=?, warranty_expiry=?, description=?,
            updated_at=datetime('now','localtime')
            WHERE id=?
        """, (
            data['name'], data.get('model', ''), data.get('serial_number', ''),
            data.get('category_id'), data.get('location', ''), data.get('status', '可用'),
            data.get('purchase_date', ''), data.get('price', 0), data.get('supplier', ''),
            data.get('warranty_expiry', ''), data.get('description', ''), eq_id
        ))
        conn.commit()
        return True, "设备更新成功"
    except sqlite3.IntegrityError:
        return False, "设备编号已存在"
    finally:
        conn.close()


def delete_equipment(eq_id):
    """删除设备"""
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM borrow_records WHERE equipment_id=?", (eq_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该设备有 {count} 条借用记录，无法删除（可将其状态设为'报废'）"
    count = conn.execute("SELECT COUNT(*) FROM maintenance_records WHERE equipment_id=?", (eq_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该设备有 {count} 条维护记录，无法删除（可将其状态设为'报废'）"
    conn.execute("DELETE FROM equipment WHERE id=?", (eq_id,))
    conn.commit()
    conn.close()
    return True, "设备删除成功"


def import_equipment_batch(records):
    """批量导入设备"""
    conn = get_connection()
    success = 0
    errors = []
    for i, rec in enumerate(records):
        try:
            conn.execute("""
                INSERT INTO equipment (name, model, serial_number, category_id, location, status,
                                       purchase_date, price, supplier, warranty_expiry, description)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rec.get('name', f'设备{i + 1}'), rec.get('model', ''), rec.get('serial_number', ''),
                rec.get('category_id'), rec.get('location', ''), rec.get('status', '可用'),
                rec.get('purchase_date', ''), rec.get('price', 0), rec.get('supplier', ''),
                rec.get('warranty_expiry', ''), rec.get('description', '')
            ))
            success += 1
        except Exception as e:
            errors.append(f"第{i + 1}行: {e}")
    conn.commit()
    conn.close()
    return success, errors


def get_equipment_for_select():
    """获取设备下拉选项"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, name, serial_number, status FROM equipment WHERE status != '报废' ORDER BY name"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 人员操作 ====================

def get_users(search=''):
    """获取人员列表"""
    conn = get_connection()
    if search:
        rows = conn.execute(
            "SELECT * FROM users WHERE name LIKE ? OR department LIKE ? ORDER BY id",
            (f'%{search}%', f'%{search}%')
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM users ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_user(data):
    """添加人员"""
    conn = get_connection()
    conn.execute(
        "INSERT INTO users (name, department, phone, email, role) VALUES (?, ?, ?, ?, ?)",
        (data['name'], data.get('department', ''), data.get('phone', ''),
         data.get('email', ''), data.get('role', '普通用户'))
    )
    conn.commit()
    conn.close()
    return True, "人员添加成功"


def update_user(user_id, data):
    """更新人员信息"""
    conn = get_connection()
    conn.execute(
        "UPDATE users SET name=?, department=?, phone=?, email=?, role=? WHERE id=?",
        (data['name'], data.get('department', ''), data.get('phone', ''),
         data.get('email', ''), data.get('role', '普通用户'), user_id)
    )
    conn.commit()
    conn.close()
    return True, "人员更新成功"


def delete_user(user_id):
    """删除人员"""
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM borrow_records WHERE user_id=?", (user_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return False, f"该人员有 {count} 条借用记录，无法删除"
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return True, "人员删除成功"


# ==================== 借用操作 ====================

def borrow_equipment(equipment_id, user_id, borrow_date, expected_return_date,
                     purpose='', notes='', test_standard='', test_start_time='', test_end_time='',
                     brand='', sku='', product_name=''):
    """使用登记（实验室内部测试预约，不改变设备状态）"""
    conn = get_connection()
    eq = conn.execute("SELECT status FROM equipment WHERE id=?", (equipment_id,)).fetchone()
    if not eq:
        conn.close()
        return False, "设备不存在"

    # 检查是否有时间冲突的使用登记
    existing = conn.execute("""
        SELECT id FROM borrow_records
        WHERE equipment_id=? AND status='借出中'
        AND borrow_date <= ? AND expected_return_date >= ?
        LIMIT 1
    """, (equipment_id, expected_return_date, borrow_date)).fetchone()

    if existing:
        conn.close()
        return False, "该设备在此时间段已有测试预约，请检查占用看板选择其他时间"

    conn.execute(
        """INSERT INTO borrow_records (equipment_id, user_id, borrow_date, expected_return_date,
           purpose, notes, test_standard, test_start_time, test_end_time, brand, sku, product_name)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (equipment_id, user_id, borrow_date, expected_return_date,
         purpose, notes, test_standard, test_start_time, test_end_time,
         brand, sku, product_name)
    )
    # 使用登记不改变设备物理状态，设备仍在实验室
    conn.commit()
    conn.close()
    return True, "使用登记成功"


def checkout_equipment(equipment_id, user_id, borrow_date, expected_return_date,
                       purpose='', notes=''):
    """借用设备出库（物理借出实验室）"""
    conn = get_connection()
    eq = conn.execute("SELECT status FROM equipment WHERE id=?", (equipment_id,)).fetchone()
    if not eq:
        conn.close()
        return False, "设备不存在"
    if eq['status'] != '可用':
        conn.close()
        return False, f"设备当前状态为「{eq['status']}」，无法借用出库"

    conn.execute(
        """INSERT INTO borrow_records (equipment_id, user_id, borrow_date, expected_return_date,
           purpose, notes, status)
           VALUES (?, ?, ?, ?, ?, ?, '已出库')""",
        (equipment_id, user_id, borrow_date, expected_return_date, purpose, notes)
    )
    conn.execute(
        "UPDATE equipment SET status='借出', updated_at=datetime('now','localtime') WHERE id=?",
        (equipment_id,)
    )
    conn.commit()
    conn.close()
    return True, "设备借用出库成功"


def return_equipment(record_id, return_date):
    """归还设备"""
    conn = get_connection()
    record = conn.execute("SELECT * FROM borrow_records WHERE id=?", (record_id,)).fetchone()
    if not record:
        conn.close()
        return False, "借用记录不存在"
    if record['status'] == '已归还':
        conn.close()
        return False, "该设备已归还"

    conn.execute(
        "UPDATE borrow_records SET actual_return_date=?, status='已归还' WHERE id=?",
        (return_date, record_id)
    )
    conn.execute(
        "UPDATE equipment SET status='可用', updated_at=datetime('now','localtime') WHERE id=?",
        (record['equipment_id'],)
    )
    conn.commit()
    conn.close()
    return True, "归还成功"


def get_active_borrows():
    """获取当前使用中的记录（实验室内部测试预约）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT b.*, e.name as equipment_name, e.serial_number, u.name as user_name, u.department
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        WHERE b.status = '借出中'
        ORDER BY b.borrow_date DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_equipment_schedule(equipment_id=None):
    """获取设备预约排期（使用登记），用于冲突检测"""
    conn = get_connection()
    if equipment_id:
        rows = conn.execute("""
            SELECT b.*, e.name as equipment_name, e.serial_number,
                   u.name as user_name, u.department
            FROM borrow_records b
            JOIN equipment e ON b.equipment_id = e.id
            JOIN users u ON b.user_id = u.id
            WHERE b.equipment_id = ? AND b.status = '借出中'
            ORDER BY b.borrow_date
        """, (equipment_id,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT b.*, e.name as equipment_name, e.serial_number,
                   u.name as user_name, u.department
            FROM borrow_records b
            JOIN equipment e ON b.equipment_id = e.id
            JOIN users u ON b.user_id = u.id
            WHERE b.status = '借出中'
            ORDER BY e.name, b.borrow_date
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def return_equipment(record_id, return_date):
    """归还设备"""
    conn = get_connection()
    record = conn.execute("SELECT * FROM borrow_records WHERE id=?", (record_id,)).fetchone()
    if not record:
        conn.close()
        return False, "借用记录不存在"
    if record['status'] == '已归还':
        conn.close()
        return False, "该设备已归还"

    conn.execute(
        "UPDATE borrow_records SET actual_return_date=?, status='已归还' WHERE id=?",
        (return_date, record_id)
    )
    conn.execute(
        "UPDATE equipment SET status='可用', updated_at=datetime('now','localtime') WHERE id=?",
        (record['equipment_id'],)
    )
    conn.commit()
    conn.close()
    return True, "归还成功"


def get_borrow_records(status=None, equipment_id=None, user_id=None, page=1, per_page=20):
    """查询借用记录"""
    conn = get_connection()
    conditions = []
    params = []

    if status:
        conditions.append("b.status = ?")
        params.append(status)
    if equipment_id:
        conditions.append("b.equipment_id = ?")
        params.append(equipment_id)
    if user_id:
        conditions.append("b.user_id = ?")
        params.append(user_id)

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    total = conn.execute(f"SELECT COUNT(*) FROM borrow_records b {where}", params).fetchone()[0]
    offset = (page - 1) * per_page

    sql = f"""
        SELECT b.*, e.name as equipment_name, e.serial_number, u.name as user_name, u.department
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        {where}
        ORDER BY b.id DESC
        LIMIT ? OFFSET ?
    """
    rows = conn.execute(sql, params + [per_page, offset]).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_active_borrows():
    """获取当前借出中的记录"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT b.*, e.name as equipment_name, e.serial_number, u.name as user_name, u.department
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        WHERE b.status = '借出中'
        ORDER BY b.borrow_date DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_equipment_schedule(equipment_id=None):
    """获取设备预约排期（当前和未来的借用），用于冲突检测"""
    conn = get_connection()
    if equipment_id:
        rows = conn.execute("""
            SELECT b.*, e.name as equipment_name, e.serial_number,
                   u.name as user_name, u.department
            FROM borrow_records b
            JOIN equipment e ON b.equipment_id = e.id
            JOIN users u ON b.user_id = u.id
            WHERE b.equipment_id = ? AND b.status = '借出中'
            ORDER BY b.borrow_date
        """, (equipment_id,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT b.*, e.name as equipment_name, e.serial_number,
                   u.name as user_name, u.department
            FROM borrow_records b
            JOIN equipment e ON b.equipment_id = e.id
            JOIN users u ON b.user_id = u.id
            WHERE b.status = '借出中'
            ORDER BY e.name, b.borrow_date
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 维护操作 ====================

def add_maintenance(data):
    """添加维护记录"""
    conn = get_connection()
    conn.execute("""
        INSERT INTO maintenance_records
        (equipment_id, maintenance_date, maintenance_type, description, cost,
         technician, next_maintenance_date, status, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data['equipment_id'], data['maintenance_date'], data.get('maintenance_type', '定期保养'),
        data.get('description', ''), data.get('cost', 0), data.get('technician', ''),
        data.get('next_maintenance_date', ''), data.get('status', '已完成'), data.get('notes', '')
    ))
    if data.get('status') == '进行中':
        conn.execute(
            "UPDATE equipment SET status='维修中', updated_at=datetime('now','localtime') WHERE id=?",
            (data['equipment_id'],)
        )
    conn.commit()
    conn.close()
    return True, "维护记录添加成功"


def update_maintenance(m_id, data):
    """更新维护记录"""
    conn = get_connection()
    conn.execute("""
        UPDATE maintenance_records SET equipment_id=?, maintenance_date=?, maintenance_type=?,
        description=?, cost=?, technician=?, next_maintenance_date=?, status=?, notes=?
        WHERE id=?
    """, (
        data['equipment_id'], data['maintenance_date'], data.get('maintenance_type', '定期保养'),
        data.get('description', ''), data.get('cost', 0), data.get('technician', ''),
        data.get('next_maintenance_date', ''), data.get('status', '已完成'), data.get('notes', ''), m_id
    ))
    if data.get('status') == '已完成':
        conn.execute(
            "UPDATE equipment SET status='可用', updated_at=datetime('now','localtime') WHERE id=?",
            (data['equipment_id'],)
        )
    conn.commit()
    conn.close()
    return True, "维护记录更新成功"


def delete_maintenance(m_id):
    """删除维护记录"""
    conn = get_connection()
    conn.execute("DELETE FROM maintenance_records WHERE id=?", (m_id,))
    conn.commit()
    conn.close()
    return True, "维护记录删除成功"


def get_maintenance_records(equipment_id=None, page=1, per_page=20):
    """查询维护记录"""
    conn = get_connection()
    if equipment_id:
        total = conn.execute(
            "SELECT COUNT(*) FROM maintenance_records WHERE equipment_id=?", (equipment_id,)
        ).fetchone()[0]
        offset = (page - 1) * per_page
        rows = conn.execute("""
            SELECT m.*, e.name as equipment_name, e.serial_number
            FROM maintenance_records m
            JOIN equipment e ON m.equipment_id = e.id
            WHERE m.equipment_id = ?
            ORDER BY m.maintenance_date DESC
            LIMIT ? OFFSET ?
        """, (equipment_id, per_page, offset)).fetchall()
    else:
        total = conn.execute("SELECT COUNT(*) FROM maintenance_records").fetchone()[0]
        offset = (page - 1) * per_page
        rows = conn.execute("""
            SELECT m.*, e.name as equipment_name, e.serial_number
            FROM maintenance_records m
            JOIN equipment e ON m.equipment_id = e.id
            ORDER BY m.maintenance_date DESC
            LIMIT ? OFFSET ?
        """, (per_page, offset)).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_upcoming_maintenance(days=30):
    """获取未来 N 天内计划的维护"""
    conn = get_connection()
    today = date.today()
    end_date = today + timedelta(days=days)
    rows = conn.execute("""
        SELECT m.*, e.name as equipment_name, e.serial_number
        FROM maintenance_records m
        JOIN equipment e ON m.equipment_id = e.id
        WHERE m.next_maintenance_date BETWEEN ? AND ?
        ORDER BY m.next_maintenance_date
    """, (str(today), str(end_date))).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 统计操作 ====================

def get_dashboard_stats():
    """获取看板统计数据"""
    conn = get_connection()
    stats = {
        'total': conn.execute("SELECT COUNT(*) FROM equipment").fetchone()[0],
        'available': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='可用'").fetchone()[0],
        'in_use': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='借出中'").fetchone()[0],
        'borrowed': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='借出'").fetchone()[0],
        'maintenance': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='维修中'").fetchone()[0],
        'scrapped': conn.execute("SELECT COUNT(*) FROM equipment WHERE status='报废'").fetchone()[0],
        'total_users': conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        'active_borrows': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='借出中'").fetchone()[0],
        'total_value': conn.execute("SELECT COALESCE(SUM(price), 0) FROM equipment").fetchone()[0],
    }
    conn.close()
    return stats


def get_status_distribution():
    """设备状态分布"""
    conn = get_connection()
    rows = conn.execute("SELECT status, COUNT(*) as count FROM equipment GROUP BY status").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_category_distribution():
    """分类设备统计"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT c.name, COUNT(e.id) as count, COALESCE(SUM(e.price), 0) as total_value
        FROM categories c
        LEFT JOIN equipment e ON c.id = e.category_id
        GROUP BY c.id
        ORDER BY count DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_borrow_stats():
    """借用统计"""
    conn = get_connection()
    stats = {
        'total': conn.execute("SELECT COUNT(*) FROM borrow_records").fetchone()[0],
        'active': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='借出中'").fetchone()[0],
        'returned': conn.execute("SELECT COUNT(*) FROM borrow_records WHERE status='已归还'").fetchone()[0],
    }
    rows = conn.execute("""
        SELECT strftime('%Y-%m', borrow_date) as month, COUNT(*) as count
        FROM borrow_records
        WHERE borrow_date >= date('now', '-12 months')
        GROUP BY month ORDER BY month
    """).fetchall()
    stats['monthly'] = [dict(r) for r in rows]
    conn.close()
    return stats


def get_maintenance_cost_stats():
    """维护费用统计"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT strftime('%Y-%m', maintenance_date) as month,
               COUNT(*) as count, COALESCE(SUM(cost), 0) as total_cost
        FROM maintenance_records
        WHERE maintenance_date >= date('now', '-12 months')
        GROUP BY month ORDER BY month
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_equipment():
    """获取所有设备（导出用）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT e.*, c.name as category_name
        FROM equipment e LEFT JOIN categories c ON e.category_id = c.id
        ORDER BY e.id
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_borrow_records_export():
    """获取所有借用记录（导出用）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT b.id, e.name as 设备名称, e.serial_number as 设备编号,
               u.name as 借用人, u.department as 部门,
               b.borrow_date as 借用日期, b.expected_return_date as 预计归还,
               b.actual_return_date as 实际归还, b.status as 状态,
               b.purpose as 用途, b.notes as 备注
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        ORDER BY b.id DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_maintenance_export():
    """获取所有维护记录（导出用）"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT m.id, e.name as 设备名称, e.serial_number as 设备编号,
               m.maintenance_date as 维护日期, m.maintenance_type as 维护类型,
               m.description as 描述, m.cost as 费用,
               m.technician as 技术人员, m.next_maintenance_date as 下次维护,
               m.status as 状态, m.notes as 备注
        FROM maintenance_records m
        JOIN equipment e ON m.equipment_id = e.id
        ORDER BY m.id DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_recent_borrows(limit=8):
    """获取最近借用记录"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT b.*, e.name as equipment_name, e.serial_number, u.name as user_name
        FROM borrow_records b
        JOIN equipment e ON b.equipment_id = e.id
        JOIN users u ON b.user_id = u.id
        ORDER BY b.id DESC
        LIMIT ?
    """, (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 测试标准库 ====================

# 设备关键词 → 关联测试标准列表
EQUIPMENT_STANDARD_MAP = {
    '按键寿命': ['IEC61058-1 (器具开关耐久性)'],
    '插拔力': ['EIA-364-09 (电子连接器插拔力)', 'USB-IF 协会规范'],
    '插拔寿命': ['EIA-364-09 (电子连接器拔插寿命)', 'USB-IF 协会规范'],
    '跌落': ['ISTA 6A (Amazon SIOC 原箱发货认证)'],
    '单翼跌落': ['ISTA 6A (Amazon SIOC 原箱发货认证)'],
    '单臂跌落': ['ISTA 6A (Amazon SIOC 原箱发货认证)'],
    '振动': ['ASTM D4728 (随机振动)', 'GB/T 4857.7 (正弦振动)'],
    '摇摆': ['IEC60335-1 §25.14 (电源线弯折)'],
    '拉力': ['ISO527 (塑料拉伸性能)', 'ASTM D3359 (附着力测试)'],
    '纽扣': ['ASTM D 相关 (纽扣拉力)'],
    '硬度': ['ISO 6508 / ASTM E18 (洛氏硬度)'],
    '盐雾': ['ISO9227 (盐雾试验)', 'ASTM B117 (盐雾试验)', 'GB/T10125 (盐雾试验)'],
    '淋雨': ['IEC60529 (IP代码)', 'GB/T4208 (IP代码)'],
    '防水': ['IEC60529 (IP代码)', 'GB/T4208 (IP代码)'],
    '温湿度': ['IEC60068-2-78 (恒定湿热)', 'IEC60068-2-30 (交变湿热)'],
    '热冲击': ['IEC60068-2-14 (温度变化)', 'GB/T2423.22 (温度变化)'],
    '冷热冲击': ['IEC60068-2-14 (温度变化)', 'GB/T2423.22 (温度变化)'],
    'ESD': ['IEC61000-4-2 (静电放电抗扰度)'],
    '静电': ['IEC61000-4-2 (静电放电抗扰度)'],
    '安规': ['IEC60335-1 (家用电器安全)', 'IEC62368-1 (音视频/IT安全)'],
    '安全性能': ['IEC60335-1 (家用电器安全)', 'IEC62368-1 (音视频/IT安全)'],
    '电池综合': ['UN38.3 (锂电池运输安全)', 'IEC62133 (锂电池安全)', 'UL2054 (北美电池安全)'],
    '模拟电池': ['IEC62368-1 (故障条件测试)'],
    '大功率': ['UL1995 (冷暖设备)', 'IEC60335-2-40 (热泵/空调)'],
    '直流电源': ['IEC62368-1 (电气安全)'],
    'PD负载': ['USB-IF PD 3.0/3.1 规范', 'EN62368-1'],
    '负载测试': ['IEC62368-1 (电气安全)'],
    '阻抗': ['IEC 相关 (阻抗测试)'],
    '直流稳压': ['IEC62368-1 (电气安全)'],
    '可调节': ['IEC62368-1 (电气安全)'],
    '二次元': ['ISO10360-7 (影像测量仪)'],
    '发热模组': ['UL130 (电热服装安全)', 'IEC60335-2-17 (电热毯/服装)'],
    '温度巡检': ['IEC60068-2-78 (恒定湿热)'],
}

# 全部标准列表（用于下拉选择）
ALL_STANDARDS = sorted(set(
    std for standards in EQUIPMENT_STANDARD_MAP.values() for std in standards
))


def find_standards_for_equipment(equipment_name):
    """根据设备名称查找关联的测试标准列表"""
    if not equipment_name:
        return []
    for keyword, standards in EQUIPMENT_STANDARD_MAP.items():
        if keyword in equipment_name:
            return standards
    return []


def get_all_test_standards():
    """获取所有测试标准列表（供下拉选择）"""
    return ALL_STANDARDS


# ==================== 品质人员导入 ====================

def import_personnel_from_excel(filepath=None):
    """从原始名单Excel导入品质人员"""
    import pandas as pd
    if filepath is None:
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                'SainStore实验室文件', '原始名单.xlsx')
    if not os.path.exists(filepath):
        return 0, "文件不存在"

    df = pd.read_excel(filepath, engine='openpyxl')
    conn = get_connection()
    count = 0
    for _, row in df.iterrows():
        name = str(row.get('品质人员', '')).strip() if pd.notna(row.get('品质人员')) else ''
        if not name or name == 'nan':
            continue
        bg = str(row.get('BG', '')).strip() if pd.notna(row.get('BG')) else ''
        bu = str(row.get('BU', '')).strip() if pd.notna(row.get('BU')) else ''
        brand = str(row.get('brand', '')).strip() if pd.notna(row.get('brand')) else ''
        dept = f"{bg}/{bu}" if bg and bu else (bg or bu or '品质部')

        existing = conn.execute("SELECT id FROM users WHERE name=?", (name,)).fetchone()
        if existing:
            conn.execute("UPDATE users SET department=?, phone=?, email=? WHERE id=?",
                         (dept, bg, bu, existing[0]))
        else:
            conn.execute("INSERT INTO users (name, department, phone, email, role) VALUES (?, ?, ?, ?, ?)",
                         (name, dept, bg, bu, '管理员'))
        count += 1
    conn.commit()
    conn.close()
    return count, "导入成功"


# ==================== 样品管理 ====================

def import_samples_from_excel(filepath=None):
    """从签样记录Excel导入样品（遍历所有BG工作表）"""
    import openpyxl
    if filepath is None:
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                'SainStore实验室文件', '签样记录 （最新）.xlsx')
    if not os.path.exists(filepath):
        return 0, "文件不存在"

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        conn = get_connection()
        count = 0
        skip = {'不是封样可参考', '原IBG'}

        # 清空旧数据重新导入
        conn.execute("DELETE FROM samples")

        for sheet_name in wb.sheetnames:
            if sheet_name in skip:
                continue
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                bg_val = str(ws.cell(r, 2).value or '').strip()
                sku_val = str(ws.cell(r, 3).value or '').strip()
                name_val = str(ws.cell(r, 4).value or '').strip()
                if not name_val or name_val.lower() == 'none':
                    continue
                sign_date = str(ws.cell(r, 5).value or '').strip()
                supplier = str(ws.cell(r, 6).value or '').strip()
                notes = str(ws.cell(r, 7).value or '').strip()
                location = str(ws.cell(r, 8).value or '').strip()

                conn.execute(
                    """INSERT INTO samples (bg, sku, sample_name, sign_date, supplier, notes, location)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (bg_val, sku_val, name_val, sign_date, supplier, notes, location)
                )
                count += 1
        wb.close()
        conn.commit()
        conn.close()
        return count, f"成功导入 {count} 条样品"
    except Exception as e:
        return 0, str(e)


def get_samples(search='', bg='', page=1, per_page=20):
    """查询样品列表"""
    conn = get_connection()
    conditions = []
    params = []
    if search:
        conditions.append("(sample_name LIKE ? OR sku LIKE ?)")
        params.extend([f'%{search}%'] * 2)
    if bg:
        conditions.append("bg = ?")
        params.append(bg)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    total = conn.execute(f"SELECT COUNT(*) FROM samples {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM samples {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def add_sample(data):
    """添加样品（自动计算一年有效期）"""
    from datetime import datetime as dt, timedelta as td
    sign_date = data.get('sign_date', '')
    expiry_date = ''
    if sign_date:
        try:
            s = sign_date.replace('.', '-')
            if len(s) == 7: s += '-01'
            d = dt.strptime(s[:10], '%Y-%m-%d')
            expiry_date = (d + td(days=365)).strftime('%Y-%m-%d')
        except Exception:
            expiry_date = ''
    conn = get_connection()
    conn.execute("""
        INSERT INTO samples (bg, sku, sample_name, sign_date, supplier, brand, notes, location, expiry_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data.get('bg', ''), data.get('sku', ''), data.get('sample_name', ''),
          sign_date, data.get('supplier', ''),
          data.get('brand', ''),
          data.get('notes', ''), data.get('location', ''), expiry_date))
    conn.commit()
    conn.close()
    return True, "样品添加成功"


def update_sample(s_id, data):
    """更新样品"""
    conn = get_connection()
    conn.execute("""
        UPDATE samples SET bg=?, sku=?, sample_name=?, sign_date=?, supplier=?,
        brand=?, notes=?, location=?
        WHERE id=?
    """, (data.get('bg', ''), data.get('sku', ''), data.get('sample_name', ''),
          data.get('sign_date', ''), data.get('supplier', ''), data.get('brand', ''),
          data.get('notes', ''), data.get('location', ''), s_id))
    conn.commit()
    conn.close()
    return True, "样品已更新"


def delete_sample(s_id):
    """删除样品"""
    conn = get_connection()
    conn.execute("DELETE FROM samples WHERE id=?", (s_id,))
    conn.commit()
    conn.close()
    return True, "样品已删除"


# ==================== 样品出库 ====================

def sample_outbound(sample_id, qty=1, out_date='', borrower='', department='', reason='', notes=''):
    """样品出库登记"""
    conn = get_connection()
    # 检查是否已出库
    current = conn.execute("SELECT out_status FROM samples WHERE id=?", (sample_id,)).fetchone()
    if current and current['out_status'] == '已出库':
        conn.close()
        return False, "该样品已出库，请先归还再操作"
    # 写入出库记录
    conn.execute("""
        INSERT INTO sample_outbound (sample_id, qty, out_date, borrower, department, reason, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (sample_id, qty, out_date, borrower, department, reason, notes))
    # 更新样品状态
    conn.execute("UPDATE samples SET out_status='已出库' WHERE id=?", (sample_id,))
    conn.commit()
    conn.close()
    return True, "出库成功"


def sample_return(sample_id):
    """样品归还（改回在库状态）"""
    conn = get_connection()
    conn.execute("UPDATE samples SET out_status='在库' WHERE id=?", (sample_id,))
    conn.commit()
    conn.close()
    return True, "已归还"


def get_outbound_records(sample_id=None, limit=100):
    """查询出库记录"""
    conn = get_connection()
    if sample_id:
        rows = conn.execute("""
            SELECT o.*, s.sample_name, s.sku, s.bg
            FROM sample_outbound o JOIN samples s ON o.sample_id = s.id
            WHERE o.sample_id = ? ORDER BY o.id DESC LIMIT ?
        """, (sample_id, limit)).fetchall()
    else:
        rows = conn.execute("""
            SELECT o.*, s.sample_name, s.sku, s.bg
            FROM sample_outbound o JOIN samples s ON o.sample_id = s.id
            ORDER BY o.id DESC LIMIT ?
        """, (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 变更管理 ====================

def import_changes_from_excel(filepath=None):
    """从产品变更汇总表Excel导入变更记录"""
    import pandas as pd
    if filepath is None:
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                'SainStore实验室文件', '产品变更汇总表 .xlsx')
    if not os.path.exists(filepath):
        return 0, "文件不存在"

    df = pd.read_excel(filepath, engine='openpyxl', header=1)
    conn = get_connection()
    count = 0
    for _, row in df.iterrows():
        bu = str(row.get('BU', '')).strip() if pd.notna(row.get('BU')) else ''
        brand = str(row.get('品牌', '')).strip() if pd.notna(row.get('品牌')) else ''
        if not brand or brand == 'nan':
            continue
        conn.execute("""
            INSERT INTO change_records (bu, brand, sku, change_reason, attachments, change_date, confirm_date, confirm_person)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            bu, brand,
            str(row.get('SKU', '')).strip() if pd.notna(row.get('SKU')) else '',
            str(row.get('变更原因及内容', '')).strip()[:500] if pd.notna(row.get('变更原因及内容')) else '',
            '',  # attachments
            str(row.get('变更时间', '')).strip()[:19] if pd.notna(row.get('变更时间')) else '',
            str(row.get('确认时间', '')).strip()[:19] if pd.notna(row.get('确认时间')) else '',
            str(row.get('确认人', '')).strip() if pd.notna(row.get('确认人')) else '',
        ))
        count += 1
    conn.commit()
    conn.close()
    return count, "导入成功"


def get_changes(search='', bu='', brand='', page=1, per_page=20, search_sku='', search_content='', search_supplier=''):
    """查询变更记录（支持多字段搜索）"""
    conn = get_connection()
    conditions = []
    params = []

    # 拼接所有搜索条件
    search_parts = []
    if search:
        search_parts.append(search)
    if search_sku:
        conditions.append("sku LIKE ?")
        params.append(f'%{search_sku}%')
    if search_content:
        conditions.append("change_reason LIKE ?")
        params.append(f'%{search_content}%')
    if search_supplier:
        conditions.append("supplier LIKE ?")
        params.append(f'%{search_supplier}%')
    if search:
        conditions.append("(brand LIKE ? OR change_reason LIKE ? OR supplier LIKE ?)")
        params.extend([f'%{search}%'] * 3)

    if bu:
        conditions.append("bu = ?")
        params.append(bu)
    if brand:
        conditions.append("brand = ?")
        params.append(brand)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    total = conn.execute(f"SELECT COUNT(*) FROM change_records {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM change_records {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def add_change(data):
    """添加变更记录"""
    conn = get_connection()
    conn.execute("""
        INSERT INTO change_records (bu, brand, sku, change_reason, supplier, attachments, change_date, confirm_date, confirm_person, rd_team)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data.get('bu', ''), data.get('brand', ''), data.get('sku', ''),
          data.get('change_reason', ''), data.get('supplier', ''), data.get('attachments', ''),
          data.get('change_date', ''), data.get('confirm_date', ''),
          data.get('confirm_person', '') or '', data.get('rd_team', '')))
    conn.commit()
    conn.close()
    return True, "变更记录添加成功"


def update_change(c_id, data):
    """更新变更记录"""
    conn = get_connection()
    conn.execute("""
        UPDATE change_records SET bu=?, brand=?, sku=?, change_reason=?, supplier=?,
        attachments=?, change_date=?, rd_team=?
        WHERE id=?
    """, (data.get('bu', ''), data.get('brand', ''), data.get('sku', ''),
          data.get('change_reason', ''), data.get('supplier', ''), data.get('attachments', ''),
          data.get('change_date', ''), data.get('rd_team', ''), c_id))
    conn.commit()
    conn.close()
    return True, "变更记录已更新"


def delete_change(c_id):
    """删除变更记录"""
    conn = get_connection()
    conn.execute("DELETE FROM change_records WHERE id=?", (c_id,))
    conn.commit()
    conn.close()
    return True, "变更记录已删除"


# ==================== 检验报告 ====================

def add_inspection_report(data):
    """添加检验报告"""
    conn = get_connection()
    conn.execute("""
        INSERT INTO inspection_reports (report_type, inspector, product_name, bg, bu, brand, sku,
                                        filename, file_path, image_paths, supplier, status, reviewer)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data.get('report_type', ''), data.get('inspector', ''),
          data.get('product_name', ''), data.get('bg', ''), data.get('bu', ''),
          data.get('brand', ''), data.get('sku', ''), data.get('filename', ''),
          data.get('file_path', ''), data.get('image_paths', ''),
          data.get('supplier', ''),
          data.get('status', '待审核'), data.get('reviewer', 'teddy.li黎晓锋')))
    conn.commit()
    conn.close()
    return True, "报告提交成功"


def get_inspection_reports(status=None, page=1, per_page=20):
    """查询检验报告"""
    conn = get_connection()
    conditions = []
    params = []
    if status:
        conditions.append("status = ?")
        params.append(status)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    total = conn.execute(f"SELECT COUNT(*) FROM inspection_reports {where}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT * FROM inspection_reports {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows], total


def get_report_daily_stats():
    """每日报告统计"""
    conn = get_connection()
    today = date.today()
    total = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE date(created_at)=?",
        (str(today),)
    ).fetchone()[0]
    pending = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE date(created_at)=? AND status='待审核'",
        (str(today),)
    ).fetchone()[0]
    approved = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE date(created_at)=? AND status='已通过'",
        (str(today),)
    ).fetchone()[0]
    conn.close()
    return {'total': total, 'pending': pending, 'approved': approved}


def _get_names_data():
    """获取原始名单数据（优先Excel，fallback到硬编码）"""
    if os.path.exists(NAMES_FILE):
        try:
            return _read_excel_all()
        except Exception:
            pass
    return _names_fallback()


def _read_excel_all():
    """从Excel读取所有列"""
    import openpyxl
    wb = openpyxl.load_workbook(NAMES_FILE, read_only=True, data_only=True)
    ws = wb['Sheet1']
    cols = {1: [], 2: [], 3: [], 4: [], 5: []}
    for r in range(2, ws.max_row + 1):
        for c in cols:
            val = ws.cell(r, c).value
            if val:
                v = str(val).strip()
                if v and v.lower() not in ('no brand', ''):
                    cols[c].append(v)
    wb.close()
    return {k: list(dict.fromkeys(v)) for k, v in cols.items()}


def _names_fallback():
    """硬编码的名单数据（当Excel文件不可用时使用）"""
    return {
        1: ['ACE', 'BOC', 'Langis LLC', 'Root', 'Z_Archived'],
        2: ['ACE', 'Aura', 'BigRock', 'Epicarry', 'KPL', 'Kronos', 'Orion', 'Parts', 'RaChat', 'Root', 'Root-Misc', 'TheUnicorn'],
        3: ['A11N Sports', 'Aireal', 'Airthereal', 'Baofeng', 'BELSIZE', 'BLIZZARD', 'BUYDEEM', 'CREALITY', 'darkFlash', 'DELI', 'Furlihong', 'Genmitsu', 'GLENCREAG', 'iPettie', 'Kronos-Misc', 'Lagute', 'LTC', 'OFFNOVA', 'OPENHEAT', 'Orion-Misc', 'ororo', 'Raddy', 'Radioddity', 'Razorri', 'Redragon', 'Root-Misc', 'Royal Kludge', 'SainSmart', 'SainSmart Jr.', 'SWONDER', 'TOSOT', 'TURBRO', 'WOODSTARTER', 'XIEGU'],
        4: ['Hibiscus', 'Ebony', 'Ace', 'Cactus', 'ET', 'ME', 'QA'],
        5: ['Carl Dong董献民', 'joung.yuan袁毅洪', 'haruna.wei韦梦婷', 'amelia.han韩亚南', 'teddy.li黎晓锋', 'colin.xu徐胜涛', 'lucy.ning宁小连', 'ken.huang黄海森', 'lainey.pan潘杨阳', 'fowler.zhai翟始福', 'leo.wu吴嘉俊', 'wenzel.chen陈文钊', 'bruce.cheng程强'],
    }


def get_bg_list():
    return _get_names_data()[1]

def get_bu_list():
    return _get_names_data()[2]

def get_brand_list():
    return _get_names_data()[3]

def get_rd_teams():
    return _get_names_data()[4]

def get_quality_users_list():
    return _get_names_data()[5]


# ==================== 活动日志 ====================

def log_activity(user_email, action, category='system', detail='', page=''):
    """记录用户活动日志"""
    import re
    conn = get_connection()
    user_name = user_email.split('@')[0] if '@' in user_email else user_email
    # 安全截断，避免超长
    detail = (detail or '')[:500]
    page = (page or '')[:200]
    conn.execute(
        """INSERT INTO activity_log (user_email, user_name, action, category, detail, page)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (user_email, user_name, action, category, detail, page)
    )
    conn.commit()
    conn.close()


def get_activity_logs(limit=200, category='', user_email='', hours=24):
    """获取活动日志"""
    conn = get_connection()
    conditions = ["created_at >= datetime('now', 'localtime', ?)"]
    params = [f'-{hours} hours']
    if category:
        conditions.append("category = ?")
        params.append(category)
    if user_email:
        conditions.append("user_email = ?")
        params.append(user_email)
    where = " AND ".join(conditions)
    sql = f"SELECT * FROM activity_log WHERE {where} ORDER BY id DESC LIMIT ?"
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_online_users(minutes=15):
    """获取最近在线的用户（X分钟内有过活动）"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT user_email, user_name, MAX(created_at) as last_active,
                  COUNT(*) as action_count
           FROM activity_log
           WHERE created_at >= datetime('now', 'localtime', ?)
           GROUP BY user_email, user_name
           ORDER BY last_active DESC""",
        (f'-{minutes} minutes',)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_login_history(limit=50):
    """获取登录历史"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT * FROM activity_log WHERE action = '登录成功'
           ORDER BY id DESC LIMIT ?""", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_daily_stats():
    """获取每日访问统计"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT date(created_at) as day,
               COUNT(DISTINCT user_email) as unique_users,
               COUNT(*) as total_actions,
               SUM(CASE WHEN action = '登录成功' THEN 1 ELSE 0 END) as logins,
               SUM(CASE WHEN category = 'page_view' THEN 1 ELSE 0 END) as page_views,
               SUM(CASE WHEN category = 'data_edit' THEN 1 ELSE 0 END) as data_edits
        FROM activity_log
        WHERE created_at >= date('now', '-7 days')
        GROUP BY day
        ORDER BY day DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_sample_bg_list():
    """获取样品表中实际存在的BG列表"""
    conn = get_connection()
    rows = conn.execute("SELECT DISTINCT bg FROM samples WHERE bg != '' ORDER BY bg").fetchall()
    conn.close()
    return [r['bg'] for r in rows]


def get_page_hotspots(limit=20):
    """获取热门页面访问统计"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT page, COUNT(*) as visit_count,
                  COUNT(DISTINCT user_email) as unique_users
           FROM activity_log WHERE category = 'page_view'
           AND created_at >= datetime('now', '-7 days')
           GROUP BY page ORDER BY visit_count DESC LIMIT ?""", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]
    """获取热门页面访问统计"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT page, COUNT(*) as visit_count,
                  COUNT(DISTINCT user_email) as unique_users
           FROM activity_log WHERE category = 'page_view'
           AND created_at >= datetime('now', '-7 days')
           GROUP BY page ORDER BY visit_count DESC LIMIT ?""", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ==================== 版本变动日志 ====================

def add_changelog(version, title, description='', changes='', category='优化', created_by=''):
    """添加版本变动记录"""
    conn = get_connection()
    conn.execute(
        """INSERT INTO changelog (version, title, description, changes, category, created_by)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (version, title, description, changes, category, created_by)
    )
    conn.commit()
    conn.close()


def get_changelogs(limit=50):
    """获取版本变动日志"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM changelog ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def delete_activity(activity_id):
    """删除活动日志"""
    conn = get_connection()
    conn.execute("DELETE FROM activity_log WHERE id = ?", (activity_id,))
    conn.commit()
    conn.close()
    return True, "已删除"


def delete_activities(activity_ids):
    """批量删除活动日志"""
    conn = get_connection()
    for aid in activity_ids:
        conn.execute("DELETE FROM activity_log WHERE id = ?", (aid,))
    conn.commit()
    conn.close()
    return True, f"已删除 {len(activity_ids)} 条记录"


# ==================== 首页看板聚合函数 ====================

def get_sample_dashboard_stats():
    """获取样品看板统计数据"""
    conn = get_connection()
    today = datetime.now().strftime("%Y-%m-%d")
    thirty_days = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")

    total = conn.execute("SELECT COUNT(*) FROM samples").fetchone()[0]
    in_stock = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE out_status != '已出库' OR out_status IS NULL"
    ).fetchone()[0]
    out_stock = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE out_status = '已出库'"
    ).fetchone()[0]
    near_expiry = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE expiry_date BETWEEN ? AND ? AND expiry_date != ''",
        (today, thirty_days)
    ).fetchone()[0]
    expired = conn.execute(
        "SELECT COUNT(*) FROM samples WHERE expiry_date < ? AND expiry_date != ''",
        (today,)
    ).fetchone()[0]
    conn.close()
    return {
        'total': total, 'in_stock': in_stock, 'out_stock': out_stock,
        'near_expiry': near_expiry, 'expired': expired
    }


def get_sample_bg_distribution():
    """获取样品 BG 分布"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT bg, COUNT(*) as count FROM samples WHERE bg != '' GROUP BY bg ORDER BY count DESC"
    ).fetchall()
    conn.close()
    return [{'bg': r[0], 'count': r[1]} for r in rows]


def get_change_dashboard_stats():
    """获取变更看板统计数据"""
    conn = get_connection()
    this_month = datetime.now().strftime("%Y-%m")
    total = conn.execute("SELECT COUNT(*) FROM change_records").fetchone()[0]
    monthly = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE change_date LIKE ?",
        (f"{this_month}%",)
    ).fetchone()[0]
    recent = conn.execute(
        "SELECT COUNT(*) FROM change_records WHERE created_at >= datetime('now', '-7 days')"
    ).fetchone()[0]
    conn.close()
    return {'total': total, 'this_month': monthly, 'recent_7d': recent}


def get_change_bu_distribution():
    """获取变更 BU 分布"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT bu, COUNT(*) as count FROM change_records WHERE bu != '' GROUP BY bu ORDER BY count DESC"
    ).fetchall()
    conn.close()
    return [{'bu': r[0], 'count': r[1]} for r in rows]


def get_inspection_dashboard_stats():
    """获取检验报告看板统计"""
    conn = get_connection()
    total = conn.execute("SELECT COUNT(*) FROM inspection_reports").fetchone()[0]
    pending = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE status = '待审核'"
    ).fetchone()[0]
    approved = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE status = '已通过'"
    ).fetchone()[0]
    rejected = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE status = '已驳回'"
    ).fetchone()[0]
    this_month = datetime.now().strftime("%Y-%m")
    monthly = conn.execute(
        "SELECT COUNT(*) FROM inspection_reports WHERE created_at LIKE ?",
        (f"{this_month}%",)
    ).fetchone()[0]
    conn.close()
    return {
        'total': total, 'pending': pending, 'approved': approved,
        'rejected': rejected, 'this_month': monthly
    }


def get_report_type_distribution():
    """获取报告类型分布"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT report_type, COUNT(*) as count FROM inspection_reports WHERE report_type != '' GROUP BY report_type ORDER BY count DESC"
    ).fetchall()
    conn.close()
    return [{'type': r[0], 'count': r[1]} for r in rows]


def get_recent_reports(limit=6):
    """获取最近检验报告"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT id, report_type, inspector, product_name, bg, bu, brand, status, created_at
           FROM inspection_reports ORDER BY id DESC LIMIT ?""",
        (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_recent_changes(limit=6):
    """获取最近变更记录"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT id, bu, brand, sku, change_reason, supplier, change_date, created_at
           FROM change_records ORDER BY id DESC LIMIT ?""",
        (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_expiring_samples(days=30):
    """获取即将过期的样品"""
    conn = get_connection()
    today = datetime.now().strftime("%Y-%m-%d")
    deadline = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    rows = conn.execute(
        """SELECT id, bg, sku, sample_name, expiry_date, sign_date, supplier, out_status
           FROM samples
           WHERE expiry_date BETWEEN ? AND ? AND expiry_date != ''
           ORDER BY expiry_date ASC""",
        (today, deadline)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_equipment_usage_trend():
    """获取设备使用趋势（最近7天）"""
    conn = get_connection()
    rows = conn.execute(
        """SELECT borrow_date, COUNT(*) as count
           FROM borrow_records
           WHERE borrow_date >= date('now', '-7 days')
           GROUP BY borrow_date ORDER BY borrow_date"""
    ).fetchall()
    conn.close()
    return [{'date': r[0], 'count': r[1]} for r in rows]
